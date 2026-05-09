# reports/excel.py
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.cogen_result import CoGenResultado

# Colores
_AZUL_HEADER = "1F4E79"
_GRIS_TOTALES = "D9E1F2"
_VERDE_EBITDA = "E2EFDA"
_GRIS_PARAMS  = "EBF0F7"

# Filas de estructura en "Análisis Mensual"
# Filas 1-6: bloque de parámetros; fila 7: separador; fila 8: encabezados; fila 9+: datos
_FILA_HEADER = 8
_FILA_DATOS  = 9


def generar_excel(resultado: CoGenResultado, output_path: Path) -> Path:
    """Genera reporte Excel con análisis mensual de cogeneración.

    Crea dos hojas:
    - 'Análisis Mensual': bloque de parámetros + tabla con meses + totales con fórmulas
    - 'Parámetros': resumen de parámetros técnicos usados

    Args:
        resultado: CoGenResultado con todos los meses calculados
        output_path: ruta donde guardar el .xlsx

    Returns:
        output_path (para encadenamiento)
    """
    output_path = Path(output_path)
    wb = openpyxl.Workbook()

    _escribir_hoja_analisis(wb, resultado)
    _escribir_hoja_parametros(wb, resultado.params)

    # Eliminar hoja vacía por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    return output_path


# ── Hoja 1: Análisis Mensual ──────────────────────────────────────────────────

_COLUMNAS = [
    ("Periodo",               "periodo_inicio",             "fecha"),    # A
    ("kWh Total",             "kwh_total",                  "numero"),   # B — valor fijo
    ("Costo CFE (MXN)",       "costo_cfe_mxn",              "moneda"),   # C — valor fijo
    ("$/kWh Promedio",        "costo_promedio_kwh",         "decimal4"), # D — valor fijo
    ("GJ Gas Real",           "gj_consumido",               "numero"),   # E — valor fijo
    ("$/GJ Gas",              "costo_unitario_gj",          "decimal4"), # F — valor fijo
    ("Costo Gas Real (MXN)",  "costo_gas_actual_mxn",       "moneda"),   # G — valor fijo
    ("kWh Cubiertos",         "kwh_cubiertos",              "numero"),   # H — fórmula
    ("GJ Cogen",              "gj_gas_cogen",               "numero"),   # I — fórmula
    ("Costo Gas Cogen (MXN)", "costo_gas_cogen_mxn",        "moneda"),   # J — fórmula
    ("Ahorro Elec. (MXN)",    "ahorro_electricidad_mxn",    "moneda"),   # K — fórmula
    ("Calor Recup. (GJ)",     "calor_recuperado_gj",        "numero"),   # L — fórmula
    ("Ahorro Caldera (MXN)",  "ahorro_caldera_mxn",         "moneda"),   # M — fórmula
    ("Ahorro Neto Mes (MXN)", "ebitda_mes_mxn",             "moneda"),   # N — fórmula
    ("Prorrateo",             "nota_prorrateo",             "texto"),    # O — valor fijo
]

# Fórmulas para columnas calculadas; {R} se reemplaza por el número de fila real.
# Referencias a parámetros del bloque superior (absolutas):
#   $B$2 = cobertura_electrica
#   $B$3 = rendimiento_electrico
#   $B$4 = rendimiento_termico
#   $B$5 = eficiencia_caldera
#   $B$6 = factor kWh→GJ (0.0036)
_FORMULAS: dict[str, str] = {
    "kwh_cubiertos":           "=B{R}*$B$2",
    "gj_gas_cogen":            "=H{R}*$B$6*1.11/$B$3",
    "costo_gas_cogen_mxn":     "=I{R}*F{R}",
    "ahorro_electricidad_mxn": "=H{R}*D{R}",
    "calor_recuperado_gj":     "=I{R}*$B$4",
    "ahorro_caldera_mxn":      "=(L{R}/$B$5)*F{R}",
    "ebitda_mes_mxn":          "=K{R}+M{R}-J{R}-H{R}*D{R}*0.3",
}

# Columnas que tienen fila de totales (=SUM).
_COLS_SUMA = {
    "kwh_total",
    "kwh_cubiertos",
    "gj_gas_cogen",
    "costo_gas_cogen_mxn",
    "ahorro_electricidad_mxn",
    "calor_recuperado_gj",
    "ahorro_caldera_mxn",
    "ebitda_mes_mxn",
}


def _escribir_bloque_params(ws, params) -> None:
    """Escribe el bloque de parámetros en filas 1-6 de la hoja."""
    # Fila 1: encabezado de sección
    ws.merge_cells("A1:B1")
    encabezado = ws["A1"]
    encabezado.value = "Parámetros del motor candidato"
    encabezado.font = Font(bold=True, color="FFFFFF")
    encabezado.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
    encabezado.alignment = Alignment(horizontal="center", vertical="center")

    # Filas 2-6: un parámetro por fila
    params_rows = [
        ("Cobertura objetivo",    float(params.cobertura_electrica),  "0%"),
        ("Rendimiento eléctrico", float(params.rendimiento_electrico), "0%"),
        ("Rendimiento térmico",   float(params.rendimiento_termico),   "0%"),
        ("Eficiencia caldera",    float(params.eficiencia_caldera),    "0%"),
        ("Factor kWh→GJ",        0.0036,                              "0.0000"),
    ]
    label_font = Font(bold=True)
    params_fill = PatternFill("solid", fgColor=_GRIS_PARAMS)
    for fila, (label, valor, fmt) in enumerate(params_rows, 2):
        celda_label = ws.cell(fila, 1, label)
        celda_label.font = label_font
        celda_label.fill = params_fill

        celda_val = ws.cell(fila, 2, valor)
        celda_val.number_format = fmt
        celda_val.fill = params_fill

    # Fila 7 queda vacía (separador visual)


def _escribir_hoja_analisis(wb: openpyxl.Workbook, resultado: CoGenResultado) -> None:
    ws = wb.create_sheet("Análisis Mensual")

    _escribir_bloque_params(ws, resultado.params)

    # Encabezados en fila 8
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_AZUL_HEADER)
    for col_idx, (titulo, _, _fmt) in enumerate(_COLUMNAS, 1):
        cell = ws.cell(_FILA_HEADER, col_idx, titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Filas de datos desde fila 9
    for row_offset, mes in enumerate(resultado.meses):
        R = _FILA_DATOS + row_offset
        for col_idx, (_titulo, attr, fmt) in enumerate(_COLUMNAS, 1):
            if attr in _FORMULAS:
                valor = _FORMULAS[attr].format(R=R)
            else:
                valor = _formatear(getattr(mes, attr), fmt)
            cell = ws.cell(R, col_idx, valor)
            if fmt == "moneda":
                cell.number_format = '#,##0.00'
            elif fmt == "numero":
                cell.number_format = '#,##0.0000'
            elif fmt == "decimal4":
                cell.number_format = '0.0000'
            if attr == "ebitda_mes_mxn":
                cell.fill = PatternFill("solid", fgColor=_VERDE_EBITDA)

    # Fila de totales con fórmulas SUM
    totales_row = _FILA_DATOS + len(resultado.meses)
    ultima_datos = totales_row - 1
    totales_fill = PatternFill("solid", fgColor=_GRIS_TOTALES)
    totales_font = Font(bold=True)

    ws.cell(totales_row, 1, "TOTAL ANUAL").font = totales_font
    ws.cell(totales_row, 1).fill = totales_fill

    for col_idx, (_titulo, attr, fmt) in enumerate(_COLUMNAS, 1):
        if attr in _COLS_SUMA:
            col_letter = get_column_letter(col_idx)
            formula = f"=SUM({col_letter}{_FILA_DATOS}:{col_letter}{ultima_datos})"
            cell = ws.cell(totales_row, col_idx, formula)
            cell.number_format = '#,##0.00' if fmt == "moneda" else '#,##0.0000'
            cell.font = totales_font
            cell.fill = totales_fill

    # Anchos de columna (A más ancho para las etiquetas de parámetros)
    anchos = [26, 14, 18, 12, 14, 10, 18, 14, 12, 18, 18, 14, 18, 18, 28]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = "B9"


def _formatear(valor: object, fmt: str) -> object:
    """Convierte Decimal/date al tipo nativo que openpyxl acepta mejor."""
    if isinstance(valor, date):
        return valor.strftime("%b %Y")
    if isinstance(valor, Decimal):
        return float(valor)
    return valor


# ── Hoja 2: Parámetros ────────────────────────────────────────────────────────

def _escribir_hoja_parametros(wb: openpyxl.Workbook, params) -> None:
    ws = wb.create_sheet("Parámetros")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_AZUL_HEADER)

    ws.cell(1, 1, "Parámetro").font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 2, "Valor").font = header_font
    ws.cell(1, 2).fill = header_fill

    filas = [
        ("Cobertura eléctrica",  f"{float(params.cobertura_electrica)*100:.0f}%"),
        ("Rendimiento eléctrico",f"{float(params.rendimiento_electrico)*100:.0f}%"),
        ("Rendimiento térmico",  f"{float(params.rendimiento_termico)*100:.0f}%"),
        ("Eficiencia caldera",   f"{float(params.eficiencia_caldera)*100:.0f}%"),
        ("Factor kWh→GJ",        "0.0036 GJ/kWh"),
    ]
    for row_idx, (nombre, valor) in enumerate(filas, 2):
        ws.cell(row_idx, 1, nombre)
        ws.cell(row_idx, 2, valor)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
