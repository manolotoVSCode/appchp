# reports/excel_modelado_chp.py
"""Generador del Excel maestro con fórmulas nativas para Modelado CHP.

Produce un archivo .xlsx con 5 hojas:
  1. "Parámetros"       — inputs editables + valores fijos de simulación + motores
  2. "KPIs Económicos"  — fórmulas que referencian Parámetros
  3. "Tabla Mensual"    — datos históricos fijos de r.meses
  4. "Flujo 15 Años"    — proyección con fórmulas dinámicas
  5. "Curva Mensual"    — datos horarios de la simulación cincominutal

El modelo financiero es "vivo": al cambiar parámetros en la hoja Parámetros
todas las celdas calculadas se actualizan automáticamente.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from models.cogen_result import CoGenResultado

# ── Nombres de hojas ──────────────────────────────────────────────────────────

_SH_PARAMS  = "Parámetros"
_SH_KPIS    = "KPIs Económicos"
_SH_MENSUAL = "Tabla Mensual"
_SH_FLUJO   = "Flujo 15 Años"
_SH_CURVA   = "Curva Mensual"

# ── Filas en hoja Parámetros (valor en columna B) ────────────────────────────

_P_COBERTURA   = 4    # B4  cobertura eléctrica (%) — fijo desde simulación
_P_REND_ELEC   = 5    # B5  rendimiento eléctrico
_P_REND_TERM   = 6    # B6  rendimiento térmico
_P_EFIC_CALD   = 7    # B7  eficiencia caldera
_P_PRECIO_GAS  = 8    # B8  precio gas ($/GJ)
_P_COSTO_OM    = 9    # B9  costo O&M ($/kWh)
_P_TIPO_CAMBIO = 10   # B10 tipo de cambio MXN/USD
_P_PRECIO_KW   = 11   # B11 precio motor USD/kW
_P_CAP_TOTAL   = 12   # B12 capacidad total (kW) — fórmula =SUM(motores)
_P_DEDUCCION   = 13   # B13 deducción fiscal (0/1)
_P_ANIOS_DED   = 14   # B14 años de deducción
_P_CONSUMO_KWH = 15   # B15 consumo cliente anual (kWh) — fijo
_P_KWH_CUB     = 16   # B16 kWh cubiertos anual — fijo
_P_GEN_BRUTA   = 17   # B17 generación bruta anual (kWh) — fijo
_P_GAS_GJ      = 18   # B18 consumo gas simulación (GJ/año) — fijo
_P_HORAS_MOTOR = 19   # B19 horas anuales por motor — fijo
_P_COSTO_PROM  = 20   # B20 costo promedio CFE ($/kWh) — fijo

# Tabla de motores (Sección C)
_P_MOTORES_SECTION  = 22   # Encabezado sección "MOTORES"
_P_MOTORES_HEADERS  = 23   # Encabezados de columna
_P_MOTORES_FIRST    = 24   # Primera fila de datos de motor (hasta +3)

# ── Colores ───────────────────────────────────────────────────────────────────

_C_SECTION_BG    = "1F7A4C"   # Fondo encabezados de sección
_C_SECTION_TXT   = "FFFFFF"   # Texto encabezados
_C_COL_HEADER_BG = "E8F4ED"   # Fondo encabezados de columna
_C_COL_HEADER_TXT= "155936"   # Texto encabezados de columna
_C_EDITABLE_BG   = "FFFBE6"   # Fondo celdas editables (amarillo suave)
_C_EDITABLE_BOR  = "E8B547"   # Borde celdas editables
_C_FIXED_BG      = "F5F5F5"   # Fondo celdas fijas de simulación
_C_FIXED_TXT     = "6C757D"   # Texto celdas fijas
_C_TOTAL_BG      = "E8F4ED"   # Fondo filas de total/resultado
_C_FLUJO_NEG     = "FDECEA"   # Flujo negativo
_C_FLUJO_POS     = "E8F4ED"   # Flujo acumulado positivo
_C_TITLE_TXT     = "1F7A4C"   # Título (A1) de cada hoja

# ── Estilos reutilizables ─────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=None, size=11, italic=False) -> Font:
    kwargs = {"bold": bold, "size": size, "italic": italic}
    if color:
        kwargs["color"] = color
    return Font(**kwargs)


def _side_medium() -> Side:
    return Side(border_style="medium", color="000000")


def _side_thin() -> Side:
    return Side(border_style="thin", color=_C_EDITABLE_BOR)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


# ── Función principal ─────────────────────────────────────────────────────────

def generar_excel_modelado_chp(
    params: dict,
    r: CoGenResultado,
    motores_config: list,
    cliente_nombre: str,
    curva: list | None = None,
    cels_mwh_anual: float | None = None,
    factor_emision_elec: float | None = None,
) -> bytes:
    """Genera el Excel maestro del Modelado CHP y retorna el contenido como bytes.

    Parámetros
    ----------
    params          : dict con todas las celdas de la hoja Parámetros (ver constantes _P_*)
    r               : CoGenResultado de calcular_cogen_desde_modelado()
    motores_config  : lista de dicts {"nombre", "capacidad_kw", "horas_anuales"}
    cliente_nombre  : str para el título A1 de cada hoja
    curva           : lista de puntos horarios {"ts", "demanda_kw", "gen_neta_kw",
                      "gen_por_motor": {id: kw}} — None = sin datos de simulación
    cels_mwh_anual  : valor fijo de CELs (de cels_resultado)
    factor_emision_elec : factor de emisión kgCO₂/kWh (para fórmula CO₂)
    """
    wb = openpyxl.Workbook()
    fecha_gen = datetime.now().strftime("%d/%m/%Y %H:%M")

    _escribir_parametros(wb, params, motores_config, cliente_nombre, fecha_gen)
    _escribir_kpis(wb, r, cels_mwh_anual, factor_emision_elec, cliente_nombre, fecha_gen)
    _escribir_mensual(wb, r, cliente_nombre, fecha_gen)
    _escribir_flujo(wb, r, cliente_nombre, fecha_gen)
    _escribir_curva(wb, motores_config, curva, cliente_nombre, fecha_gen)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 1: Parámetros
# ─────────────────────────────────────────────────────────────────────────────

def _escribir_parametros(
    wb: openpyxl.Workbook,
    params: dict,
    motores_config: list,
    cliente_nombre: str,
    fecha_gen: str,
) -> None:
    ws = wb.create_sheet(_SH_PARAMS)

    # A1: título
    ws["A1"] = f"{cliente_nombre} — Modelado CHP — {fecha_gen}"
    ws["A1"].font = _font(bold=True, color=_C_TITLE_TXT, size=12)

    # Fila 2: sección A header
    _seccion(ws, 2, "INPUTS DEL MODELADO")

    # Fila 3: blank (gap visual antes de row 4)
    # (nada)

    # Sección A — inputs editables (filas 4-14)
    _params_row_fijo(  ws, _P_COBERTURA,   "Cobertura eléctrica (%)",
                       params["cobertura_pct"], "0.00%",
                       nota="Este valor proviene de la simulación cincominutal. "
                            "Para recalcular, ajusta los motores en el dashboard y descarga de nuevo.")
    _params_row_editable(ws, _P_REND_ELEC,   "Rendimiento eléctrico (%)",   params["rendimiento_electrico"], "0.00%")
    _params_row_editable(ws, _P_REND_TERM,   "Rendimiento térmico (%)",     params["rendimiento_termico"],   "0.00%")
    _params_row_editable(ws, _P_EFIC_CALD,   "Eficiencia caldera (%)",      params["eficiencia_caldera"],    "0.00%")
    _params_row_editable(ws, _P_PRECIO_GAS,  "Precio gas ($/GJ)",           params["precio_gas_gj"],         '#,##0.00')
    _params_row_editable(ws, _P_COSTO_OM,    "Costo O&M ($/kWh)",           params["costo_om_kwh"],          '0.0000')
    _params_row_editable(ws, _P_TIPO_CAMBIO, "Tipo de cambio (MXN/USD)",    params["tipo_cambio"],           '#,##0.00')
    _params_row_editable(ws, _P_PRECIO_KW,   "Precio motor (USD/kW)",       params["precio_kw_usd"],         '#,##0.00')

    # B12: fórmula SUM de motores (columna B de tabla motores = col 2 de la hoja)
    # La tabla de motores usa columna B para capacidad_kw
    _params_row_formula(ws, _P_CAP_TOTAL,  "Capacidad total (kW)",
                        f"=SUM(B{_P_MOTORES_FIRST}:B{_P_MOTORES_FIRST+3})", '#,##0')

    _params_row_editable(ws, _P_DEDUCCION, "Deducción fiscal (1=Sí, 0=No)", float(params["deduccion_fiscal"]), "0")
    _params_row_editable(ws, _P_ANIOS_DED, "Años deducción",                 float(params["anios_deduccion"]),  "0")

    # Separador visual antes de sección B
    ws.row_dimensions[_P_CONSUMO_KWH - 1].height = 4

    # Sección B — valores fijos de simulación (filas 15-20)
    nota_b = ("Estos valores provienen de la simulación cincominutal. "
              "Solo cambian al recalcular en el dashboard.")
    _params_row_fijo(ws, _P_CONSUMO_KWH, "Consumo cliente anual (kWh)",   params["consumo_cliente_anual_kwh"], '#,##0', nota=nota_b)
    _params_row_fijo(ws, _P_KWH_CUB,     "kWh cubiertos anual (gen neta)", params["kwh_cubiertos_anual"],       '#,##0')
    _params_row_fijo(ws, _P_GEN_BRUTA,   "Generación bruta anual (kWh)",   params["gen_bruta_anual_kwh"],       '#,##0')
    _params_row_fijo(ws, _P_GAS_GJ,      "Consumo gas simulación (GJ/año)",params["consumo_gas_anual_gj"],      '#,##0.0')
    _params_row_fijo(ws, _P_HORAS_MOTOR, "Horas anuales por motor (h)",    params["horas_anuales_motor"],       '#,##0.0')
    _params_row_fijo(ws, _P_COSTO_PROM,  "Costo promedio CFE ($/kWh)",     params["kwh_costo_promedio_cfe"],    '0.0000')

    # Sección C — tabla de motores
    ws.row_dimensions[21].height = 8  # gap

    _seccion(ws, _P_MOTORES_SECTION, "MOTORES (CONFIGURACIÓN SIMULACIÓN)")

    # Encabezados tabla motores
    hdr_cols = ["Motor", "Capacidad (kW)", "Horas anuales (h)", "% tiempo activo"]
    for col_i, h in enumerate(hdr_cols, 1):
        c = ws.cell(_P_MOTORES_HEADERS, col_i, h)
        c.font  = _font(bold=True, color=_C_COL_HEADER_TXT)
        c.fill  = _fill(_C_COL_HEADER_BG)
        c.alignment = _center()

    # Datos de motores (filas 24-27, hasta 4 motores)
    for idx, m in enumerate(motores_config[:4]):
        fila = _P_MOTORES_FIRST + idx
        horas = float(m.get("horas_anuales", 0))
        cap   = float(m.get("capacidad_kw", 0))
        pct   = round(horas / 8760 * 100, 1) if horas > 0 else 0

        ws.cell(fila, 1, m.get("nombre", f"Motor {idx+1}"))
        c_cap = ws.cell(fila, 2, cap)
        c_cap.number_format = '#,##0'
        ws.cell(fila, 3, horas).number_format = '#,##0.0'
        ws.cell(fila, 4, pct).number_format   = '0.0"%"'

        for col_i in range(1, 5):
            ws.cell(fila, col_i).fill = _fill(_C_FIXED_BG)
            ws.cell(fila, col_i).font = _font(color=_C_FIXED_TXT)

    # Nota motores
    nota_m_row = _P_MOTORES_FIRST + 4 + 1
    ws.cell(nota_m_row, 1,
            "Datos fijos de la simulación. Solo cambian al modificar la configuración de motores y recalcular.")
    ws.cell(nota_m_row, 1).font = _font(italic=True, color=_C_FIXED_TXT, size=9)

    # Anchos de columna
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.freeze_panes = "A2"


def _params_row_editable(ws, row: int, label: str, value, fmt: str) -> None:
    """Escribe una fila de parámetro editable (fondo amarillo, borde naranja)."""
    cA = ws.cell(row, 1, label)
    cA.font = _font(bold=True, size=10)
    cA.alignment = _left()

    cB = ws.cell(row, 2, float(value))
    cB.number_format = fmt
    cB.fill = _fill(_C_EDITABLE_BG)
    cB.alignment = _right()
    thin = _side_thin()
    cB.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _params_row_fijo(ws, row: int, label: str, value, fmt: str, nota: str | None = None) -> None:
    """Escribe una fila de valor fijo de simulación (fondo gris, no editable)."""
    cA = ws.cell(row, 1, label)
    cA.font = _font(color=_C_FIXED_TXT, size=10)
    cA.fill = _fill(_C_FIXED_BG)
    cA.alignment = _left()

    cB = ws.cell(row, 2, float(value))
    cB.number_format = fmt
    cB.fill = _fill(_C_FIXED_BG)
    cB.font = _font(color=_C_FIXED_TXT)
    cB.alignment = _right()

    if nota:
        cB.comment = Comment(nota, "CHP App")


def _params_row_formula(ws, row: int, label: str, formula: str, fmt: str) -> None:
    """Escribe una fila cuyo valor en B es una fórmula."""
    cA = ws.cell(row, 1, label)
    cA.font = _font(bold=True, size=10)
    cA.alignment = _left()

    cB = ws.cell(row, 2, formula)
    cB.number_format = fmt
    cB.fill = _fill(_C_FIXED_BG)
    cB.alignment = _right()


def _seccion(ws, row: int, titulo: str) -> None:
    """Escribe una fila de encabezado de sección (fondo verde, texto blanco)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row, 1, titulo)
    c.font  = _font(bold=True, color=_C_SECTION_TXT, size=10)
    c.fill  = _fill(_C_SECTION_BG)
    c.alignment = _left()
    ws.row_dimensions[row].height = 18


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 2: KPIs Económicos
# ─────────────────────────────────────────────────────────────────────────────

# Filas de cada KPI en la hoja KPIs Económicos
_K = {
    "sxn_inversion": 3,
    "inversion_usd": 4,
    "inversion_mxn": 5,
    "benef_fiscal":  6,
    "sxn_gen":       8,
    "gen_neta":      9,
    "gen_bruta":     10,
    "cobertura":     11,
    "sxn_ingresos":  13,
    "ah_elec":       14,
    "ah_caldera":    15,
    "total_ingresos":16,
    "sxn_gastos":    18,
    "costo_gas":     19,
    "om":            20,
    "total_gastos":  21,
    "sxn_resultado": 23,
    "ahorro_neto":   24,
    "payback":       25,
    "payback_benef": 26,
    "sxn_ambiental": 28,
    "co2_red":       29,
    "co2_red_pct":   30,
    "cels":          31,
    "sxn_notas":     33,
}

# Referencia a Parámetros (nombre de hoja con apóstrofe para seguridad)
_P = f"'{_SH_PARAMS}'"


def _p(row: int) -> str:
    """Referencia absoluta a celda B de Parámetros."""
    return f"{_P}!$B${row}"


def _escribir_kpis(
    wb: openpyxl.Workbook,
    r: CoGenResultado,
    cels_mwh_anual: float | None,
    factor_emision_elec: float | None,
    cliente_nombre: str,
    fecha_gen: str,
) -> None:
    ws = wb.create_sheet(_SH_KPIS)

    # A1: título
    ws["A1"] = f"{cliente_nombre} — KPIs Económicos — {fecha_gen}"
    ws["A1"].font = _font(bold=True, color=_C_TITLE_TXT, size=12)

    # ── SECCIÓN INVERSIÓN ────────────────────────────────────────────────────
    _kpi_seccion(ws, _K["sxn_inversion"], "INVERSIÓN")

    # Inversión USD = Capacidad total × Precio/kW
    inv_usd_row = _K["inversion_usd"]
    _kpi_row(ws, inv_usd_row, "Inversión (USD)",
             f"={_p(_P_CAP_TOTAL)}*{_p(_P_PRECIO_KW)}", '$#,##0')

    # Inversión MXN = Inv USD × tipo_cambio
    inv_mxn_row = _K["inversion_mxn"]
    _kpi_row(ws, inv_mxn_row, "Inversión (MXN)",
             f"=B{inv_usd_row}*{_p(_P_TIPO_CAMBIO)}", '$#,##0')

    # Beneficio Fiscal año 1 = SI(deduccion=1, inv_mxn*0.3/anios, 0)
    bf_row = _K["benef_fiscal"]
    _kpi_row(ws, bf_row, "Beneficio fiscal año 1 (MXN)",
             f"=SI({_p(_P_DEDUCCION)}=1,B{inv_mxn_row}*0.3/{_p(_P_ANIOS_DED)},0)", '$#,##0')

    # ── SECCIÓN GENERACIÓN ───────────────────────────────────────────────────
    _kpi_seccion(ws, _K["sxn_gen"], "GENERACIÓN")

    _kpi_row(ws, _K["gen_neta"],   "Gen. neta anual (kWh)",   f"={_p(_P_KWH_CUB)}",     '#,##0')
    _kpi_row(ws, _K["gen_bruta"],  "Gen. bruta anual (kWh)",  f"={_p(_P_GEN_BRUTA)}",   '#,##0')
    _kpi_row(ws, _K["cobertura"],  "Cobertura (%)",           f"={_p(_P_COBERTURA)}",   '0.00%')

    # ── SECCIÓN INGRESOS ─────────────────────────────────────────────────────
    _kpi_seccion(ws, _K["sxn_ingresos"], "INGRESOS")

    ah_elec_row = _K["ah_elec"]
    # Ahorro electricidad = kWh cubiertos × costo promedio CFE
    _kpi_row(ws, ah_elec_row, "Ahorro electricidad (MXN)",
             f"={_p(_P_KWH_CUB)}*{_p(_P_COSTO_PROM)}", '$#,##0')

    ah_cal_row = _K["ah_caldera"]
    # Ahorro caldera = Gen bruta (GJ) × rend_termico × (1 - 1/efic_caldera) × precio_gas
    # GJ = kWh × 0.0036
    _kpi_row(ws, ah_cal_row, "Ahorro caldera (MXN)",
             f"={_p(_P_GEN_BRUTA)}*0.0036*{_p(_P_REND_TERM)}"
             f"*(1-1/{_p(_P_EFIC_CALD)})*{_p(_P_PRECIO_GAS)}", '$#,##0')

    tot_ing_row = _K["total_ingresos"]
    _kpi_row(ws, tot_ing_row, "Total ingresos (MXN)",
             f"=B{ah_elec_row}+B{ah_cal_row}", '$#,##0', total=True)

    # ── SECCIÓN GASTOS ───────────────────────────────────────────────────────
    _kpi_seccion(ws, _K["sxn_gastos"], "GASTOS")

    gas_row = _K["costo_gas"]
    # Costo gas cogen = Gen bruta (GJ) / rend_electrico × precio_gas
    _kpi_row(ws, gas_row, "Costo gas cogeneración (MXN)",
             f"={_p(_P_GEN_BRUTA)}*0.0036/{_p(_P_REND_ELEC)}*{_p(_P_PRECIO_GAS)}", '$#,##0')

    om_row = _K["om"]
    # O&M = Gen bruta × costo_om_kwh
    _kpi_row(ws, om_row, "O&M estimado (MXN)",
             f"={_p(_P_GEN_BRUTA)}*{_p(_P_COSTO_OM)}", '$#,##0')

    tot_gas_row = _K["total_gastos"]
    _kpi_row(ws, tot_gas_row, "Total gastos (MXN)",
             f"=B{gas_row}+B{om_row}", '$#,##0', total=True)

    # ── SECCIÓN RESULTADO ────────────────────────────────────────────────────
    _kpi_seccion(ws, _K["sxn_resultado"], "RESULTADO")

    neto_row = _K["ahorro_neto"]
    _kpi_row(ws, neto_row, "Ahorro neto / EBITDA (MXN)",
             f"=B{tot_ing_row}-B{tot_gas_row}", '$#,##0', total=True)

    pb_row = _K["payback"]
    _kpi_row(ws, pb_row, "Payback (años)",
             f'=SI(B{neto_row}>0,B{inv_mxn_row}/B{neto_row},"N/A")', '0.00')

    pb_bf_row = _K["payback_benef"]
    _kpi_row(ws, pb_bf_row, "Payback con benef. fiscal (años)",
             f'=SI(B{neto_row}>0,'
             f'SI(B{bf_row}>0,(B{inv_mxn_row}-B{bf_row})/B{neto_row},B{inv_mxn_row}/B{neto_row}),'
             f'"N/A")', '0.00')

    # ── SECCIÓN AMBIENTAL ────────────────────────────────────────────────────
    _kpi_seccion(ws, _K["sxn_ambiental"], "AMBIENTAL")

    fe = round(factor_emision_elec, 4) if factor_emision_elec else 0.435
    co2_row = _K["co2_red"]
    # CO₂ reducción = kWh cubiertos / 1000 × factor_emision (toneladas)
    _kpi_row(ws, co2_row, "CO₂ reducción (t CO₂/año)",
             f"={_p(_P_KWH_CUB)}/1000*{fe}", '#,##0.0')
    ws.cell(co2_row, 3,
            f"Factor emisión: {fe} kgCO₂/kWh").font = _font(italic=True, color=_C_FIXED_TXT, size=9)

    co2pct_row = _K["co2_red_pct"]
    _kpi_row(ws, co2pct_row, "CO₂ reducción (%)",
             f"=SI({_p(_P_CONSUMO_KWH)}>0,B{co2_row}/({_p(_P_CONSUMO_KWH)}/1000*{fe}),0)",
             '0.0%')

    cels_row = _K["cels"]
    cels_val = round(cels_mwh_anual, 2) if cels_mwh_anual is not None else 0
    _kpi_row(ws, cels_row, "CELs (MWh/año)",
             cels_val, '#,##0.0', fixed=True)
    ws.cell(cels_row, 3,
            "Calculado según tabla CRE con RefE. Solo cambia al recalcular en el dashboard."
            ).font = _font(italic=True, color=_C_FIXED_TXT, size=9)

    # ── NOTAS DE FÓRMULAS ────────────────────────────────────────────────────
    _kpi_seccion(ws, _K["sxn_notas"], "NOTAS DE FÓRMULAS")

    notas = [
        "Ahorro electricidad = kWh cubiertos (B16) × costo promedio CFE (B20)",
        "Ahorro caldera      = Gen bruta (GJ) × rend_térmico × (1 - 1/efic_caldera) × precio_gas",
        "Costo gas           = Gen bruta (GJ) / rend_eléctrico × precio_gas",
        "O&M                 = Gen bruta (kWh) × costo_om_kwh",
        "Inversión           = Cap total (kW) × precio_kw (USD) × tipo_cambio",
        "Payback             = Inversión MXN / Ahorro Neto Anual",
        f"CO₂ reducción      = kWh cubiertos × {fe} kgCO₂/kWh ÷ 1000",
    ]
    for i, nota in enumerate(notas):
        fila = _K["sxn_notas"] + 1 + i
        c = ws.cell(fila, 1, nota)
        c.font = _font(color=_C_FIXED_TXT, size=9, italic=True)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)

    # Anchos
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 45
    ws.freeze_panes = "A2"


def _kpi_seccion(ws, row: int, titulo: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row, 1, titulo)
    c.font  = _font(bold=True, color=_C_SECTION_TXT, size=10)
    c.fill  = _fill(_C_SECTION_BG)
    c.alignment = _left()
    ws.row_dimensions[row].height = 18


def _kpi_row(ws, row: int, label: str, value_or_formula,
             fmt: str = '#,##0', total: bool = False, fixed: bool = False) -> None:
    """Escribe etiqueta en A y valor/fórmula en B con formato."""
    cA = ws.cell(row, 1, label)
    cA.font = _font(bold=total, size=10)
    if total:
        cA.fill = _fill(_C_TOTAL_BG)

    cB = ws.cell(row, 2, value_or_formula)
    cB.number_format = fmt
    cB.alignment = _right()
    if total:
        cB.fill  = _fill(_C_TOTAL_BG)
        cB.font  = _font(bold=True)
    elif fixed:
        cB.fill = _fill(_C_FIXED_BG)
        cB.font = _font(color=_C_FIXED_TXT)


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 3: Tabla Mensual
# ─────────────────────────────────────────────────────────────────────────────

_M_COLS = [
    ("Período",                   "periodo_inicio",           "fecha"),
    ("kWh cubiertos",             "kwh_cubiertos",            '#,##0'),
    ("Ahorro Elec. (MXN)",        "ahorro_electricidad_mxn",  '$#,##0'),
    ("Ahorro Caldera (MXN)",      "ahorro_caldera_mxn",       '$#,##0'),
    ("Costo Gas (MXN)",           "costo_gas_cogen_mxn",      '$#,##0'),
    ("O&M (MXN)",                 "gasto_om_mes_mxn",         '$#,##0'),
    ("Ahorro Neto (MXN)",         "ebitda_mes_mxn",           '$#,##0'),
]

_M_SUMA_COLS = {1, 2, 3, 4, 5, 6}  # índices 0-based de columnas con SUM (excl. fecha)


def _escribir_mensual(
    wb: openpyxl.Workbook,
    r: CoGenResultado,
    cliente_nombre: str,
    fecha_gen: str,
) -> None:
    ws = wb.create_sheet(_SH_MENSUAL)

    ws["A1"] = f"{cliente_nombre} — Tabla Mensual — {fecha_gen}"
    ws["A1"].font = _font(bold=True, color=_C_TITLE_TXT, size=12)

    nota_row = 2
    ws.cell(nota_row, 1,
            "Datos históricos de facturas reales. Esta tabla no se recalcula al cambiar parámetros.")
    ws.cell(nota_row, 1).font = _font(italic=True, color=_C_FIXED_TXT, size=9)

    hdr_row = 3
    for col_i, (titulo, _, _f) in enumerate(_M_COLS, 1):
        c = ws.cell(hdr_row, col_i, titulo)
        c.font  = _font(bold=True, color=_C_COL_HEADER_TXT)
        c.fill  = _fill(_C_COL_HEADER_BG)
        c.alignment = _center()

    if not r.meses:
        ws.cell(hdr_row + 1, 1, "Sin datos disponibles").font = _font(italic=True, color=_C_FIXED_TXT)
    else:
        datos_inicio = hdr_row + 1
        for offset, mes in enumerate(r.meses):
            fila = datos_inicio + offset
            for col_i, (_t, attr, fmt) in enumerate(_M_COLS, 1):
                if fmt == "fecha":
                    val = mes.periodo_inicio.strftime("%b %Y")
                else:
                    val = float(getattr(mes, attr))
                c = ws.cell(fila, col_i, val)
                if fmt != "fecha":
                    c.number_format = fmt
                    c.alignment = _right()

        # Fila de totales
        total_row = datos_inicio + len(r.meses)
        ws.cell(total_row, 1, "TOTAL ANUAL").font = _font(bold=True)
        ws.cell(total_row, 1).fill = _fill(_C_TOTAL_BG)

        ultima = total_row - 1
        for col_i, (_t, attr, fmt) in enumerate(_M_COLS, 1):
            if col_i - 1 in _M_SUMA_COLS:
                col_letter = get_column_letter(col_i)
                formula = f"=SUM({col_letter}{datos_inicio}:{col_letter}{ultima})"
                c = ws.cell(total_row, col_i, formula)
                c.number_format = fmt
                c.font = _font(bold=True)
                c.fill = _fill(_C_TOTAL_BG)
                c.alignment = _right()

    ws.column_dimensions["A"].width = 14
    for col_i in range(2, len(_M_COLS) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 22
    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 4: Flujo 15 Años
# ─────────────────────────────────────────────────────────────────────────────

def _escribir_flujo(
    wb: openpyxl.Workbook,
    r: CoGenResultado,
    cliente_nombre: str,
    fecha_gen: str,
) -> None:
    ws = wb.create_sheet(_SH_FLUJO)

    ws["A1"] = f"{cliente_nombre} — Flujo 15 Años — {fecha_gen}"
    ws["A1"].font = _font(bold=True, color=_C_TITLE_TXT, size=12)

    # Encabezados
    hdr_row = 2
    hdrs = ["Año", "Flujo anual (MXN)", "Flujo acumulado (MXN)", "Nota"]
    for col_i, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, col_i, h)
        c.font  = _font(bold=True, color=_C_COL_HEADER_TXT)
        c.fill  = _fill(_C_COL_HEADER_BG)
        c.alignment = _center()

    datos_inicio = hdr_row + 1
    inv_mxn_row  = _K["inversion_mxn"]
    bf_row       = _K["benef_fiscal"]
    neto_row     = _K["ahorro_neto"]
    kpis_name    = _SH_KPIS

    for anio in range(0, 16):
        fila = datos_inicio + anio

        # Columna A: Año
        ws.cell(fila, 1, anio).alignment = _center()
        ws.cell(fila, 1).number_format = "0"

        # Columna B: Flujo anual
        if anio == 0:
            # Inversión inicial (negativa)
            flujo_formula = f"=-'{kpis_name}'!$B${inv_mxn_row}"
            nota = "Inversión inicial"
        elif anio == 1:
            # Año 1 incluye beneficio fiscal
            flujo_formula = (
                f"='{kpis_name}'!$B${neto_row}"
                f"+'{kpis_name}'!$B${bf_row}"
            )
            nota = "Con beneficio fiscal ISR"
        else:
            flujo_formula = f"='{kpis_name}'!$B${neto_row}"
            nota = f"Año {anio}"

        cB = ws.cell(fila, 2, flujo_formula)
        cB.number_format = '$#,##0'
        cB.alignment = _right()

        # Columna C: Flujo acumulado
        if anio == 0:
            acum_formula = f"=B{fila}"
        else:
            acum_formula = f"=C{fila-1}+B{fila}"

        cC = ws.cell(fila, 3, acum_formula)
        cC.number_format = '$#,##0'
        cC.alignment = _right()

        # Columna D: nota
        ws.cell(fila, 4, nota).font = _font(color=_C_FIXED_TXT, size=9)

        # Formato condicional básico: colores por contenido numérico
        # (openpyxl no evalúa fórmulas, así que usamos el conocimiento del dominio)
        if anio == 0:
            # Inversión = negativo
            for col_i in range(1, 4):
                ws.cell(fila, col_i).fill = _fill(_C_FLUJO_NEG)

    # Anchos
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 28
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 5: Curva Mensual
# ─────────────────────────────────────────────────────────────────────────────

def _escribir_curva(
    wb: openpyxl.Workbook,
    motores_config: list,
    curva: list | None,
    cliente_nombre: str,
    fecha_gen: str,
) -> None:
    ws = wb.create_sheet(_SH_CURVA)

    ws["A1"] = f"{cliente_nombre} — Curva Mensual — {fecha_gen}"
    ws["A1"].font = _font(bold=True, color=_C_TITLE_TXT, size=12)

    nota_row = 2
    ws.cell(nota_row, 1,
            "Datos de la simulación cincominutal agregados por hora. "
            "Solo cambian al recalcular en el dashboard.")
    ws.cell(nota_row, 1).font = _font(italic=True, color=_C_FIXED_TXT, size=9)

    # Determinar columnas de motores
    nombres_motores = [m.get("nombre", f"Motor {i+1}") for i, m in enumerate(motores_config[:4])]
    motor_ids = [str(m.get("id", i + 1)) for i, m in enumerate(motores_config[:4])]

    hdr_row = 3
    hdrs = ["Timestamp", "Demanda (kW)", "Gen Total (kW)"] + [f"{n} (kW)" for n in nombres_motores]
    for col_i, h in enumerate(hdrs, 1):
        c = ws.cell(hdr_row, col_i, h)
        c.font  = _font(bold=True, color=_C_COL_HEADER_TXT)
        c.fill  = _fill(_C_COL_HEADER_BG)
        c.alignment = _center()

    if not curva:
        ws.cell(hdr_row + 1, 1,
                "Sin datos de simulación disponibles.").font = _font(italic=True, color=_C_FIXED_TXT)
    else:
        # Agregar por hora
        from collections import defaultdict
        hourly: dict[str, dict] = defaultdict(lambda: {"dem": [], "gen": [],
                                                        "motores": defaultdict(list)})
        for p in curva:
            ts = str(p.get("ts", ""))
            hora = ts[:13]  # "YYYY-MM-DD HH"
            hourly[hora]["dem"].append(float(p.get("demanda_kw", 0)))
            hourly[hora]["gen"].append(float(p.get("gen_neta_kw", 0)))
            gpm = p.get("gen_por_motor") or {}
            for mid in motor_ids:
                hourly[hora]["motores"][mid].append(float(gpm.get(mid, 0)))

        datos_inicio = hdr_row + 1
        for row_offset, (hora_key, vals) in enumerate(sorted(hourly.items())):
            fila = datos_inicio + row_offset
            n = len(vals["dem"]) or 1
            dem_avg = sum(vals["dem"]) / n
            gen_avg = sum(vals["gen"]) / n

            ws.cell(fila, 1, hora_key + ":00").alignment = _left()
            ws.cell(fila, 2, round(dem_avg, 2)).number_format = '#,##0.0'
            ws.cell(fila, 3, round(gen_avg, 2)).number_format = '#,##0.0'

            for m_col, mid in enumerate(motor_ids, 4):
                m_vals = vals["motores"][mid]
                m_avg  = sum(m_vals) / len(m_vals) if m_vals else 0
                ws.cell(fila, m_col, round(m_avg, 2)).number_format = '#,##0.0'

    # Anchos
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    for i in range(4, 4 + len(nombres_motores)):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "A4"
