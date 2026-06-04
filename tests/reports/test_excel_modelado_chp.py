# tests/reports/test_excel_modelado_chp.py
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest

from models.cogen_result import CoGenMes, CoGenParams, CoGenResultado
from reports.excel_modelado_chp import (
    generar_excel_modelado_chp,
    _SH_PARAMS,
    _SH_KPIS,
    _SH_MENSUAL,
    _SH_FLUJO,
    _SH_CURVA,
    _P_COBERTURA,
    _P_REND_ELEC,
    _P_REND_TERM,
    _P_EFIC_CALD,
    _P_PRECIO_GAS,
    _P_COSTO_OM,
    _P_TIPO_CAMBIO,
    _P_PRECIO_KW,
    _P_CAP_TOTAL,
    _P_DEDUCCION,
    _P_ANIOS_DED,
    _P_CONSUMO_KWH,
    _P_KWH_CUB,
    _P_GEN_BRUTA,
    _P_GAS_GJ,
    _P_HORAS_MOTOR,
    _P_COSTO_PROM,
)


# ── Fixtures ──────────────────────────────────────────────────────────────────

def _mes(year: int, month: int, **kw) -> CoGenMes:
    d = dict(
        periodo_inicio=date(year, month, 1),
        periodo_fin=date(year, month, 28),
        kwh_total=Decimal("100000"),
        costo_cfe_mxn=Decimal("280000"),
        costo_promedio_kwh=Decimal("2.8"),
        gj_consumido=Decimal("500"),
        costo_unitario_gj=Decimal("220"),
        costo_gas_actual_mxn=Decimal("110000"),
        kwh_cubiertos=Decimal("75000"),
        kwh_punta_cubierto=Decimal("20000"),
        kwh_intermedia_cubierto=Decimal("30000"),
        kwh_base_cubierto=Decimal("25000"),
        ahorro_energia_mes_mxn=Decimal("150000"),
        ahorro_capacidad_mes_mxn=Decimal("40000"),
        ahorro_distribucion_mes_mxn=Decimal("20000"),
        gj_gas_cogen=Decimal("748"),
        costo_gas_cogen_mxn=Decimal("164560"),
        ahorro_electricidad_mxn=Decimal("210000"),
        calor_recuperado_gj=Decimal("187"),
        ahorro_caldera_mxn=Decimal("44000"),
        gasto_om_mes_mxn=Decimal("22500"),
        ebitda_mes_mxn=Decimal("66940"),
        prorrateado=False,
        nota_prorrateo="",
    )
    d.update(kw)
    return CoGenMes(**d)


@pytest.fixture()
def resultado() -> CoGenResultado:
    params = CoGenParams(
        cobertura_electrica=Decimal("0.75"),
        rendimiento_electrico=Decimal("0.40"),
        rendimiento_termico=Decimal("0.25"),
        eficiencia_caldera=Decimal("0.85"),
    )
    meses = [_mes(2024, m) for m in range(1, 13)]
    return CoGenResultado(
        params=params,
        meses=meses,
        kwh_total_anual=Decimal("1200000"),
        kwh_cubiertos_anual=Decimal("900000"),
        gj_gas_cogen_anual=Decimal("8976"),
        costo_gas_cogen_anual_mxn=Decimal("1974720"),
        ahorro_electricidad_anual_mxn=Decimal("2520000"),
        ahorro_caldera_anual_mxn=Decimal("528000"),
        ebitda_anual_mxn=Decimal("803280"),
        gasto_om_anual_mxn=Decimal("270000"),
        inversion_usd=Decimal("1330000"),
        inversion_mxn=Decimal("23275000"),
        tipo_cambio_mxn_usd=Decimal("17.50"),
        beneficio_fiscal_anio_1_mxn=Decimal("6982500"),
        flujo_anio_1_con_beneficio_mxn=Decimal("7785780"),
    )


@pytest.fixture()
def params_dict() -> dict:
    return {
        "cobertura_pct": 0.75,
        "rendimiento_electrico": 0.40,
        "rendimiento_termico": 0.25,
        "eficiencia_caldera": 0.85,
        "precio_gas_gj": 220.0,
        "costo_om_kwh": 0.30,
        "tipo_cambio": 17.50,
        "precio_kw_usd": 1400.0,
        "deduccion_fiscal": 1,
        "anios_deduccion": 1,
        "consumo_cliente_anual_kwh": 1200000.0,
        "kwh_cubiertos_anual": 900000.0,
        "gen_bruta_anual_kwh": 920000.0,
        "consumo_gas_anual_gj": 8976.0,
        "horas_anuales_motor": 7200.0,
        "kwh_costo_promedio_cfe": 2.80,
    }


@pytest.fixture()
def motores_config() -> list:
    return [
        {"nombre": "Motor 1", "capacidad_kw": 950, "horas_anuales": 7200},
    ]


@pytest.fixture()
def wb(resultado, params_dict, motores_config) -> openpyxl.Workbook:
    buf = generar_excel_modelado_chp(
        params=params_dict,
        r=resultado,
        motores_config=motores_config,
        cliente_nombre="IBERICA TILES",
    )
    return openpyxl.load_workbook(BytesIO(buf))


# ── Test 1: Estructura — 5 hojas con nombres correctos ───────────────────────

def test_cinco_hojas(wb):
    assert set(wb.sheetnames) == {_SH_PARAMS, _SH_KPIS, _SH_MENSUAL, _SH_FLUJO, _SH_CURVA}


# ── Test 2: Hoja Parámetros — celdas B correctas ─────────────────────────────

def test_parametros_b4_cobertura(wb, params_dict):
    ws = wb[_SH_PARAMS]
    val = ws.cell(_P_COBERTURA, 2).value
    # B4 es valor fijo (float), no fórmula
    assert isinstance(val, float), f"B4 esperado float, obtenido {val!r}"
    assert abs(val - params_dict["cobertura_pct"]) < 1e-6


def test_parametros_secciones_editables(wb, params_dict):
    """B5-B11, B13-B14 son valores numéricos editables (no fórmulas de texto)."""
    ws = wb[_SH_PARAMS]
    editables = [
        (_P_REND_ELEC,   params_dict["rendimiento_electrico"]),
        (_P_REND_TERM,   params_dict["rendimiento_termico"]),
        (_P_EFIC_CALD,   params_dict["eficiencia_caldera"]),
        (_P_PRECIO_GAS,  params_dict["precio_gas_gj"]),
        (_P_COSTO_OM,    params_dict["costo_om_kwh"]),
        (_P_TIPO_CAMBIO, params_dict["tipo_cambio"]),
        (_P_PRECIO_KW,   params_dict["precio_kw_usd"]),
        (_P_DEDUCCION,   float(params_dict["deduccion_fiscal"])),
        (_P_ANIOS_DED,   float(params_dict["anios_deduccion"])),
    ]
    for row, expected in editables:
        val = ws.cell(row, 2).value
        assert isinstance(val, (int, float)), f"Fila {row}: esperado numérico, obtenido {val!r}"
        assert abs(float(val) - expected) < 1e-4, f"Fila {row}: esperado {expected}, obtenido {val}"


def test_parametros_b12_es_formula(wb):
    """B12 (capacidad total) es fórmula SUM de la tabla de motores."""
    ws = wb[_SH_PARAMS]
    val = ws.cell(_P_CAP_TOTAL, 2).value
    assert isinstance(val, str) and val.startswith("="), f"B12 debe ser fórmula, obtenido {val!r}"


def test_parametros_seccion_fija(wb, params_dict):
    """B15-B20 son valores numéricos fijos (simulación)."""
    ws = wb[_SH_PARAMS]
    fijos = [
        (_P_CONSUMO_KWH, params_dict["consumo_cliente_anual_kwh"]),
        (_P_KWH_CUB,     params_dict["kwh_cubiertos_anual"]),
        (_P_GEN_BRUTA,   params_dict["gen_bruta_anual_kwh"]),
        (_P_GAS_GJ,      params_dict["consumo_gas_anual_gj"]),
        (_P_HORAS_MOTOR, params_dict["horas_anuales_motor"]),
        (_P_COSTO_PROM,  params_dict["kwh_costo_promedio_cfe"]),
    ]
    for row, expected in fijos:
        val = ws.cell(row, 2).value
        assert isinstance(val, (int, float)), f"Fila {row}: esperado numérico, obtenido {val!r}"
        assert abs(float(val) - expected) < 1e-3, f"Fila {row}: esperado {expected}, obtenido {val}"


def test_parametros_a1_tiene_titulo(wb):
    ws = wb[_SH_PARAMS]
    a1 = ws["A1"].value or ""
    assert "IBERICA TILES" in a1, f"A1 debe incluir nombre cliente: {a1!r}"


# ── Test 3: Hoja KPIs Económicos — fórmulas clave ────────────────────────────

def test_kpis_inversion_usd_formula(wb):
    """La celda de Inversión USD contiene fórmula referenciando Parámetros B12 y B11."""
    ws = wb[_SH_KPIS]
    # Buscar la celda con "Inversión" en columna A y valor con fórmula en B
    found = False
    for row in ws.iter_rows():
        label = row[0].value or ""
        if "Invers" in str(label) and "USD" in str(label):
            formula = row[1].value or ""
            assert isinstance(formula, str) and formula.startswith("="), \
                f"Inversión USD debe ser fórmula, obtenido {formula!r}"
            assert f"$B${_P_CAP_TOTAL}" in formula, f"Debe referenciar B{_P_CAP_TOTAL}: {formula!r}"
            assert f"$B${_P_PRECIO_KW}" in formula, f"Debe referenciar B{_P_PRECIO_KW}: {formula!r}"
            found = True
            break
    assert found, "No se encontró celda 'Inversión USD' en hoja KPIs Económicos"


def test_kpis_ahorro_elec_formula(wb):
    """Ahorro electricidad referencia B16 (kwh_cubiertos) y B20 (costo_promedio)."""
    ws = wb[_SH_KPIS]
    found = False
    for row in ws.iter_rows():
        label = row[0].value or ""
        if "Ahorro elec" in str(label) or "Ahorro Elec" in str(label):
            formula = row[1].value or ""
            assert isinstance(formula, str) and formula.startswith("="), \
                f"Ahorro electricidad debe ser fórmula: {formula!r}"
            assert f"$B${_P_KWH_CUB}" in formula, f"Debe referenciar B{_P_KWH_CUB}: {formula!r}"
            assert f"$B${_P_COSTO_PROM}" in formula, f"Debe referenciar B{_P_COSTO_PROM}: {formula!r}"
            found = True
            break
    assert found, "No se encontró celda 'Ahorro electricidad' en hoja KPIs Económicos"


def test_kpis_ahorro_neto_formula(wb):
    """Ahorro neto es diferencia entre ingresos y gastos."""
    ws = wb[_SH_KPIS]
    found = False
    for row in ws.iter_rows():
        label = row[0].value or ""
        if "Ahorro neto" in str(label) or "EBITDA" in str(label):
            formula = row[1].value or ""
            assert isinstance(formula, str) and formula.startswith("="), \
                f"Ahorro neto debe ser fórmula: {formula!r}"
            found = True
            break
    assert found, "No se encontró celda 'Ahorro neto' en hoja KPIs"


def test_kpis_payback_formula(wb):
    """Payback usa fórmula SI() — no N/A hardcodeado."""
    ws = wb[_SH_KPIS]
    found = False
    for row in ws.iter_rows():
        label = row[0].value or ""
        if "Payback" in str(label) and "benef" not in str(label).lower():
            formula = row[1].value or ""
            assert isinstance(formula, str) and formula.startswith("=SI("), \
                f"Payback debe ser fórmula SI(): {formula!r}"
            found = True
            break
    assert found, "No se encontró celda 'Payback' en hoja KPIs"


def test_kpis_a1_tiene_titulo(wb):
    ws = wb[_SH_KPIS]
    a1 = ws["A1"].value or ""
    assert "IBERICA TILES" in a1


# ── Test 4: Tabla Mensual — datos y totales ───────────────────────────────────

def test_tabla_mensual_filas(wb, resultado):
    """Tabla mensual tiene tantas filas de datos como r.meses."""
    ws = wb[_SH_MENSUAL]
    # Contar filas con datos no vacíos (excluyendo título, encabezados, total)
    n_meses = len(resultado.meses)
    # Buscar fila de total
    total_row = None
    for row in ws.iter_rows():
        label = row[0].value or ""
        if "TOTAL" in str(label).upper() or "Total" in str(label):
            total_row = row[0].row
            break
    assert total_row is not None, "Debe haber fila de total en Tabla Mensual"
    # Las filas de datos deben ser al menos n_meses
    assert total_row >= n_meses + 2  # al menos n_meses + header rows


def test_tabla_mensual_totales_son_formulas(wb):
    """La fila TOTAL debe tener fórmulas =SUM() en columnas de valores."""
    ws = wb[_SH_MENSUAL]
    for row in ws.iter_rows():
        label = row[0].value or ""
        if "TOTAL" in str(label).upper() or "Total" in str(label):
            # Al menos una celda de suma debe ser fórmula
            suma_cells = [c.value for c in row[1:] if isinstance(c.value, str) and c.value.startswith("=SUM")]
            assert len(suma_cells) >= 3, f"Fila Total debe tener ≥3 =SUM(), encontrado: {suma_cells}"
            break


def test_tabla_mensual_a1_titulo(wb):
    ws = wb[_SH_MENSUAL]
    a1 = ws["A1"].value or ""
    assert "IBERICA TILES" in a1


# ── Test 5: Flujo 15 Años ─────────────────────────────────────────────────────

def test_flujo_16_filas_datos(wb):
    """Flujo 15 Años debe tener filas para años 0-15 (16 filas de datos)."""
    ws = wb[_SH_FLUJO]
    # Contar filas con valor numérico o fórmula en col A (año)
    anio_rows = []
    for row in ws.iter_rows(min_col=1, max_col=1):
        val = row[0].value
        if isinstance(val, (int, float)) and 0 <= val <= 15:
            anio_rows.append(val)
    assert len(anio_rows) == 16, f"Deben existir 16 filas (años 0-15), encontrado: {anio_rows}"


def test_flujo_anio0_es_negativo(wb):
    """Año 0 tiene flujo negativo (inversión)."""
    ws = wb[_SH_FLUJO]
    for row in ws.iter_rows():
        anio_val = row[0].value
        if anio_val == 0:
            flujo = row[1].value
            # Puede ser valor negativo o fórmula que empieza con =-
            if isinstance(flujo, (int, float)):
                assert flujo < 0, f"Año 0 debe ser negativo, obtenido {flujo}"
            elif isinstance(flujo, str):
                assert flujo.startswith("=-") or "-" in flujo, f"Fórmula año 0 debe dar negativo: {flujo!r}"
            break


def test_flujo_anios_2_15_referencian_ahorro_neto(wb):
    """Años 2-15 referencian la celda de ahorro neto (fórmula)."""
    ws = wb[_SH_FLUJO]
    for row in ws.iter_rows():
        anio_val = row[0].value
        if anio_val == 2:
            flujo = row[1].value
            assert isinstance(flujo, str) and flujo.startswith("="), \
                f"Año 2 debe ser fórmula, obtenido {flujo!r}"
            # Debe referenciar hoja KPIs Económicos
            kpis_name = _SH_KPIS.replace(" ", " ")
            assert "KPIs" in flujo or "EBITDA" in flujo or "ahorro" in flujo.lower(), \
                f"Año 2 debe referenciar KPIs: {flujo!r}"
            break


def test_flujo_acumulado_formula(wb):
    """Flujo acumulado (col C) desde año 1 es fórmula C_anterior + B_actual."""
    ws = wb[_SH_FLUJO]
    for row in ws.iter_rows():
        anio_val = row[0].value
        if anio_val == 1:
            acum = row[2].value  # col C = flujo acumulado
            assert isinstance(acum, str) and acum.startswith("="), \
                f"Flujo acumulado año 1 debe ser fórmula, obtenido {acum!r}"
            break


def test_flujo_a1_titulo(wb):
    ws = wb[_SH_FLUJO]
    a1 = ws["A1"].value or ""
    assert "IBERICA TILES" in a1


# ── Test 6: Curva Mensual — encabezados presentes ─────────────────────────────

def test_curva_mensual_encabezados(wb):
    """Curva Mensual tiene encabezados con Timestamp, Demanda, Gen Total."""
    ws = wb[_SH_CURVA]
    # Buscar fila de encabezados
    found_ts = found_dem = found_gen = False
    for row in ws.iter_rows(max_row=10):
        for cell in row:
            v = str(cell.value or "")
            if "Timestamp" in v or "timestamp" in v:
                found_ts = True
            if "Demanda" in v or "demanda" in v:
                found_dem = True
            if "Gen" in v or "gen" in v:
                found_gen = True
    assert found_ts, "Falta columna Timestamp en Curva Mensual"
    assert found_dem, "Falta columna Demanda en Curva Mensual"
    assert found_gen, "Falta columna Gen Total en Curva Mensual"


def test_curva_sin_datos_muestra_nota(wb):
    """Sin curva (curva=None), la hoja muestra nota en lugar de datos."""
    # El fixture no pasa curva, así que la hoja debe tener nota
    ws = wb[_SH_CURVA]
    has_nota = False
    for row in ws.iter_rows(max_row=15):
        for cell in row:
            v = str(cell.value or "")
            if "Sin datos" in v or "simulación" in v:
                has_nota = True
    assert has_nota, "Sin curva, la hoja debe mostrar nota informativa"


# ── Test 7: retorna bytes ─────────────────────────────────────────────────────

def test_retorna_bytes(resultado, params_dict, motores_config):
    buf = generar_excel_modelado_chp(
        params=params_dict,
        r=resultado,
        motores_config=motores_config,
        cliente_nombre="Test",
    )
    assert isinstance(buf, bytes) and len(buf) > 1000


# ── Test 8: sin meses — tabla mensual muestra mensaje ─────────────────────────

def test_sin_meses_tabla_mensaje(params_dict, motores_config):
    params = CoGenParams()
    r_vacio = CoGenResultado(
        params=params,
        meses=[],
        kwh_total_anual=Decimal("0"),
        kwh_cubiertos_anual=Decimal("0"),
        gj_gas_cogen_anual=Decimal("0"),
        costo_gas_cogen_anual_mxn=Decimal("0"),
        ahorro_electricidad_anual_mxn=Decimal("0"),
        ahorro_caldera_anual_mxn=Decimal("0"),
        ebitda_anual_mxn=Decimal("0"),
    )
    buf = generar_excel_modelado_chp(
        params=params_dict,
        r=r_vacio,
        motores_config=motores_config,
        cliente_nombre="Test",
    )
    wb = openpyxl.load_workbook(BytesIO(buf))
    ws = wb[_SH_MENSUAL]
    has_msg = any(
        "Sin datos" in str(cell.value or "") or "disponibles" in str(cell.value or "")
        for row in ws.iter_rows(max_row=20)
        for cell in row
    )
    assert has_msg, "Sin meses, la hoja debe mostrar mensaje informativo"
