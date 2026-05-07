# tests/reports/test_excel_formulas.py
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from models.cogen_result import CoGenMes, CoGenParams, CoGenResultado
from reports.excel import generar_excel, _FILA_DATOS, _FILA_HEADER


# ── Fixture ───────────────────────────────────────────────────────────────────

def _mes(year: int, month: int, **kwargs) -> CoGenMes:
    defaults = dict(
        periodo_inicio=date(year, month, 1),
        periodo_fin=date(year, month, 28),
        kwh_total=Decimal("1000"),
        costo_cfe_mxn=Decimal("2500"),
        costo_promedio_kwh=Decimal("2.5"),
        gj_consumido=Decimal("50"),
        costo_unitario_gj=Decimal("200"),
        costo_gas_actual_mxn=Decimal("10000"),
        kwh_cubiertos=Decimal("750"),
        gj_gas_cogen=Decimal("6.75"),
        costo_gas_cogen_mxn=Decimal("1350"),
        ahorro_electricidad_mxn=Decimal("1875"),
        calor_recuperado_gj=Decimal("1.6875"),
        ahorro_caldera_mxn=Decimal("397.06"),
        ebitda_mes_mxn=Decimal("922.06"),
        prorrateado=False,
        nota_prorrateo="",
    )
    defaults.update(kwargs)
    return CoGenMes(**defaults)


@pytest.fixture()
def resultado():
    params = CoGenParams(
        cobertura_electrica=Decimal("0.75"),
        rendimiento_electrico=Decimal("0.40"),
        rendimiento_termico=Decimal("0.25"),
        eficiencia_caldera=Decimal("0.85"),
    )
    meses = [
        _mes(2024, 1),
        _mes(2024, 2, prorrateado=True, nota_prorrateo="Prorrateado a 30 días"),
    ]
    return CoGenResultado(
        params=params,
        meses=meses,
        kwh_total_anual=Decimal("2000"),
        kwh_cubiertos_anual=Decimal("1500"),
        gj_gas_cogen_anual=Decimal("13.5"),
        costo_gas_cogen_anual_mxn=Decimal("2700"),
        ahorro_electricidad_anual_mxn=Decimal("3750"),
        ahorro_caldera_anual_mxn=Decimal("794.12"),
        ebitda_anual_mxn=Decimal("1844.12"),
    )


@pytest.fixture()
def wb_path(resultado, tmp_path):
    path = tmp_path / "test.xlsx"
    generar_excel(resultado, path)
    return path


@pytest.fixture()
def ws(wb_path):
    wb = openpyxl.load_workbook(wb_path)
    return wb["Análisis Mensual"]


# ── Test 1: celdas calculadas contienen fórmulas, no valores ─────────────────

def test_celdas_calculadas_contienen_formulas(ws):
    """Las columnas H–N en la primera fila de datos deben ser cadenas con '='."""
    # Primera fila de datos
    R = _FILA_DATOS
    columnas_formula = ["H", "I", "J", "K", "L", "M", "N"]
    for col in columnas_formula:
        valor = ws[f"{col}{R}"].value
        assert isinstance(valor, str) and valor.startswith("="), (
            f"Celda {col}{R}: se esperaba fórmula, se obtuvo {valor!r}"
        )


def test_celdas_fijas_contienen_valores(ws):
    """Las columnas B–G en la primera fila de datos deben ser valores numéricos."""
    R = _FILA_DATOS
    columnas_fijas = ["B", "C", "D", "E", "F", "G"]
    for col in columnas_fijas:
        valor = ws[f"{col}{R}"].value
        assert isinstance(valor, (int, float)), (
            f"Celda {col}{R}: se esperaba valor numérico, se obtuvo {valor!r}"
        )


# ── Test 2: fórmulas referencian las celdas correctas ────────────────────────

def test_formulas_primera_fila(ws):
    """Las fórmulas de la primera fila de datos (fila 9) son exactamente las esperadas."""
    R = _FILA_DATOS
    esperadas = {
        f"H{R}": f"=B{R}*$B$2",
        f"I{R}": f"=H{R}*$B$6/$B$3",
        f"J{R}": f"=I{R}*F{R}",
        f"K{R}": f"=H{R}*D{R}",
        f"L{R}": f"=I{R}*$B$4",
        f"M{R}": f"=(L{R}/$B$5)*F{R}",
        f"N{R}": f"=K{R}+M{R}-J{R}",
    }
    for celda, formula_esperada in esperadas.items():
        assert ws[celda].value == formula_esperada, (
            f"Celda {celda}: esperado {formula_esperada!r}, obtenido {ws[celda].value!r}"
        )


def test_formulas_segunda_fila_distinta_de_primera(ws):
    """Las fórmulas de la segunda fila referencian su propia fila, no la primera."""
    R1 = _FILA_DATOS
    R2 = _FILA_DATOS + 1
    assert ws[f"H{R2}"].value == f"=B{R2}*$B$2"
    assert ws[f"H{R2}"].value != ws[f"H{R1}"].value


def test_formulas_referencian_parametros_absolutos(ws):
    """Las referencias a parámetros usan $ para ser absolutas."""
    R = _FILA_DATOS
    formula_h = ws[f"H{R}"].value   # =B9*$B$2
    formula_i = ws[f"I{R}"].value   # =H9*$B$6/$B$3
    formula_l = ws[f"L{R}"].value   # =I9*$B$4
    formula_m = ws[f"M{R}"].value   # =(L9/$B$5)*F9

    assert "$B$2" in formula_h
    assert "$B$3" in formula_i
    assert "$B$6" in formula_i
    assert "$B$4" in formula_l
    assert "$B$5" in formula_m


# ── Test 3: fila de totales contiene fórmulas SUM ────────────────────────────

def test_totales_son_formulas_suma(ws, resultado):
    """La fila de totales contiene =SUM(...) en todas las columnas sumadas."""
    totales_row = _FILA_DATOS + len(resultado.meses)
    columnas_suma = ["B", "H", "I", "J", "K", "L", "M", "N"]
    for col in columnas_suma:
        valor = ws[f"{col}{totales_row}"].value
        assert isinstance(valor, str) and valor.startswith("=SUM("), (
            f"Total {col}{totales_row}: se esperaba =SUM(...), se obtuvo {valor!r}"
        )


def test_totales_suman_rango_correcto(ws, resultado):
    """El rango de la fórmula SUM cubre exactamente las filas de datos."""
    totales_row = _FILA_DATOS + len(resultado.meses)
    ultima_datos = totales_row - 1
    formula_ebitda = ws[f"N{totales_row}"].value
    rango_esperado = f"N{_FILA_DATOS}:N{ultima_datos}"
    assert rango_esperado in formula_ebitda, (
        f"Rango incorrecto en EBITDA total: {formula_ebitda!r}"
    )


# ── Test 4: valores de fórmulas coinciden con el cálculo Python ──────────────

def test_valores_formula_vs_python(ws, resultado):
    """Evalúa manualmente las fórmulas de cada mes y compara con los valores Python.

    No se usa data_only (openpyxl no cachea valores de fórmulas al escribir).
    En su lugar se leen las celdas de entrada y se evalúa la aritmética aquí.
    """
    params = resultado.params
    cobertura  = float(params.cobertura_electrica)
    rend_elec  = float(params.rendimiento_electrico)
    rend_term  = float(params.rendimiento_termico)
    ef_caldera = float(params.eficiencia_caldera)
    factor     = 0.0036

    for offset, mes in enumerate(resultado.meses):
        R = _FILA_DATOS + offset

        kwh_total    = ws[f"B{R}"].value
        cu_kwh       = ws[f"D{R}"].value
        cu_gj        = ws[f"F{R}"].value

        # Evaluar fórmulas
        kwh_cubiertos  = kwh_total * cobertura
        gj_cogen       = kwh_cubiertos * factor / rend_elec
        costo_gas_cogen= gj_cogen * cu_gj
        ahorro_elec    = kwh_cubiertos * cu_kwh
        calor_recup    = gj_cogen * rend_term
        ahorro_caldera = (calor_recup / ef_caldera) * cu_gj
        ebitda         = ahorro_elec + ahorro_caldera - costo_gas_cogen

        assert abs(kwh_cubiertos   - float(mes.kwh_cubiertos))        < 0.01, f"Mes {R}: kwh_cubiertos"
        assert abs(gj_cogen        - float(mes.gj_gas_cogen))          < 0.01, f"Mes {R}: gj_gas_cogen"
        assert abs(costo_gas_cogen - float(mes.costo_gas_cogen_mxn))   < 0.01, f"Mes {R}: costo_gas_cogen"
        assert abs(ahorro_elec     - float(mes.ahorro_electricidad_mxn)) < 0.01, f"Mes {R}: ahorro_elec"
        assert abs(calor_recup     - float(mes.calor_recuperado_gj))   < 0.01, f"Mes {R}: calor_recup"
        assert abs(ebitda          - float(mes.ebitda_mes_mxn))        < 0.5,  f"Mes {R}: ebitda"


# ── Test 5: bloque de parámetros está presente ───────────────────────────────

def test_bloque_parametros_presente(ws, resultado):
    """El bloque de parámetros en filas 1-6 contiene los valores correctos."""
    params = resultado.params
    assert ws["A1"].value == "Parámetros del motor candidato"
    assert abs(ws["B2"].value - float(params.cobertura_electrica))  < 1e-6
    assert abs(ws["B3"].value - float(params.rendimiento_electrico)) < 1e-6
    assert abs(ws["B4"].value - float(params.rendimiento_termico))   < 1e-6
    assert abs(ws["B5"].value - float(params.eficiencia_caldera))    < 1e-6
    assert abs(ws["B6"].value - 0.0036) < 1e-6
