# tests/reports/test_excel.py
from __future__ import annotations
import pytest
from decimal import Decimal
from datetime import date
from pathlib import Path
import tempfile

from models.cogen_result import CoGenParams, CoGenMes, CoGenResultado
from reports.excel import generar_excel


def _resultado_fixture() -> CoGenResultado:
    params = CoGenParams()
    meses = [
        CoGenMes(
            periodo_inicio=date(2023, 11, 1),
            periodo_fin=date(2023, 11, 30),
            kwh_total=Decimal("380800"),
            costo_cfe_mxn=Decimal("1369072.01"),
            costo_promedio_kwh=Decimal("3.60"),
            gj_consumido=Decimal("106445.18"),
            costo_unitario_gj=Decimal("79.48"),
            costo_gas_actual_mxn=Decimal("8460263.13"),
            kwh_cubiertos=Decimal("285600.00"),
            gj_gas_cogen=Decimal("2570.40"),
            costo_gas_cogen_mxn=Decimal("204254.59"),
            ahorro_electricidad_mxn=Decimal("1026804.01"),
            calor_recuperado_gj=Decimal("642.60"),
            ahorro_caldera_mxn=Decimal("60134.56"),
            ebitda_mes_mxn=Decimal("882683.98"),
        ),
        CoGenMes(
            periodo_inicio=date(2023, 12, 1),
            periodo_fin=date(2023, 12, 31),
            kwh_total=Decimal("616000"),
            costo_cfe_mxn=Decimal("1901763.84"),
            costo_promedio_kwh=Decimal("3.09"),
            gj_consumido=Decimal("98199.58"),
            costo_unitario_gj=Decimal("79.48"),
            costo_gas_actual_mxn=Decimal("7804062.04"),
            kwh_cubiertos=Decimal("462000.00"),
            gj_gas_cogen=Decimal("4158.00"),
            costo_gas_cogen_mxn=Decimal("330530.40"),
            ahorro_electricidad_mxn=Decimal("1426322.88"),
            calor_recuperado_gj=Decimal("1039.50"),
            ahorro_caldera_mxn=Decimal("97255.30"),
            ebitda_mes_mxn=Decimal("1193047.78"),
        ),
    ]
    return CoGenResultado(
        params=params,
        meses=meses,
        kwh_total_anual=sum(m.kwh_total for m in meses),
        kwh_cubiertos_anual=sum(m.kwh_cubiertos for m in meses),
        gj_gas_cogen_anual=sum(m.gj_gas_cogen for m in meses),
        costo_gas_cogen_anual_mxn=sum(m.costo_gas_cogen_mxn for m in meses),
        ahorro_electricidad_anual_mxn=sum(m.ahorro_electricidad_mxn for m in meses),
        ahorro_caldera_anual_mxn=sum(m.ahorro_caldera_mxn for m in meses),
        ebitda_anual_mxn=sum(m.ebitda_mes_mxn for m in meses),
    )


@pytest.fixture
def xlsx_path(tmp_path):
    resultado = _resultado_fixture()
    path = tmp_path / "test_analisis.xlsx"
    generar_excel(resultado, path)
    return path


def test_archivo_creado(xlsx_path):
    assert xlsx_path.exists()
    assert xlsx_path.stat().st_size > 0


def test_hojas_esperadas(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    assert "Análisis Mensual" in wb.sheetnames
    assert "Parámetros" in wb.sheetnames


def test_filas_de_datos(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Análisis Mensual"]
    # Fila 1 = encabezado, filas 2-3 = 2 meses, fila 4 = totales → mínimo 4 filas
    assert ws.max_row >= 4


def test_ebitda_anual_en_totales(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Análisis Mensual"]
    # Buscar celda con el EBITDA anual total en la última fila
    last_row = ws.max_row
    # Verificar que la última fila tiene algún valor numérico > 0
    valores = [ws.cell(last_row, c).value for c in range(1, ws.max_column + 1)]
    numericos = [v for v in valores if isinstance(v, (int, float)) and v > 0]
    assert len(numericos) > 0
