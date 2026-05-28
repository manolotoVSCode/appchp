# web/mediciones_parser.py
"""
Parser de archivos Excel de mediciones cincominutal.

Headers esperados (case-insensitive, strip): Fecha | kWh E
Conversión: potencia_kw = kWh_E × 12  (energía 5 min → potencia media en kW)
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path


def parse_cincominutal(filepath: Path) -> list[dict]:
    """Parsea un Excel de mediciones cincominutal y retorna lista de dicts.

    Returns:
        Lista de {"ts": datetime, "potencia_kw": float}

    Raises:
        ValueError: si no encuentra columnas requeridas, no hay datos válidos,
                    o hay menos de 100 filas válidas.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active

    # ── Detectar fila de headers ──────────────────────────────────────────────
    # Primera fila donde alguna celda tiene valor no vacío.
    header_row_idx = None
    col_fecha: int | None = None
    col_kwhe: int | None = None

    for row in ws.iter_rows():
        non_empty = [c for c in row if c.value is not None]
        if not non_empty:
            continue
        # Mapear nombre de columna (lower + strip) → índice 0-based
        mapping: dict[str, int] = {}
        for cell in row:
            if cell.value is not None:
                key = str(cell.value).strip().lower()
                mapping[key] = cell.column - 1  # 0-based

        fecha_idx = mapping.get("fecha")
        kwhe_idx  = mapping.get("kwh e")

        if fecha_idx is not None and kwhe_idx is not None:
            col_fecha = fecha_idx
            col_kwhe  = kwhe_idx
            header_row_idx = row[0].row
            break

    wb.close()

    if col_fecha is None or col_kwhe is None:
        raise ValueError(
            "No se encontraron las columnas 'Fecha' y 'kWh E' en el archivo. "
            "Verifica que los headers existan y estén escritos exactamente así."
        )

    # ── Segunda pasada: leer datos ────────────────────────────────────────────
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active

    datos: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Ignorar filas más cortas que los índices requeridos
        if len(row) <= max(col_fecha, col_kwhe):
            continue

        ts   = row[col_fecha]
        kwhe = row[col_kwhe]

        if not isinstance(ts, datetime):
            continue
        try:
            kwhe_float = float(kwhe)
        except (TypeError, ValueError):
            continue

        datos.append({
            "ts":          ts,
            "potencia_kw": round(kwhe_float * 12, 3),
        })

    wb.close()

    if not datos:
        raise ValueError(
            "El archivo no contiene filas de datos válidos. "
            "Asegúrate de que las columnas Fecha y kWh E tengan valores."
        )

    if len(datos) < 100:
        raise ValueError("Archivo con muy pocos datos")

    return datos
