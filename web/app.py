# web/app.py
from __future__ import annotations

import logging
import os
import sys
from dataclasses import asdict
from datetime import date, timedelta
from pathlib import Path
from time import time

from flask import Flask, abort, flash, redirect, render_template, request, send_file, session, url_for
from flask_wtf.csrf import CSRFProtect

from storage.repository import (
    get_facturas_para_dashboard,
    get_facturas_ppa_y_gas_para_dashboard,
    get_ultimas_cfe_invoices,
    get_ultimas_gas_invoices,
    get_ultimas_ppa_invoices,
    get_tipo_suministro_electrico_seleccionado,
    get_contratos_por_cliente,
    get_configuracion,
    get_configuracion_row,
    list_configuracion,
    set_configuracion,
    get_mediciones_por_cliente,
    get_cliente_chp_params,
)
from models.contrato import TIPO_ELECTRICO_CALIFICADO
from web.error_logger import log_error
from calc.cogen import calcular_cogen, calcular_cogen_ppa, calcular_cogen_precio_manual, calcular_payback_decimal, calcular_flujo_acumulado
from calc.historico import calcular_historico_cfe, calcular_tablas_cfe, calcular_historico_gas
from calc.nombre_canonico import generar_nombre_canonico
from calc.periodo import mes_asociado, UMBRAL_PRORRATEO_DIAS
from models.cogen_result import CoGenParams

logger = logging.getLogger(__name__)
csrf = CSRFProtect()

try:
    import re as _re
    _changelog = (Path(__file__).resolve().parent.parent / "CHANGELOG.md").read_text(encoding="utf-8")
    _m = _re.search(r"^## \[(\d+\.\d+\.\d+)\]", _changelog, _re.MULTILINE)
    _APP_VERSION = _m.group(1) if _m else ""
except Exception:
    _APP_VERSION = ""


_SECTOR_SVG_MAP = {
    "hotelero": "hotelero.svg",
    "manufactura": "manufactura.svg",
    "alimentos y bebidas": "alimentos-y-bebidas.svg",
    "químico": "quimico.svg",
    "textil": "textil.svg",
    "pesquero": "pesquero.svg",
    "forestal": "forestal.svg",
    "cerámico": "ceramico.svg",
    "plásticos": "plasticos.svg",
    "metalúrgico": "metalurgico.svg",
    "otro": "otro.svg",
}


def obtener_logo_cliente(cliente: dict) -> str:
    """Devuelve la URL del logo a mostrar para un cliente.

    Si el cliente tiene logo_url personalizado, lo devuelve.
    Si no, devuelve el logo SVG por defecto del sector.
    Si no hay sector o el sector no tiene SVG, devuelve 'otro.svg'.
    """
    from flask import url_for
    if cliente.get("logo_url"):
        return cliente["logo_url"]
    sector = (cliente.get("sector_industrial") or "").strip().lower()
    archivo = _SECTOR_SVG_MAP.get(sector, "otro.svg")
    return url_for("static", filename=f"img/sectores/{archivo}")


def _verificar_cliente_activo(cliente_id: int):
    """Verifica que cliente_id coincida con la sesión activa y exista en BD.

    Para usuario_normal: si el cliente_id está en su lista asignada,
    actualiza cliente_activo_id en sesión y continúa sin flash de error.

    Retorna (cliente_dict, None) si todo está bien.
    Retorna (None, response) si hay error; el caller debe retornar esa response.
    """
    from flask import flash
    from storage.repository import get_cliente_con_conteos
    from web.auth import get_current_user as _gcu

    activo_id = session.get("cliente_activo_id")
    if activo_id != cliente_id:
        current_user_data = _gcu()
        clientes_ids = session.get("_clientes_ids", [])
        if (current_user_data
                and current_user_data.get("rol") == "usuario_normal"
                and cliente_id in clientes_ids):
            session["cliente_activo_id"] = cliente_id
            session["_empresa_id"] = cliente_id
        else:
            flash(
                "El dashboard solicitado no corresponde al cliente activo. "
                "Selecciona el cliente desde el listado.",
                "warning",
            )
            return None, redirect(url_for("clientes.listado"))

    cliente = get_cliente_con_conteos(cliente_id)
    if cliente is None:
        session.pop("cliente_activo_id", None)
        session.pop("cliente_activo_nombre", None)
        flash("El cliente ya no existe.", "warning")
        return None, redirect(url_for("clientes.listado"))

    return cliente, None


_MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _formatear_meses(meses: list[tuple[int, int]]) -> str:
    """Formatea lista de (anio, mes) en string descriptivo en español."""
    if not meses:
        return ""
    ms = sorted(set(meses))
    if len(ms) == 1:
        a, m = ms[0]
        return f"{_MESES_ES[m]} {a}"
    indices = [a * 12 + m for a, m in ms]
    consecutive = all(indices[i + 1] == indices[i] + 1 for i in range(len(indices) - 1))
    anios = {a for a, _ in ms}
    if consecutive:
        a1, m1 = ms[0]
        a2, m2 = ms[-1]
        if a1 == a2:
            return f"{_MESES_ES[m1]} a {_MESES_ES[m2].lower()} {a1}"
        return f"{_MESES_ES[m1]} {a1} a {_MESES_ES[m2].lower()} {a2}"
    if len(anios) == 1:
        anio = next(iter(anios))
        nombres = [_MESES_ES[m].lower() for _, m in ms]
        nombres[0] = nombres[0].capitalize()
        return ", ".join(nombres) + f" {anio}"
    return ", ".join(f"{_MESES_ES[m]} {a}" for a, m in ms)


def _describir_periodo_contabilidad(cfe_invoices, gas_invoices, ppa_invoices, tipo_suministro) -> str:
    """Descripción detallada del período para el header del dashboard Contabilidad."""
    elec_meses: list[tuple[int, int]] = []
    if ppa_invoices:
        for inv in ppa_invoices:
            elec_meses.append((inv.anio, inv.mes))
        tipo_label = "[PPA Calificado]"
    else:
        for inv in cfe_invoices:
            elec_meses.append(mes_asociado(inv.periodo_inicio, inv.periodo_fin))
        tipo_label = "[CFE GDMTH]"
    gas_meses = [mes_asociado(inv.periodo_inicio, inv.periodo_fin) for inv in gas_invoices]
    partes = []
    if elec_meses:
        partes.append(f"{_formatear_meses(elec_meses)} {tipo_label}")
    if gas_meses:
        partes.append(f"{_formatear_meses(gas_meses)} [Gas]")
    return " + ".join(partes)


def _calcular_periodo_label(cfe_invoices, gas_invoices, ppa_invoices=None) -> str:
    """Retorna etiqueta del periodo cubierto por las facturas seleccionadas."""
    anios: set[int] = set()
    for inv in cfe_invoices:
        anio, _ = mes_asociado(inv.periodo_inicio, inv.periodo_fin)
        anios.add(anio)
    for inv in gas_invoices:
        anio, _ = mes_asociado(inv.periodo_inicio, inv.periodo_fin)
        anios.add(anio)
    if ppa_invoices:
        for inv in ppa_invoices:
            anios.add(inv.anio)
    if not anios:
        return ""
    ordenados = sorted(anios)
    if len(ordenados) == 1:
        return str(ordenados[0])
    if len(ordenados) == 2 and ordenados[1] == ordenados[0] + 1:
        return f"{ordenados[0]}–{ordenados[1]}"
    return "Múltiples años"


def _cargar_facturas_seleccionadas(cliente_id: int):
    """Carga facturas CFE y gas seleccionadas del cliente y las formatea para templates.

    Retorna (cfe_invoices, gas_invoices, facturas_cfe, facturas_gas).
    Usa get_facturas_para_dashboard para compartir la query de meses seleccionados (4 queries en total).
    """
    cfe_invoices, gas_invoices = get_facturas_para_dashboard(cliente_id)

    facturas_cfe = [
        {
            "nombre_canonico": generar_nombre_canonico(inv),
            "periodo": f"{inv.periodo_inicio.strftime('%d %b %Y')} – {inv.periodo_fin.strftime('%d %b %Y')}",
            "mes_asociado": date(*mes_asociado(inv.periodo_inicio, inv.periodo_fin), 1).strftime("%b %Y"),
            "kwh_total": float(sum(p.consumo_kwh for p in inv.periodos)),
            "costo_mxn": float(inv.subtotal_mxn),
            "prorrateado": (inv.periodo_fin - inv.periodo_inicio).days < UMBRAL_PRORRATEO_DIAS,
        }
        for inv in sorted(cfe_invoices, key=lambda x: x.periodo_inicio)
    ]

    facturas_gas = [
        {
            "nombre_canonico": generar_nombre_canonico(inv),
            "periodo": f"{inv.periodo_inicio.strftime('%d %b %Y')} – {inv.periodo_fin.strftime('%d %b %Y')}",
            "mes_asociado": date(*mes_asociado(inv.periodo_inicio, inv.periodo_fin), 1).strftime("%b %Y"),
            "gj_total": float(inv.consumo_total_gj),
            "costo_mxn": float(inv.subtotal_mxn),
            "prorrateado": (inv.periodo_fin - inv.periodo_inicio).days < UMBRAL_PRORRATEO_DIAS,
        }
        for inv in sorted(gas_invoices, key=lambda x: x.periodo_inicio)
    ]

    return cfe_invoices, gas_invoices, facturas_cfe, facturas_gas


def _cargar_facturas_ppa(cliente_id: int):
    """Carga facturas PPA y gas seleccionadas. Retorna (ppa_invoices, gas_invoices, facturas_ppa, facturas_gas)."""
    ppa_invoices, gas_invoices = get_facturas_ppa_y_gas_para_dashboard(cliente_id)

    facturas_ppa = [
        {
            "nombre_canonico": inv.nombre_canonico or f"CALIFICADO-{inv.anio}-{inv.mes:02d}",
            "periodo": f"{inv.periodo_inicio.strftime('%d %b %Y')} – {inv.periodo_fin.strftime('%d %b %Y')}",
            "mes_asociado": date(inv.anio, inv.mes, 1).strftime("%b %Y"),
            "kwh_total": float(inv.consumo_kwh),
            "costo_mxn": float(inv.subtotal_mxn),
            "precio_unitario_mxn_kwh": float(inv.precio_unitario_mxn_kwh),
            "suministrador": inv.suministrador or "",
        }
        for inv in sorted(ppa_invoices, key=lambda x: (x.anio, x.mes))
    ]

    facturas_gas = [
        {
            "nombre_canonico": generar_nombre_canonico(inv),
            "periodo": f"{inv.periodo_inicio.strftime('%d %b %Y')} – {inv.periodo_fin.strftime('%d %b %Y')}",
            "mes_asociado": date(*mes_asociado(inv.periodo_inicio, inv.periodo_fin), 1).strftime("%b %Y"),
            "gj_total": float(inv.consumo_total_gj),
            "costo_mxn": float(inv.subtotal_mxn),
            "prorrateado": (inv.periodo_fin - inv.periodo_inicio).days < UMBRAL_PRORRATEO_DIAS,
        }
        for inv in sorted(gas_invoices, key=lambda x: x.periodo_inicio)
    ]

    return ppa_invoices, gas_invoices, facturas_ppa, facturas_gas


_MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}


def _calcular_rango_cogen(cfe_invoices, gas_invoices, ppa_invoices=None) -> str:
    """Rango de meses usados en el cálculo de cogeneración.

    Formato: 'Cálculo basado en los últimos 12 meses: junio 2024 a mayo 2025'
    """
    meses: list[date] = []
    if ppa_invoices:
        for inv in ppa_invoices:
            meses.append(date(inv.anio, inv.mes, 1))
    else:
        for inv in cfe_invoices:
            a, m = mes_asociado(inv.periodo_inicio, inv.periodo_fin)
            meses.append(date(a, m, 1))
    for inv in gas_invoices:
        a, m = mes_asociado(inv.periodo_inicio, inv.periodo_fin)
        meses.append(date(a, m, 1))
    if not meses:
        return ""
    min_mes = min(meses)
    max_mes = max(meses)
    if min_mes == max_mes:
        return f"Cálculo basado en los últimos 12 meses: {_MESES_ES[min_mes.month]} {min_mes.year}"
    return (
        f"Cálculo basado en los últimos 12 meses: "
        f"{_MESES_ES[min_mes.month]} {min_mes.year} a "
        f"{_MESES_ES[max_mes.month]} {max_mes.year}"
    )


def _cargar_ultimas_facturas_cogen(cliente_id: int):
    """Carga las últimas 12 facturas CFE y gas para el dashboard de Cogeneración.

    Ignora la selección manual del sidebar; siempre usa las facturas más recientes.
    Retorna (cfe_invoices, gas_invoices, facturas_cfe, facturas_gas).
    """
    cfe_invoices = sorted(get_ultimas_cfe_invoices(cliente_id, n=12), key=lambda x: x.periodo_inicio)
    gas_invoices = sorted(get_ultimas_gas_invoices(cliente_id, n=12), key=lambda x: x.periodo_inicio)

    facturas_cfe = [
        {
            "nombre_canonico": generar_nombre_canonico(inv),
            "periodo": f"{inv.periodo_inicio.strftime('%d %b %Y')} – {inv.periodo_fin.strftime('%d %b %Y')}",
            "mes_asociado": date(*mes_asociado(inv.periodo_inicio, inv.periodo_fin), 1).strftime("%b %Y"),
            "kwh_total": float(sum(p.consumo_kwh for p in inv.periodos)),
            "costo_mxn": float(inv.subtotal_mxn),
            "prorrateado": (inv.periodo_fin - inv.periodo_inicio).days < UMBRAL_PRORRATEO_DIAS,
        }
        for inv in cfe_invoices
    ]

    facturas_gas = [
        {
            "nombre_canonico": generar_nombre_canonico(inv),
            "periodo": f"{inv.periodo_inicio.strftime('%d %b %Y')} – {inv.periodo_fin.strftime('%d %b %Y')}",
            "mes_asociado": date(*mes_asociado(inv.periodo_inicio, inv.periodo_fin), 1).strftime("%b %Y"),
            "gj_total": float(inv.consumo_total_gj),
            "costo_mxn": float(inv.subtotal_mxn),
            "prorrateado": (inv.periodo_fin - inv.periodo_inicio).days < UMBRAL_PRORRATEO_DIAS,
        }
        for inv in gas_invoices
    ]

    return cfe_invoices, gas_invoices, facturas_cfe, facturas_gas


def _cargar_ultimas_ppa_cogen(cliente_id: int):
    """Carga las últimas 12 facturas PPA y gas para el dashboard de Cogeneración.

    Ignora la selección manual del sidebar; siempre usa las facturas más recientes.
    Retorna (ppa_invoices, gas_invoices, facturas_ppa, facturas_gas).
    """
    ppa_invoices = sorted(get_ultimas_ppa_invoices(cliente_id, n=12), key=lambda x: (x.anio, x.mes))
    gas_invoices = sorted(get_ultimas_gas_invoices(cliente_id, n=12), key=lambda x: x.periodo_inicio)

    facturas_ppa = [
        {
            "nombre_canonico": inv.nombre_canonico or f"CALIFICADO-{inv.anio}-{inv.mes:02d}",
            "periodo": f"{inv.periodo_inicio.strftime('%d %b %Y')} – {inv.periodo_fin.strftime('%d %b %Y')}",
            "mes_asociado": date(inv.anio, inv.mes, 1).strftime("%b %Y"),
            "kwh_total": float(inv.consumo_kwh),
            "costo_mxn": float(inv.subtotal_mxn),
            "precio_unitario_mxn_kwh": float(inv.precio_unitario_mxn_kwh),
            "suministrador": inv.suministrador or "",
        }
        for inv in ppa_invoices
    ]

    facturas_gas = [
        {
            "nombre_canonico": generar_nombre_canonico(inv),
            "periodo": f"{inv.periodo_inicio.strftime('%d %b %Y')} – {inv.periodo_fin.strftime('%d %b %Y')}",
            "mes_asociado": date(*mes_asociado(inv.periodo_inicio, inv.periodo_fin), 1).strftime("%b %Y"),
            "gj_total": float(inv.consumo_total_gj),
            "costo_mxn": float(inv.subtotal_mxn),
            "prorrateado": (inv.periodo_fin - inv.periodo_inicio).days < UMBRAL_PRORRATEO_DIAS,
        }
        for inv in gas_invoices
    ]

    return ppa_invoices, gas_invoices, facturas_ppa, facturas_gas


def _serial(obj):
    """Convierte Decimal/date recursivamente a tipos JSON-safe."""
    from decimal import Decimal
    from datetime import date as _date
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, _date):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {k: _serial(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_serial(v) for v in obj]
    return obj


def _calcular_queso(tablas: dict) -> dict | None:
    """Calcula datos para gráfica de composición de costo (pie chart) a partir de tablas CFE."""
    filas_mes = [f for f in tablas.get("costos_detallados", []) if f.get("mes") != "ANUAL"]
    if not filas_mes:
        return None
    tot_e = sum(f["ce_total"] for f in filas_mes)
    tot_d = sum(f["costo_dem"] for f in filas_mes)
    tot_s = sum(f["subtotal"] for f in filas_mes)
    return {
        "agregado": {
            "energia": tot_e,
            "demanda": tot_d,
            "otros": max(0.0, tot_s - tot_e - tot_d),
            "total": tot_s,
        },
        "por_mes": [
            {
                "label": f["mes"],
                "energia": f["ce_total"],
                "demanda": f["costo_dem"],
                "otros": max(0.0, f["subtotal"] - f["ce_total"] - f["costo_dem"]),
                "total": f["subtotal"],
            }
            for f in filas_mes
        ],
    }


def _cels_to_dict(cels) -> dict | None:
    """Convierte CELsResultado a dict JSON-safe. None si cels es None."""
    if cels is None:
        return None
    return _serial({
        "es_eficiente": cels.es_eficiente,
        "cels_mwh_anual": cels.cels_mwh_anual,
        "capacidad_kw": cels.capacidad_kw,
        "capacidad_es_estimada": cels.capacidad_es_estimada,
        "medio_termico": cels.medio_termico,
        "nivel_tension_kv": cels.nivel_tension_kv,
        "altitud_msnm": cels.altitud_msnm,
        "tipo_motor": cels.tipo_motor,
        "E_mwh": cels.E_mwh,
        "F_mwh": cels.F_mwh,
        "H_mwh": cels.H_mwh,
        "RefE": cels.RefE,
        "RefH": cels.RefH,
        "fp": cels.fp,
        "RefE_prima": cels.RefE_prima,
        "Fh": cels.Fh,
        "Fe": cels.Fe,
        "EE": cels.EE,
        "EP": cels.EP,
        "AEP": cels.AEP,
        "APEP": cels.APEP,
        "AREL": cels.AREL,
        "ELC": cels.ELC,
        "porcentaje_ELC": cels.porcentaje_ELC,
    })


def create_app() -> Flask:
    """Flask app factory."""

    logging.basicConfig(
        stream=sys.stdout,
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        force=True,
    )

    app = Flask(__name__)

    secret = os.environ.get("SECRET_KEY")
    if not secret:
        raise RuntimeError(
            "SECRET_KEY no configurada en variables de entorno. "
            "Genera una con: python -c \"import secrets; print(secrets.token_hex(32))\" "
            "y configúrala en Render Dashboard → Environment."
        )
    app.config["SECRET_KEY"] = secret
    app.config["FASE2_HABILITADA"] = os.getenv("FASE2_HABILITADA", "false").lower() == "true"

    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SECURE"] = not app.debug
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

    # Autenticación, CSRF y blueprints
    from web.auth import init_auth, get_current_user, is_authenticated, clear_user_session, _verificar_activo_con_cache, _verificar_session_version_con_cache
    from web.clientes import clientes_bp
    init_auth(app)
    csrf.init_app(app)
    app.register_blueprint(clientes_bp)

    @app.template_filter("label_rol")
    def _label_rol(rol: str) -> str:
        return {"master_admin": "Super Admin", "admin": "Administrador", "usuario_normal": "Cliente"}.get(rol, rol)

    @app.template_filter("abreviar_con_cliente")
    def _abreviar_con_cliente(nombre_contrato: str, nombre_cliente: str) -> str:
        """Reemplaza el prefijo del nombre del cliente por sus iniciales.

        'IBÉRICA TILES Planta 1', 'IBÉRICA TILES' → 'IT Planta 1'
        Artículos ignorados: de, del, la, los, las, y.
        Comparación sin acentos ni distinción de mayúsculas (NFKD).
        """
        import unicodedata

        ARTICULOS = {"de", "del", "la", "los", "las", "y"}
        iniciales = "".join(
            p[0].upper() for p in nombre_cliente.split()
            if p.lower() not in ARTICULOS and p
        )

        def _norm(s: str) -> str:
            return "".join(
                c for c in unicodedata.normalize("NFKD", s.lower())
                if unicodedata.category(c) != "Mn"
            )

        if _norm(nombre_contrato).startswith(_norm(nombre_cliente)):
            return iniciales + nombre_contrato[len(nombre_cliente):]
        return nombre_contrato

    # Rutas exentas de autenticación
    _PUBLIC_PREFIXES = ("/auth/", "/static")
    _PUBLIC_EXACT = {"/healthz", "/health", "/privacidad"}

    @app.before_request
    def _require_login():
        path = request.path
        if path in _PUBLIC_EXACT:
            return None
        for prefix in _PUBLIC_PREFIXES:
            if path.startswith(prefix):
                return None
        if not is_authenticated():
            return redirect(url_for("auth.login", next=path))
        user = get_current_user()
        # Verificar que la cuenta sigue activa (cache de 5 minutos en sesión)
        if user and not _verificar_activo_con_cache(user["user_id"]):
            clear_user_session()
            flash("Tu cuenta ha sido desactivada. Contacta con tu administrador.", "warning")
            return redirect(url_for("auth.login"))
        if user and not _verificar_session_version_con_cache(user["user_id"]):
            clear_user_session()
            flash("Tu sesión ha sido invalidada por seguridad. Por favor inicia sesión de nuevo.", "warning")
            return redirect(url_for("auth.login"))
        # usuario_normal solo puede ver su empresa asignada
        if user and user["rol"] == "usuario_normal":
            from web.auth_permissions import usuario_puede_ver_empresa
            empresa_id = user.get("empresa_id")
            # Bloquear acceso a rutas de clientes que no sean su empresa
            import re as _re
            m = _re.match(r"^/clientes/(\d+)", path)
            if m:
                cid = int(m.group(1))
                if not usuario_puede_ver_empresa(cid, user):
                    flash("No tienes acceso a ese cliente.", "danger")
                    return redirect(url_for("clientes.listado"))

    @app.context_processor
    def _inject_globals():
        from storage.repository import get_cliente_con_conteos as _get_cliente
        from storage.repository import get_mediciones_por_cliente as _get_mediciones
        current_user_data = get_current_user()
        id_ = session.get("cliente_activo_id")
        base = {
            "current_user_data": current_user_data,
            "app_version": _APP_VERSION,
        }
        try:
            base["mediciones_sidebar"] = _get_mediciones(id_) if id_ else []
        except Exception:
            base["mediciones_sidebar"] = []
        # Inyectar lista de clientes para usuario_normal (sidebar dinámico)
        if current_user_data and current_user_data.get("rol") == "usuario_normal":
            from storage.repository import get_clientes_de_usuario as _gcdu
            uid = current_user_data.get("user_id")
            try:
                base["clientes_usuario"] = _gcdu(uid) if uid else []
            except Exception:
                base["clientes_usuario"] = []
        else:
            base["clientes_usuario"] = []
        # Nombre de empresa para usuario_normal: prefijo común de los clientes asignados
        _clst = base["clientes_usuario"]
        if len(_clst) > 1:
            _nombres = [c["nombre"] for c in _clst]
            _prefix = _nombres[0]
            for _n in _nombres[1:]:
                _i = 0
                while _i < len(_prefix) and _i < len(_n) and _prefix[_i] == _n[_i]:
                    _i += 1
                _prefix = _prefix[:_i]
            _prefix = _prefix.rstrip()
            _sp = _prefix.rfind(" ")
            base["nombre_empresa_usuario"] = _prefix[:_sp].strip() if _sp != -1 else _prefix
        elif _clst:
            base["nombre_empresa_usuario"] = _clst[0]["nombre"]
        else:
            base["nombre_empresa_usuario"] = ""
        if not id_:
            return {**base, "cliente_activo": None}

        # Usar valor cacheado si es fresco (TTL 60s) y corresponde al mismo cliente
        cached = session.get("_cp_cache")
        if (cached and cached.get("id") == id_
                and time() - cached.get("ts", 0) < 60):
            return {**base, "cliente_activo": cached["data"]}

        # Cache miss: consultar BD
        cliente = _get_cliente(id_)
        if cliente is None:
            session.pop("cliente_activo_id", None)
            session.pop("cliente_activo_nombre", None)
            session.pop("cliente_activo_logo_url", None)
            session.pop("_cp_cache", None)
            return {**base, "cliente_activo": None}
        contratos = [asdict(c) for c in get_contratos_por_cliente(id_)]
        data = {
            "id": id_,
            "nombre": cliente["nombre"],
            "contratos": contratos,
            "logo_url": cliente.get("logo_url"),
        }
        session["_cp_cache"] = {"id": id_, "ts": time(), "data": data}
        return {**base, "cliente_activo": data}

    @app.context_processor
    def _inject_logo_helper():
        return {"obtener_logo_cliente": obtener_logo_cliente}

    @app.context_processor
    def inject_fase2_flag():
        return {"fase2_habilitada": app.config["FASE2_HABILITADA"]}

    @app.route("/")
    def dashboard():
        """Redirige al listado de clientes."""
        return redirect(url_for("clientes.listado"))

    @app.route("/clientes/<int:cliente_id>/dashboard")
    def cliente_dashboard(cliente_id: int):
        """Redirige a Contabilidad Energética (primera vista del análisis del cliente)."""
        return redirect(url_for("cliente_dashboard_contabilidad", cliente_id=cliente_id))

    @app.route("/clientes/<int:cliente_id>/dashboard/contabilidad")
    def cliente_dashboard_contabilidad(cliente_id: int):
        """Vista de Contabilidad Energética: histórico eléctrico del cliente."""
        cliente, err = _verificar_cliente_activo(cliente_id)
        if err:
            return err

        tipo_suministro = get_tipo_suministro_electrico_seleccionado(cliente_id)

        try:
            if tipo_suministro == TIPO_ELECTRICO_CALIFICADO:
                ppa_invoices, gas_invoices, facturas_ppa, facturas_gas = _cargar_facturas_ppa(cliente_id)
                historico = None
                tablas = {}
                queso = None
                historico_gas = calcular_historico_gas(gas_invoices)
                kwh_total_periodo = sum(f["kwh_total"] for f in facturas_ppa)
                costo_total_periodo = sum(f["costo_mxn"] for f in facturas_ppa)
                costo_unit_promedio = costo_total_periodo / kwh_total_periodo if kwh_total_periodo > 0 else 0.0
                periodo_label = _calcular_periodo_label([], gas_invoices, ppa_invoices)
                suministrador_ppa = facturas_ppa[0]["suministrador"] if facturas_ppa else ""
                num_elec_sel = len(facturas_ppa)
                num_gas_sel = len(facturas_gas)
                facturas_cfe_tmpl = facturas_ppa
            else:
                cfe_invoices, gas_invoices, facturas_cfe, facturas_gas = _cargar_facturas_seleccionadas(cliente_id)
                historico = calcular_historico_cfe(cfe_invoices)
                tablas = calcular_tablas_cfe(cfe_invoices)
                queso = _calcular_queso(tablas)
                historico_gas = calcular_historico_gas(gas_invoices)
                kwh_total_periodo = sum(f["kwh_total"] for f in facturas_cfe)
                costo_total_periodo = sum(f["costo_mxn"] for f in facturas_cfe)
                costo_unit_promedio = costo_total_periodo / kwh_total_periodo if kwh_total_periodo > 0 else 0.0
                periodo_label = _calcular_periodo_label(cfe_invoices, gas_invoices)
                ppa_invoices = []
                facturas_ppa = []
                suministrador_ppa = ""
                num_elec_sel = len(facturas_cfe)
                num_gas_sel = len(facturas_gas)
                facturas_cfe_tmpl = facturas_cfe
        except Exception as e:
            import traceback
            traceback.print_exc()
            return (
                "<html><head><title>Error</title></head>"
                "<body style='font-family:sans-serif;padding:2rem'>"
                f"<h2>&#9888; Error al cargar datos</h2>"
                f"<pre>{e}</pre>"
                "</body></html>",
                500,
            )

        if tipo_suministro == TIPO_ELECTRICO_CALIFICADO:
            if cliente["num_electricidad"] == 0 and cliente["num_gas"] == 0:
                aviso_datos = {"tipo": "sin_facturas", "num_cfe": 0, "num_gas": 0, "cliente_id": cliente_id}
            elif num_elec_sel == 0:
                aviso_datos = {"tipo": "sin_seleccion", "cliente_id": cliente_id}
            else:
                aviso_datos = None
        else:
            num_cfe_total = cliente["num_cfe"]
            num_gas_total = cliente["num_gas"]
            if num_cfe_total == 0 and num_gas_total == 0:
                aviso_datos = {"tipo": "sin_facturas", "num_cfe": 0, "num_gas": 0, "cliente_id": cliente_id}
            elif num_elec_sel == 0 and num_gas_sel == 0:
                aviso_datos = {"tipo": "sin_seleccion", "cliente_id": cliente_id}
            elif num_elec_sel == 0 or num_gas_sel == 0:
                aviso_datos = {"tipo": "sin_par", "num_cfe": num_elec_sel, "num_gas": num_gas_sel}
            else:
                aviso_datos = None

        mediciones_cont = get_mediciones_por_cliente(cliente_id)
        if mediciones_cont and not session.get("medicion_activa_id"):
            session["medicion_activa_id"] = mediciones_cont[0]["id"]

        return render_template(
            "dashboard_contabilidad.html",
            aviso_datos=aviso_datos,
            cliente_id=cliente_id,
            cliente_nombre=cliente["nombre"],
            logo_url=obtener_logo_cliente(cliente),
            facturas_cfe=facturas_cfe_tmpl,
            facturas_gas=facturas_gas,
            historico=historico,
            tablas=tablas,
            historico_gas=historico_gas,
            queso=queso,
            num_meses_analizados=num_elec_sel,
            kwh_total_periodo=kwh_total_periodo,
            costo_total_periodo=costo_total_periodo,
            costo_unit_promedio=costo_unit_promedio,
            periodo_label=periodo_label,
            tipo_suministro_electrico=tipo_suministro,
            suministrador_ppa=suministrador_ppa,
            facturas_ppa=facturas_ppa,
        )

    @app.route("/clientes/<int:cliente_id>/dashboard/cogeneracion")
    def cliente_dashboard_cogeneracion(cliente_id: int):
        """Vista de Proyecto Cogeneración: análisis de oportunidad de cogeneración."""
        cliente, err = _verificar_cliente_activo(cliente_id)
        if err:
            return err

        tipo_suministro = get_tipo_suministro_electrico_seleccionado(cliente_id)

        try:
            from decimal import Decimal as _D
            _cfg = {r["clave"]: r["valor"] for r in list_configuracion()}
            tc_str = _cfg.get("tipo_cambio_mxn_usd")
            tipo_cambio = _D(tc_str) if tc_str else _D("17.50")
            fe_elec_str = _cfg.get("factor_emision_electricidad_kg_co2_kwh")
            fe_gas_str  = _cfg.get("factor_emision_gas_kg_co2_gj")
            fe_elec = _D(fe_elec_str) if fe_elec_str else None
            fe_gas  = _D(fe_gas_str)  if fe_gas_str  else None

            if tipo_suministro == TIPO_ELECTRICO_CALIFICADO:
                ppa_invoices, gas_invoices, facturas_ppa, facturas_gas = _cargar_ultimas_ppa_cogen(cliente_id)
                r = calcular_cogen_ppa(
                    ppa_invoices, gas_invoices, CoGenParams(),
                    tipo_cambio=tipo_cambio,
                    factor_emision_elec=fe_elec,
                    factor_emision_gas=fe_gas,
                )
                cfe_invoices = []
                facturas_cfe = []
                precio_gas_fuente = "ppa"
            else:
                cfe_invoices, gas_invoices, facturas_cfe, facturas_gas = _cargar_ultimas_facturas_cogen(cliente_id)
                ppa_invoices = []
                facturas_ppa = []
                _precio_manual_str = cliente.get("precio_gas_manual_mxn_gj_pcs")
                _precio_manual = _D(_precio_manual_str) if _precio_manual_str else None
                if len(cfe_invoices) >= 12 and len(gas_invoices) < 12 and _precio_manual:
                    r = calcular_cogen_precio_manual(
                        cfe_invoices, _precio_manual, CoGenParams(),
                        tipo_cambio=tipo_cambio,
                        factor_emision_elec=fe_elec,
                        factor_emision_gas=fe_gas,
                    )
                    precio_gas_fuente = "manual"
                else:
                    r = calcular_cogen(
                        cfe_invoices, gas_invoices, CoGenParams(),
                        tipo_cambio=tipo_cambio,
                        factor_emision_elec=fe_elec,
                        factor_emision_gas=fe_gas,
                    )
                    precio_gas_fuente = "real"
        except Exception as e:
            import traceback
            traceback.print_exc()
            return (
                "<html><head><title>Error</title></head>"
                "<body style='font-family:sans-serif;padding:2rem'>"
                f"<h2>&#9888; Error al cargar datos</h2>"
                f"<pre>{e}</pre>"
                "</body></html>",
                500,
            )

        # ── CELs (aplica igual para GDMTH y PPA — fórmula CRE Caso I) ─────────────
        cels_resultado = None
        try:
            from calc.cels import calcular_cels as _calcular_cels
            calor_recuperado_anual = sum(m.calor_recuperado_gj for m in r.meses)
            cels_resultado = _calcular_cels(
                kwh_cubiertos_anual=r.kwh_cubiertos_anual,
                gj_gas_cogen_pci_anual=r.gj_gas_cogen_pci_anual,
                calor_recuperado_gj_anual=calor_recuperado_anual,
                capacidad_nominal_kw=r.capacidad_nominal_kw,
                medio_termico=cliente.get("medio_termico"),
                nivel_tension_kv=cliente.get("nivel_tension_kv"),
                altitud_msnm=cliente.get("altitud_msnm"),
                tipo_motor=cliente.get("tipo_motor"),
                medio_termico_vapor_pct=cliente.get("medio_termico_vapor_pct"),
            )
        except Exception as _e_cels:
            import logging as _logging
            _logging.getLogger(__name__).error("Error calculando CELs: %s", _e_cels)

        num_cfe_total = cliente["num_cfe"]
        num_gas_total = cliente["num_gas"]
        num_elec_sel = len(facturas_ppa) if tipo_suministro == TIPO_ELECTRICO_CALIFICADO else len(facturas_cfe)
        num_gas_sel = len(facturas_gas)

        # Validación 12 meses
        if tipo_suministro != TIPO_ELECTRICO_CALIFICADO:
            _precio_manual_str = cliente.get("precio_gas_manual_mxn_gj_pcs")
            _precio_manual = _D(_precio_manual_str) if _precio_manual_str else None
            if num_elec_sel < 12:
                aviso_datos = {"tipo": "insuficiente_elec", "n_elec": num_elec_sel, "cliente_id": cliente_id}
            elif num_gas_sel < 12 and not _precio_manual:
                aviso_datos = {"tipo": "insuficiente_gas", "n_gas": num_gas_sel, "cliente_id": cliente_id}
            elif len(r.meses) < 12:
                aviso_datos = {"tipo": "meses_no_coinciden", "n_pares": len(r.meses),
                               "n_elec": num_elec_sel, "n_gas": num_gas_sel}
            else:
                aviso_datos = None
        elif num_cfe_total == 0 and num_gas_total == 0:
            aviso_datos = {"tipo": "sin_facturas", "num_cfe": 0, "num_gas": 0, "cliente_id": cliente_id}
        elif num_elec_sel == 0 and num_gas_sel == 0:
            aviso_datos = {"tipo": "sin_seleccion", "cliente_id": cliente_id}
        elif num_elec_sel == 0 or num_gas_sel == 0:
            aviso_datos = {"tipo": "sin_par", "num_cfe": num_elec_sel, "num_gas": num_gas_sel}
        elif not r.meses:
            aviso_datos = {"tipo": "sin_pares_mes", "num_cfe": num_elec_sel, "num_gas": num_gas_sel}
        else:
            aviso_datos = None

        rango_cogen = _calcular_rango_cogen(cfe_invoices, gas_invoices, ppa_invoices if ppa_invoices else None)

        chart_labels = [m.periodo_inicio.strftime("%b %Y") for m in r.meses]
        chart_ebitda = [float(m.ebitda_mes_mxn) for m in r.meses]
        chart_ahorro_elec = [float(m.ahorro_electricidad_mxn) for m in r.meses]
        chart_ahorro_caldera = [float(m.ahorro_caldera_mxn) for m in r.meses]
        chart_costo_gas = [float(m.costo_gas_cogen_mxn) for m in r.meses]
        chart_om = [float(m.gasto_om_mes_mxn) for m in r.meses]
        meses_raw = [
            {
                "periodo": m.periodo_inicio.strftime("%b %Y"),
                "kwh_total": float(m.kwh_total),
                "costo_cfe_mxn": float(m.costo_cfe_mxn),
                "costo_promedio_kwh": float(m.costo_promedio_kwh),
                "gj_consumido": float(m.gj_consumido),
                "costo_unitario_gj": float(m.costo_unitario_gj),
                "costo_gas_actual_mxn": float(m.costo_gas_actual_mxn),
                "kwh_punta": float(m.kwh_punta_total),
                "kwh_intermedia": float(m.kwh_intermedia_total),
                "kwh_base": float(m.kwh_base_total),
                "cu_punta": float(m.cu_punta_kwh),
                "cu_intermedia": float(m.cu_intermedia_kwh),
                "cu_base": float(m.cu_base_kwh),
                "kwh_punta_cubierto": float(m.kwh_punta_cubierto),
                "kwh_intermedia_cubierto": float(m.kwh_intermedia_cubierto),
                "kwh_base_cubierto": float(m.kwh_base_cubierto),
                "kw_max": float(m.kw_max),
                "kw_punta": float(m.kw_punta_orig),
                "dias_mes": m.dias_facturados,
                "kwh_total_orig": float(m.kwh_total_orig),
                "precio_capacidad_kw": float(m.precio_capacidad_kw),
                "precio_distribucion_kw": float(m.precio_distribucion_kw),
                "kw_facturado_capacidad": float(m.kw_facturado_capacidad),
                "kw_facturado_distribucion": float(m.kw_facturado_distribucion),
                "precio_otros_kwh": float(m.precio_otros_mxn_kwh),
            }
            for m in r.meses
        ]

        # Payback y flujo a 15 años (solo si hay inversión calculable)
        if r.inversion_mxn is not None and r.inversion_mxn > 0:
            payback_inicial = calcular_payback_decimal(r.inversion_mxn, r.ebitda_anual_mxn, r.ebitda_anual_mxn)
            flujo_acum_15 = [float(v) for v in calcular_flujo_acumulado(r.inversion_mxn, r.ebitda_anual_mxn)]
            flujo_anual_15 = [-float(r.inversion_mxn)] + [float(r.ebitda_anual_mxn)] * 15
        else:
            payback_inicial = None
            flujo_acum_15 = []
            flujo_anual_15 = []

        periodo_label = _calcular_periodo_label(cfe_invoices, gas_invoices, ppa_invoices if ppa_invoices else None)
        suministrador_ppa = facturas_ppa[0]["suministrador"] if facturas_ppa else ""

        mediciones_cogen = get_mediciones_por_cliente(cliente_id)
        if mediciones_cogen and not session.get("medicion_activa_id"):
            session["medicion_activa_id"] = mediciones_cogen[0]["id"]

        return render_template(
            "dashboard_cogeneracion.html",
            r=r,
            aviso_datos=aviso_datos,
            cliente_id=cliente_id,
            cliente_nombre=cliente["nombre"],
            logo_url=obtener_logo_cliente(cliente),
            periodo_label=periodo_label,
            rango_cogen=rango_cogen,
            chart_labels=chart_labels,
            chart_ebitda=chart_ebitda,
            chart_ahorro_elec=chart_ahorro_elec,
            chart_ahorro_caldera=chart_ahorro_caldera,
            chart_costo_gas=chart_costo_gas,
            chart_om=chart_om,
            meses_raw=meses_raw,
            payback_inicial=payback_inicial,
            flujo_acum_15=flujo_acum_15,
            flujo_anual_15=flujo_anual_15,
            factor_emision_elec=float(fe_elec) if fe_elec is not None else None,
            factor_emision_gas=float(fe_gas)   if fe_gas  is not None else None,
            cels=cels_resultado,
            cliente_ficha_url=url_for("clientes.ficha", cliente_id=cliente_id),
            tipo_suministro_electrico=tipo_suministro,
            suministrador_ppa=suministrador_ppa,
            precio_gas_fuente=precio_gas_fuente,
        )

    # ── Endpoints JSON para dashboards (client-side rendering) ─────────────────

    @app.route("/clientes/<int:cliente_id>/dashboard/contabilidad/data")
    def cliente_dashboard_contabilidad_data(cliente_id: int):
        """JSON con todos los datos del dashboard de Contabilidad Energética."""
        from flask import jsonify
        activo_id = session.get("cliente_activo_id")
        if activo_id != cliente_id:
            return jsonify({"error": "no_autorizado"}), 403
        from storage.repository import get_cliente_con_conteos as _gcc
        cliente = _gcc(cliente_id)
        if cliente is None:
            return jsonify({"error": "no_encontrado"}), 404

        tipo_suministro = get_tipo_suministro_electrico_seleccionado(cliente_id)

        try:
            if tipo_suministro == TIPO_ELECTRICO_CALIFICADO:
                ppa_invoices, gas_invoices, facturas_ppa, facturas_gas = _cargar_facturas_ppa(cliente_id)
                historico = None
                tablas = {}
                queso = None
                historico_gas = calcular_historico_gas(gas_invoices)
                kwh_total = sum(f["kwh_total"] for f in facturas_ppa)
                costo_total = sum(f["costo_mxn"] for f in facturas_ppa)
                costo_unit = costo_total / kwh_total if kwh_total > 0 else 0.0
                periodo_label = _describir_periodo_contabilidad([], gas_invoices, ppa_invoices, tipo_suministro)
                num_elec_sel = len(facturas_ppa)
                num_gas_sel = len(facturas_gas)
                historico_ppa = [
                    {
                        "mes": f["mes_asociado"],
                        "kwh_total": f["kwh_total"],
                        "precio_unitario_mxn_kwh": f["precio_unitario_mxn_kwh"],
                        "costo_mxn": f["costo_mxn"],
                        "suministrador": f["suministrador"],
                    }
                    for f in facturas_ppa
                ]
                suministrador_ppa = facturas_ppa[0]["suministrador"] if facturas_ppa else ""
                facturas_elec = facturas_ppa
            else:
                cfe_invoices, gas_invoices, facturas_cfe, facturas_gas = _cargar_facturas_seleccionadas(cliente_id)
                historico = calcular_historico_cfe(cfe_invoices)
                tablas = calcular_tablas_cfe(cfe_invoices)
                historico_gas = calcular_historico_gas(gas_invoices)
                queso = _calcular_queso(tablas)
                kwh_total = sum(f["kwh_total"] for f in facturas_cfe)
                costo_total = sum(f["costo_mxn"] for f in facturas_cfe)
                costo_unit = costo_total / kwh_total if kwh_total > 0 else 0.0
                periodo_label = _describir_periodo_contabilidad(cfe_invoices, gas_invoices, [], tipo_suministro)
                num_elec_sel = len(facturas_cfe)
                num_gas_sel = len(facturas_gas)
                historico_ppa = []
                suministrador_ppa = ""
                facturas_elec = facturas_cfe
        except Exception as _e:
            logger.exception("Error en contabilidad/data: %s", _e)
            return jsonify({"error": "error_calculo", "mensaje": str(_e)}), 500

        if tipo_suministro == TIPO_ELECTRICO_CALIFICADO:
            if cliente["num_electricidad"] == 0 and cliente["num_gas"] == 0:
                aviso_datos = {"tipo": "sin_facturas", "num_cfe": 0, "num_gas": 0}
            elif num_elec_sel == 0:
                aviso_datos = {"tipo": "sin_seleccion"}
            else:
                aviso_datos = None
        else:
            num_cfe_total = cliente["num_cfe"]
            num_gas_total = cliente["num_gas"]
            if num_cfe_total == 0 and num_gas_total == 0:
                aviso_datos = {"tipo": "sin_facturas", "num_cfe": 0, "num_gas": 0}
            elif num_elec_sel == 0 and num_gas_sel == 0:
                aviso_datos = {"tipo": "sin_seleccion"}
            elif num_elec_sel == 0 or num_gas_sel == 0:
                aviso_datos = {"tipo": "sin_par", "num_cfe": num_elec_sel, "num_gas": num_gas_sel}
            else:
                aviso_datos = None

        return jsonify({
            "estado": "ok",
            "tipo_suministro_electrico": tipo_suministro,
            "suministrador_ppa": suministrador_ppa,
            "aviso_datos": aviso_datos,
            "cliente": {"id": cliente_id, "nombre": cliente["nombre"], "periodo_label": periodo_label},
            "kpis": {
                "num_meses": num_elec_sel,
                "kwh_total": kwh_total,
                "costo_total": costo_total,
                "costo_unit": costo_unit,
            },
            "facturas_cfe": facturas_elec,
            "facturas_gas": facturas_gas,
            "historico": historico,
            "tablas": tablas,
            "queso": queso,
            "historico_gas": historico_gas,
            "historico_ppa": historico_ppa,
        })

    @app.route("/clientes/<int:cliente_id>/dashboard/contabilidad/desglose-costo-total")
    def cliente_dashboard_contabilidad_desglose(cliente_id: int):
        """Desglose en 4 categorías del costo total del periodo seleccionado.

        Solo aplica a CFE GDMTH. Devuelve agregado sobre las facturas
        seleccionadas (mismo conjunto que usa el dashboard principal).
        """
        from flask import jsonify
        from decimal import Decimal
        cliente, err = _verificar_cliente_activo(cliente_id)
        if err:
            return err

        tipo_suministro = get_tipo_suministro_electrico_seleccionado(cliente_id)
        if tipo_suministro == TIPO_ELECTRICO_CALIFICADO:
            return jsonify({"error": "No aplica a PPA"}), 400

        cfe_invoices, _, _, _ = _cargar_facturas_seleccionadas(cliente_id)

        energia = Decimal("0")
        capacidad = Decimal("0")
        distribucion = Decimal("0")
        otros = Decimal("0")

        for inv in cfe_invoices:
            comp = {c.nombre: c for c in inv.componentes_mem}
            energia += (
                (comp["Generación B"].importe_mxn if "Generación B" in comp else Decimal("0"))
                + (comp["Generación I"].importe_mxn if "Generación I" in comp else Decimal("0"))
                + (comp["Generación P"].importe_mxn if "Generación P" in comp else Decimal("0"))
            )
            capacidad += comp["Capacidad"].importe_mxn if "Capacidad" in comp else Decimal("0")
            distribucion += comp["Distribución"].importe_mxn if "Distribución" in comp else Decimal("0")
            otros += (
                (comp["Transmisión"].importe_mxn if "Transmisión" in comp else Decimal("0"))
                + (comp["CENACE"].importe_mxn if "CENACE" in comp else Decimal("0"))
                + (comp["SCnMEM"].importe_mxn if "SCnMEM" in comp else Decimal("0"))
                + (comp["Suministro"].importe_mxn if "Suministro" in comp else Decimal("0"))
                + (inv.cargo_factor_potencia_mxn or Decimal("0"))
            )

        total = energia + capacidad + distribucion + otros

        def _pct(v, t):
            return round(float(v / t * 100)) if t > 0 else 0

        # Serie mensual — un registro por factura, ordenado cronológicamente
        mensual = []
        for inv in sorted(cfe_invoices, key=lambda x: x.periodo_inicio):
            comp_m = {c.nombre: c for c in inv.componentes_mem}
            e_m = (
                (comp_m["Generación B"].importe_mxn if "Generación B" in comp_m else Decimal("0"))
                + (comp_m["Generación I"].importe_mxn if "Generación I" in comp_m else Decimal("0"))
                + (comp_m["Generación P"].importe_mxn if "Generación P" in comp_m else Decimal("0"))
            )
            c_m = comp_m["Capacidad"].importe_mxn if "Capacidad" in comp_m else Decimal("0")
            d_m = comp_m["Distribución"].importe_mxn if "Distribución" in comp_m else Decimal("0")
            o_m = (
                (comp_m["Transmisión"].importe_mxn if "Transmisión" in comp_m else Decimal("0"))
                + (comp_m["CENACE"].importe_mxn if "CENACE" in comp_m else Decimal("0"))
                + (comp_m["SCnMEM"].importe_mxn if "SCnMEM" in comp_m else Decimal("0"))
                + (comp_m["Suministro"].importe_mxn if "Suministro" in comp_m else Decimal("0"))
                + (inv.cargo_factor_potencia_mxn or Decimal("0"))
            )
            t_m = e_m + c_m + d_m + o_m
            kwh_m = float(sum(p.consumo_kwh for p in inv.periodos))
            mes_label = date(*mes_asociado(inv.periodo_inicio, inv.periodo_fin), 1).strftime("%b %Y")
            mensual.append({
                "mes":        mes_label,
                "energia":    float(e_m),
                "capacidad":  float(c_m),
                "distribucion": float(d_m),
                "otros":      float(o_m),
                "total":      float(t_m),
                "kwh":        kwh_m,
                "costo_unit": float(t_m / Decimal(str(kwh_m))) if kwh_m > 0 else 0.0,
            })

        return jsonify({
            "lineas": [
                {"nombre": "Energía",         "monto": float(energia),      "pct": _pct(energia, total)},
                {"nombre": "Capacidad",       "monto": float(capacidad),    "pct": _pct(capacidad, total)},
                {"nombre": "Distribución",    "monto": float(distribucion), "pct": _pct(distribucion, total)},
                {"nombre": "Otros Servicios", "monto": float(otros),        "pct": _pct(otros, total)},
            ],
            "total": float(total),
            "mensual": mensual,
        })

    @app.route("/clientes/<int:cliente_id>/dashboard/cogeneracion/data")
    def cliente_dashboard_cogeneracion_data(cliente_id: int):
        """JSON con todos los datos del dashboard de Cogeneración."""
        from flask import jsonify
        from decimal import Decimal as _D

        activo_id = session.get("cliente_activo_id")
        if activo_id != cliente_id:
            return jsonify({"error": "no_autorizado"}), 403
        from storage.repository import get_cliente_con_conteos as _gcc
        cliente = _gcc(cliente_id)
        if cliente is None:
            return jsonify({"error": "no_encontrado"}), 404

        tipo_suministro = get_tipo_suministro_electrico_seleccionado(cliente_id)

        try:
            _cfg = {row["clave"]: row["valor"] for row in list_configuracion()}
            tc_str = _cfg.get("tipo_cambio_mxn_usd")
            tipo_cambio = _D(tc_str) if tc_str else _D("17.50")
            fe_elec_str = _cfg.get("factor_emision_electricidad_kg_co2_kwh")
            fe_gas_str  = _cfg.get("factor_emision_gas_kg_co2_gj")
            fe_elec = _D(fe_elec_str) if fe_elec_str else None
            fe_gas  = _D(fe_gas_str)  if fe_gas_str  else None

            if tipo_suministro == TIPO_ELECTRICO_CALIFICADO:
                ppa_invoices, gas_invoices, facturas_ppa, facturas_gas = _cargar_ultimas_ppa_cogen(cliente_id)
                r = calcular_cogen_ppa(
                    ppa_invoices, gas_invoices, CoGenParams(),
                    tipo_cambio=tipo_cambio,
                    factor_emision_elec=fe_elec,
                    factor_emision_gas=fe_gas,
                )
                cfe_invoices = []
                facturas_cfe = []
                precio_gas_fuente = "ppa"
            else:
                cfe_invoices, gas_invoices, facturas_cfe, facturas_gas = _cargar_ultimas_facturas_cogen(cliente_id)
                ppa_invoices = []
                facturas_ppa = []
                _precio_manual_str = cliente.get("precio_gas_manual_mxn_gj_pcs")
                _precio_manual = _D(_precio_manual_str) if _precio_manual_str else None
                if len(cfe_invoices) >= 12 and len(gas_invoices) < 12 and _precio_manual:
                    r = calcular_cogen_precio_manual(
                        cfe_invoices, _precio_manual, CoGenParams(),
                        tipo_cambio=tipo_cambio,
                        factor_emision_elec=fe_elec,
                        factor_emision_gas=fe_gas,
                    )
                    precio_gas_fuente = "manual"
                else:
                    r = calcular_cogen(
                        cfe_invoices, gas_invoices, CoGenParams(),
                        tipo_cambio=tipo_cambio,
                        factor_emision_elec=fe_elec,
                        factor_emision_gas=fe_gas,
                    )
                    precio_gas_fuente = "real"
        except Exception as _e:
            logger.exception("Error en cogeneracion/data: %s", _e)
            return jsonify({"error": "error_calculo", "mensaje": str(_e)}), 500

        # CELs — aplica igual para GDMTH y PPA (fórmula CRE Caso I independiente del suministro)
        cels_resultado = None
        try:
            from calc.cels import calcular_cels as _calcular_cels
            calor_recuperado_anual = sum(m.calor_recuperado_gj for m in r.meses)
            cels_resultado = _calcular_cels(
                kwh_cubiertos_anual=r.kwh_cubiertos_anual,
                gj_gas_cogen_pci_anual=r.gj_gas_cogen_pci_anual,
                calor_recuperado_gj_anual=calor_recuperado_anual,
                capacidad_nominal_kw=r.capacidad_nominal_kw,
                medio_termico=cliente.get("medio_termico"),
                nivel_tension_kv=cliente.get("nivel_tension_kv"),
                altitud_msnm=cliente.get("altitud_msnm"),
                tipo_motor=cliente.get("tipo_motor"),
                medio_termico_vapor_pct=cliente.get("medio_termico_vapor_pct"),
            )
        except Exception as _e_cels:
            logger.error("Error calculando CELs en data endpoint: %s", _e_cels)

        # Energía limpia generada — aplica igual para GDMTH y PPA
        if (cels_resultado is not None and cels_resultado.es_eficiente
                and cels_resultado.cels_mwh_anual is not None
                and r.kwh_total_anual > 0):
            from decimal import Decimal as _D2, ROUND_HALF_UP as _RHU
            _pct = (_D2(str(cels_resultado.cels_mwh_anual)) * _D2("1000") / r.kwh_total_anual * _D2("100"))
            r.energia_limpia_pct = _pct.quantize(_D2("0.01"), _RHU)
        else:
            r.energia_limpia_pct = None

        num_cfe_total = cliente["num_cfe"]
        num_gas_total = cliente["num_gas"]
        num_elec_sel = len(facturas_ppa) if tipo_suministro == TIPO_ELECTRICO_CALIFICADO else len(facturas_cfe)
        num_gas_sel = len(facturas_gas)

        # Validación 12 meses
        if tipo_suministro != TIPO_ELECTRICO_CALIFICADO:
            _precio_manual_str = cliente.get("precio_gas_manual_mxn_gj_pcs")
            _precio_manual = _D(_precio_manual_str) if _precio_manual_str else None
            if num_elec_sel < 12:
                aviso_datos = {"tipo": "insuficiente_elec", "n_elec": num_elec_sel}
            elif num_gas_sel < 12 and not _precio_manual:
                aviso_datos = {"tipo": "insuficiente_gas", "n_gas": num_gas_sel}
            elif len(r.meses) < 12:
                aviso_datos = {"tipo": "meses_no_coinciden", "n_pares": len(r.meses),
                               "n_elec": num_elec_sel, "n_gas": num_gas_sel}
            else:
                aviso_datos = None
        elif num_cfe_total == 0 and num_gas_total == 0:
            aviso_datos = {"tipo": "sin_facturas", "num_cfe": 0, "num_gas": 0}
        elif num_elec_sel == 0 and num_gas_sel == 0:
            aviso_datos = {"tipo": "sin_seleccion"}
        elif num_elec_sel == 0 or num_gas_sel == 0:
            aviso_datos = {"tipo": "sin_par", "num_cfe": num_elec_sel, "num_gas": num_gas_sel}
        elif not r.meses:
            aviso_datos = {"tipo": "sin_pares_mes", "num_cfe": num_elec_sel, "num_gas": num_gas_sel}
        else:
            aviso_datos = None

        rango_cogen = _calcular_rango_cogen(cfe_invoices, gas_invoices, ppa_invoices if ppa_invoices else None)

        chart_labels = [m.periodo_inicio.strftime("%b %Y") for m in r.meses]
        chart_ebitda = [float(m.ebitda_mes_mxn) for m in r.meses]
        chart_ahorro_elec = [float(m.ahorro_electricidad_mxn) for m in r.meses]
        chart_ahorro_caldera = [float(m.ahorro_caldera_mxn) for m in r.meses]
        chart_costo_gas = [float(m.costo_gas_cogen_mxn) for m in r.meses]
        chart_om = [float(m.gasto_om_mes_mxn) for m in r.meses]

        meses_raw = [
            {
                "periodo": m.periodo_inicio.strftime("%b %Y"),
                "kwh_total": float(m.kwh_total),
                "costo_cfe_mxn": float(m.costo_cfe_mxn),
                "costo_promedio_kwh": float(m.costo_promedio_kwh),
                "gj_consumido": float(m.gj_consumido),
                "costo_unitario_gj": float(m.costo_unitario_gj),
                "costo_gas_actual_mxn": float(m.costo_gas_actual_mxn),
                "kwh_punta": float(m.kwh_punta_total),
                "kwh_intermedia": float(m.kwh_intermedia_total),
                "kwh_base": float(m.kwh_base_total),
                "cu_punta": float(m.cu_punta_kwh),
                "cu_intermedia": float(m.cu_intermedia_kwh),
                "cu_base": float(m.cu_base_kwh),
                "kw_max": float(m.kw_max),
                "kw_punta": float(m.kw_punta_orig),
                "dias_mes": m.dias_facturados,
                "kwh_total_orig": float(m.kwh_total_orig),
                "precio_capacidad_kw": float(m.precio_capacidad_kw),
                "precio_distribucion_kw": float(m.precio_distribucion_kw),
                "kw_facturado_capacidad": float(m.kw_facturado_capacidad),
                "kw_facturado_distribucion": float(m.kw_facturado_distribucion),
                "precio_otros_kwh": float(m.precio_otros_mxn_kwh),
            }
            for m in r.meses
        ]

        tabla_mensual = [
            {
                "periodo": m.periodo_inicio.strftime("%b %Y"),
                "prorrateado": m.prorrateado,
                "nota_prorrateo": m.nota_prorrateo,
                "kwh_total": float(m.kwh_total),
                "costo_cfe_mxn": float(m.costo_cfe_mxn),
                "costo_promedio_kwh": float(m.costo_promedio_kwh),
                "gj_consumido": float(m.gj_consumido),
                "costo_unitario_gj": float(m.costo_unitario_gj),
                "costo_gas_actual_mxn": float(m.costo_gas_actual_mxn),
                "kwh_cubiertos": float(m.kwh_cubiertos),
                "kwh_punta_cubierto": float(m.kwh_punta_cubierto),
                "kwh_intermedia_cubierto": float(m.kwh_intermedia_cubierto),
                "kwh_base_cubierto": float(m.kwh_base_cubierto),
                "ahorro_energia_mes_mxn": float(m.ahorro_energia_mes_mxn),
                "ahorro_capacidad_mes_mxn": float(m.ahorro_capacidad_mes_mxn),
                "ahorro_distribucion_mes_mxn": float(m.ahorro_distribucion_mes_mxn),
                "ahorro_otros_servicios_mes_mxn": float(m.ahorro_otros_servicios_mes_mxn),
                "gj_gas_cogen": float(m.gj_gas_cogen),
                "costo_gas_cogen_mxn": float(m.costo_gas_cogen_mxn),
                "ahorro_electricidad_mxn": float(m.ahorro_electricidad_mxn),
                "calor_recuperado_gj": float(m.calor_recuperado_gj),
                "ahorro_caldera_mxn": float(m.ahorro_caldera_mxn),
                "gasto_om_mes_mxn": float(m.gasto_om_mes_mxn),
                "ebitda_mes_mxn": float(m.ebitda_mes_mxn),
            }
            for m in r.meses
        ]

        if r.inversion_mxn is not None and r.inversion_mxn > 0:
            payback_inicial = calcular_payback_decimal(r.inversion_mxn, r.ebitda_anual_mxn, r.ebitda_anual_mxn)
            flujo_acum_15   = [float(v) for v in calcular_flujo_acumulado(r.inversion_mxn, r.ebitda_anual_mxn)]
            flujo_anual_15  = [-float(r.inversion_mxn)] + [float(r.ebitda_anual_mxn)] * 15
            # Flujo con beneficio fiscal año 1 (depreciación inmediata Art. 34 XIII LISR)
            if r.flujo_anio_1_con_beneficio_mxn is not None:
                flujo_anual_15_fiscal = (
                    [-float(r.inversion_mxn), float(r.flujo_anio_1_con_beneficio_mxn)]
                    + [float(r.ebitda_anual_mxn)] * 14
                )
                _acum = 0.0
                flujo_acum_15_fiscal = []
                for v in flujo_anual_15_fiscal:
                    _acum += v
                    flujo_acum_15_fiscal.append(_acum)
                # Payback con beneficio fiscal (decimal, interpolación lineal)
                payback_con_beneficio = calcular_payback_decimal(
                    r.inversion_mxn, r.flujo_anio_1_con_beneficio_mxn, r.ebitda_anual_mxn
                )
                if payback_con_beneficio is None:
                    payback_con_beneficio = -1  # > 15 años (sin retorno)
                else:
                    payback_con_beneficio = float(payback_con_beneficio)
            else:
                flujo_anual_15_fiscal = flujo_anual_15
                flujo_acum_15_fiscal = flujo_acum_15
                payback_con_beneficio = float(payback_inicial) if payback_inicial is not None else None
            payback_inicial = float(payback_inicial) if payback_inicial is not None else None
        else:
            payback_inicial = None
            flujo_acum_15 = []
            flujo_anual_15 = []
            flujo_anual_15_fiscal = []
            flujo_acum_15_fiscal = []
            payback_con_beneficio = None

        co2 = None
        if r.co2_reduccion_kg_anual is not None:
            reduccion_t = float(r.co2_reduccion_kg_anual) / 1000
            co2 = {
                "actual_total_t": float(r.co2_actual_total_kg_anual) / 1000 if r.co2_actual_total_kg_anual else None,
                "reduccion_t": reduccion_t,
                "reduccion_pct": float(r.co2_reduccion_porcentaje) if r.co2_reduccion_porcentaje else 0.0,
                "arboles": int(reduccion_t * 50),
                "factor_emision_elec": float(fe_elec) if fe_elec else None,
                "factor_emision_gas": float(fe_gas) if fe_gas else None,
            }

        periodo_label = _calcular_periodo_label(cfe_invoices, gas_invoices, ppa_invoices if ppa_invoices else None)
        suministrador_ppa = facturas_ppa[0]["suministrador"] if facturas_ppa else ""

        return jsonify({
            "estado": "ok",
            "tipo_suministro_electrico": tipo_suministro,
            "suministrador_ppa": suministrador_ppa,
            "precio_gas_fuente": precio_gas_fuente,
            "aviso_datos": aviso_datos,
            "rango_cogen": rango_cogen,
            "cliente": {"id": cliente_id, "nombre": cliente["nombre"], "periodo_label": periodo_label},
            "kpis": {
                "ahorro_electricidad_anual": float(r.ahorro_electricidad_anual_mxn),
                "ahorro_caldera_anual": float(r.ahorro_caldera_anual_mxn),
                "costo_gas_cogen_anual": float(r.costo_gas_cogen_anual_mxn),
                "gasto_om_anual": float(r.gasto_om_anual_mxn),
                "ebitda_anual": float(r.ebitda_anual_mxn),
                "kwh_total_anual": float(r.kwh_total_anual),
                "kwh_cubiertos_anual": float(r.kwh_cubiertos_anual),
                "gj_gas_cogen_anual": float(r.gj_gas_cogen_anual),
                "capacidad_nominal_kw": float(r.capacidad_nominal_kw) if r.capacidad_nominal_kw else None,
                "inversion_usd": float(r.inversion_usd) if r.inversion_usd else None,
                "inversion_mxn": float(r.inversion_mxn) if r.inversion_mxn else None,
                "tipo_cambio": float(r.tipo_cambio_mxn_usd) if r.tipo_cambio_mxn_usd else None,
                "ahorro_energia_anual": float(r.ahorro_energia_anual_mxn),
                "ahorro_capacidad_anual": float(r.ahorro_capacidad_anual_mxn),
                "ahorro_distribucion_anual": float(r.ahorro_distribucion_anual_mxn),
                "ahorro_otros_servicios_anual": float(r.ahorro_otros_servicios_anual_mxn),
                "beneficio_fiscal_anio_1_mxn": float(r.beneficio_fiscal_anio_1_mxn) if r.beneficio_fiscal_anio_1_mxn else None,
                "energia_limpia_pct": float(r.energia_limpia_pct) if r.energia_limpia_pct else None,
            },
            "co2": co2,
            "cels": _cels_to_dict(cels_resultado),
            "chart_labels": chart_labels,
            "chart_ebitda": chart_ebitda,
            "chart_ahorro_elec": chart_ahorro_elec,
            "chart_ahorro_caldera": chart_ahorro_caldera,
            "chart_costo_gas": chart_costo_gas,
            "chart_om": chart_om,
            "meses_raw": meses_raw,
            "tabla_mensual": tabla_mensual,
            "totales": {
                "kwh_total_anual": float(r.kwh_total_anual),
                "kwh_cubiertos_anual": float(r.kwh_cubiertos_anual),
                "gj_gas_cogen_anual": float(r.gj_gas_cogen_anual),
                "costo_gas_cogen_anual_mxn": float(r.costo_gas_cogen_anual_mxn),
                "ahorro_electricidad_anual_mxn": float(r.ahorro_electricidad_anual_mxn),
                "ahorro_caldera_anual_mxn": float(r.ahorro_caldera_anual_mxn),
                "gasto_om_anual_mxn": float(r.gasto_om_anual_mxn),
                "ebitda_anual_mxn": float(r.ebitda_anual_mxn),
                "ahorro_energia_anual_mxn": float(r.ahorro_energia_anual_mxn),
                "ahorro_capacidad_anual_mxn": float(r.ahorro_capacidad_anual_mxn),
                "ahorro_distribucion_anual_mxn": float(r.ahorro_distribucion_anual_mxn),
                "ahorro_otros_servicios_anual_mxn": float(r.ahorro_otros_servicios_anual_mxn),
            },
            "payback_inicial": payback_inicial,
            "flujo_acum_15": flujo_acum_15,
            "flujo_anual_15": flujo_anual_15,
            "flujo_anual_15_fiscal": flujo_anual_15_fiscal,
            "flujo_acum_15_fiscal": flujo_acum_15_fiscal,
            "payback_con_beneficio": payback_con_beneficio,
            "params": {
                "cobertura_electrica": float(r.params.cobertura_electrica),
                "rendimiento_electrico": float(r.params.rendimiento_electrico),
                "rendimiento_termico": float(r.params.rendimiento_termico),
                "eficiencia_caldera": float(r.params.eficiencia_caldera),
            },
            "cliente_ficha_url": url_for("clientes.ficha", cliente_id=cliente_id),
        })

    @app.route("/clientes/<int:cliente_id>/dashboard/contabilidad/export-datos")
    def cliente_dashboard_contabilidad_export(cliente_id: int):
        """Descarga Excel con los datos del dashboard de Contabilidad Energética."""
        import io
        from decimal import Decimal
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter

        cliente, err = _verificar_cliente_activo(cliente_id)
        if err:
            return err

        tipo_suministro = get_tipo_suministro_electrico_seleccionado(cliente_id)

        try:
            if tipo_suministro == TIPO_ELECTRICO_CALIFICADO:
                ppa_invoices, gas_invoices, facturas_ppa, facturas_gas = _cargar_facturas_ppa(cliente_id)
                cfe_invoices = []
                facturas_cfe = []
                tablas = {}
                historico = None
                historico_gas = calcular_historico_gas(gas_invoices)
            else:
                cfe_invoices, gas_invoices, facturas_cfe, facturas_gas = _cargar_facturas_seleccionadas(cliente_id)
                tablas = calcular_tablas_cfe(cfe_invoices)
                historico = calcular_historico_cfe(cfe_invoices)
                historico_gas = calcular_historico_gas(gas_invoices)
                ppa_invoices = []
                facturas_ppa = []
        except Exception as _e:
            logger.exception("Error en contabilidad/export-datos: %s", _e)
            return f"Error generando Excel: {_e}", 500

        wb = Workbook()
        _HDR_FILL = PatternFill("solid", fgColor="E8F4ED")
        _HDR_FONT = Font(bold=True, color="155936")
        _TOT_FILL = PatternFill("solid", fgColor="F0F4F1")
        _TOT_FONT = Font(bold=True)
        _FMT_MXN  = '$#,##0'
        _FMT_KWH  = '#,##0'
        _FMT_GJ   = '#,##0.00'
        _FMT_PCT  = '0.0%'

        def _hdr_row(ws, cols, row=1):
            for c, title in enumerate(cols, 1):
                cell = ws.cell(row=row, column=c, value=title)
                cell.font = _HDR_FONT
                cell.fill = _HDR_FILL
                cell.alignment = Alignment(horizontal="center")

        def _autofit(ws):
            for col in ws.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 32)

        _PERIODOS = ("base", "intermedio", "punta")

        def _to_excel_value(v):
            """Convierte un valor a tipo que openpyxl acepta en celda."""
            if v is None:
                return None
            if isinstance(v, bool):
                return v
            if isinstance(v, (int, float, str)):
                return v
            if hasattr(v, "isoformat"):   # date / datetime
                return v
            if hasattr(v, "__float__"):
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return str(v)
            return str(v)   # fallback seguro para dict, list, etc.

        def _expandir_kpis(kpis):
            """Expande entradas cuyo valor sea dict por periodo horario a 3 filas."""
            resultado = []
            for label, v in kpis:
                if isinstance(v, dict) and any(k in _PERIODOS for k in v):
                    for p in _PERIODOS:
                        if p in v:
                            resultado.append((f"{label} ({p.capitalize()})", v[p]))
                else:
                    resultado.append((label, v))
            return resultado

        nombre_cliente = cliente["nombre"]

        if tipo_suministro == TIPO_ELECTRICO_CALIFICADO:
            # ── Hoja 1: KPIs PPA ───────────────────────────────────────────────
            ws = wb.active
            ws.title = "KPIs"
            kwh_total_ppa = sum(f["kwh_total"] for f in facturas_ppa)
            costo_total_ppa = sum(f["costo_mxn"] for f in facturas_ppa)
            kpis = [
                ("Cliente", nombre_cliente),
                ("Meses seleccionados", len(facturas_ppa)),
                ("kWh total (PPA)", kwh_total_ppa),
                ("Costo total PPA (MXN)", costo_total_ppa),
                ("Costo unitario prom. (MXN/kWh)", costo_total_ppa / kwh_total_ppa if kwh_total_ppa else 0),
            ]
            for r_i, (k, v) in enumerate(kpis, 1):
                ws.cell(r_i, 1, k).font = Font(bold=True)
                ws.cell(r_i, 2, v)
            _autofit(ws)

            # ── Hoja 2: Facturas PPA ───────────────────────────────────────────
            ws2 = wb.create_sheet("Facturas PPA")
            cols2 = ["Mes", "Suministrador", "kWh Total", "Precio Unit. (MXN/kWh)", "Costo (MXN)"]
            _hdr_row(ws2, cols2)
            for i, f in enumerate(facturas_ppa, 2):
                ws2.cell(i, 1, f["mes_asociado"])
                ws2.cell(i, 2, f["suministrador"])
                ws2.cell(i, 3, f["kwh_total"]).number_format = _FMT_KWH
                ws2.cell(i, 4, f["precio_unitario_mxn_kwh"]).number_format = '$#,##0.0000'
                ws2.cell(i, 5, f["costo_mxn"]).number_format = _FMT_MXN
            tot_row = len(facturas_ppa) + 2
            ws2.cell(tot_row, 1, "TOTAL").font = _TOT_FONT
            ws2.cell(tot_row, 1).fill = _TOT_FILL
            ws2.cell(tot_row, 3, kwh_total_ppa).number_format = _FMT_KWH
            ws2.cell(tot_row, 3).font = _TOT_FONT
            ws2.cell(tot_row, 3).fill = _TOT_FILL
            ws2.cell(tot_row, 5, costo_total_ppa).number_format = _FMT_MXN
            ws2.cell(tot_row, 5).font = _TOT_FONT
            ws2.cell(tot_row, 5).fill = _TOT_FILL
            _autofit(ws2)

        else:
            # ── Hoja 1: KPIs CFE ───────────────────────────────────────────────
            ws = wb.active
            ws.title = "KPIs"
            ind = tablas.get("indicadores", [])
            kwh_total_cfe = sum(f["kwh_total"] for f in facturas_cfe)
            costo_total_cfe = sum(f["costo_mxn"] for f in facturas_cfe)
            kpis = _expandir_kpis([
                ("Cliente", nombre_cliente),
                ("Meses seleccionados", len(facturas_cfe)),
                ("kWh total", kwh_total_cfe),
                ("Costo total CFE (MXN)", costo_total_cfe),
                ("Costo unitario prom. (MXN/kWh)", tablas.get("costo_unit_promedio_total", 0)),
            ])
            for r_i, (k, v) in enumerate(kpis, 1):
                ws.cell(r_i, 1, k).font = Font(bold=True)
                ws.cell(r_i, 2, _to_excel_value(v))
            _autofit(ws)

            # ── Hoja 2: Consumos y demandas ────────────────────────────────────
            ws2 = wb.create_sheet("Consumos y demandas")
            cols2 = ["Mes", "kWh Base", "kWh Intermedio", "kWh Punta", "kWh Total",
                     "kW Base", "kW Intermedio", "kW Punta"]
            _hdr_row(ws2, cols2)
            cd = tablas.get("consumos_demandas", [])
            for i, row in enumerate(cd, 2):
                ws2.cell(i, 1, row["mes"])
                ws2.cell(i, 2, float(row["kwh_base"])).number_format = _FMT_KWH
                ws2.cell(i, 3, float(row["kwh_inter"])).number_format = _FMT_KWH
                ws2.cell(i, 4, float(row["kwh_punta"])).number_format = _FMT_KWH
                ws2.cell(i, 5, float(row["kwh_total"])).number_format = _FMT_KWH
                ws2.cell(i, 6, float(row.get("kw_base") or 0)).number_format = '#,##0.0'
                ws2.cell(i, 7, float(row.get("kw_inter") or 0)).number_format = '#,##0.0'
                ws2.cell(i, 8, float(row.get("kw_punta") or 0)).number_format = '#,##0.0'
            _autofit(ws2)

            # ── Hoja 3: Costos detallados ──────────────────────────────────────
            ws3 = wb.create_sheet("Costos detallados")
            cols3 = ["Mes", "CE Base (MXN)", "CE Inter (MXN)", "CE Punta (MXN)", "CE Total (MXN)",
                     "Costo Dist. (MXN)", "Costo Cap. (MXN)", "Cargo FP (MXN)",
                     "Subtotal (MXN)", "CU Base ($/kWh)", "CU Inter ($/kWh)", "CU Punta ($/kWh)"]
            _hdr_row(ws3, cols3)
            cos = tablas.get("costos_detallados", [])
            for i, row in enumerate(cos, 2):
                ws3.cell(i, 1, row["mes"])
                ws3.cell(i, 2, float(row["ce_base"])).number_format = _FMT_MXN
                ws3.cell(i, 3, float(row["ce_inter"])).number_format = _FMT_MXN
                ws3.cell(i, 4, float(row["ce_punta"])).number_format = _FMT_MXN
                ws3.cell(i, 5, float(row["ce_total"])).number_format = _FMT_MXN
                ws3.cell(i, 6, float(row.get("costo_dist", 0))).number_format = _FMT_MXN
                ws3.cell(i, 7, float(row.get("costo_cap", 0))).number_format = _FMT_MXN
                ws3.cell(i, 8, float(row.get("cargo_fp", 0))).number_format = _FMT_MXN
                ws3.cell(i, 9, float(row.get("subtotal", 0))).number_format = _FMT_MXN
                ws3.cell(i, 10, float(row.get("cu_base_total") or 0)).number_format = '$#,##0.0000'
                ws3.cell(i, 11, float(row.get("cu_inter_total") or 0)).number_format = '$#,##0.0000'
                ws3.cell(i, 12, float(row.get("cu_punta_total") or 0)).number_format = '$#,##0.0000'
            _autofit(ws3)

            # ── Hoja 4: Indicadores ────────────────────────────────────────────
            ws4 = wb.create_sheet("Indicadores")
            cols4 = ["Mes", "Costo Unit. (MXN/kWh)", "% Energía", "% Demanda",
                     "Factor de Carga", "Demanda Prom. (kW)"]
            _hdr_row(ws4, cols4)
            for i, row in enumerate(ind, 2):
                ws4.cell(i, 1, row["mes"])
                ws4.cell(i, 2, float(row["costo_unit"])).number_format = '$#,##0.0000'
                ws4.cell(i, 3, float(row["pct_energia"]) / 100).number_format = _FMT_PCT
                ws4.cell(i, 4, float(row["pct_demanda"]) / 100).number_format = _FMT_PCT
                ws4.cell(i, 5, float(row.get("factor_carga", 0))).number_format = '0.00'
                ws4.cell(i, 6, float(row.get("demanda_prom", 0))).number_format = '#,##0.0'
            _autofit(ws4)

        # ── Hoja Gas ───────────────────────────────────────────────────────────
        ws_gas = wb.create_sheet("Gas natural")
        cols_g = ["Mes", "Consumo (GJ)", "Costo Unit. (MXN/GJ)", "Costo Total (MXN)"]
        _hdr_row(ws_gas, cols_g)
        gas_filas = historico_gas["filas"] if historico_gas else []
        for i, row in enumerate(gas_filas, 2):
            ws_gas.cell(i, 1, row.get("mes", ""))
            ws_gas.cell(i, 2, float(row.get("consumo_gj") or 0)).number_format = _FMT_GJ
            ws_gas.cell(i, 3, float(row.get("costo_unit_gj") or 0)).number_format = '$#,##0.00'
            ws_gas.cell(i, 4, float(row.get("costo_total_mxn") or 0)).number_format = _FMT_MXN
        if historico_gas:
            tot = historico_gas["total"]
            tr = len(gas_filas) + 2
            ws_gas.cell(tr, 1, "TOTAL").font = _TOT_FONT
            ws_gas.cell(tr, 1).fill = _TOT_FILL
            ws_gas.cell(tr, 2, float(tot.get("consumo_gj") or 0)).number_format = _FMT_GJ
            ws_gas.cell(tr, 2).font = _TOT_FONT; ws_gas.cell(tr, 2).fill = _TOT_FILL
            ws_gas.cell(tr, 3, float(tot.get("costo_unit_gj") or 0)).number_format = '$#,##0.00'
            ws_gas.cell(tr, 3).font = _TOT_FONT; ws_gas.cell(tr, 3).fill = _TOT_FILL
            ws_gas.cell(tr, 4, float(tot.get("costo_total_mxn") or 0)).number_format = _FMT_MXN
            ws_gas.cell(tr, 4).font = _TOT_FONT; ws_gas.cell(tr, 4).fill = _TOT_FILL
        _autofit(ws_gas)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        slug = nombre_cliente.lower().replace(" ", "_")
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"contabilidad_{slug}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/clientes/<int:cliente_id>/dashboard/cogeneracion/export-datos")
    def cliente_dashboard_cogeneracion_export(cliente_id: int):
        """Descarga Excel con los datos del dashboard de Cogeneración."""
        import io
        from decimal import Decimal as _D
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        cliente, err = _verificar_cliente_activo(cliente_id)
        if err:
            return err

        # Leer parámetros de slider desde query string
        def _qparam(name, default, lo=0.01, hi=1.0):
            try:
                v = float(request.args.get(name, default))
                return _D(str(max(lo, min(hi, v))))
            except (TypeError, ValueError):
                return _D(str(default))

        params = CoGenParams(
            cobertura_electrica=_qparam("cobertura", 0.75, 0.50, 0.95),
            rendimiento_electrico=_qparam("rend_elec", 0.40, 0.10, 0.60),
            rendimiento_termico=_qparam("rend_term", 0.25, 0.05, 0.50),
            eficiencia_caldera=_qparam("eficiencia_caldera", 0.85, 0.50, 0.99),
        )

        tipo_suministro = get_tipo_suministro_electrico_seleccionado(cliente_id)

        try:
            _cfg = {row["clave"]: row["valor"] for row in list_configuracion()}
            tc_str = _cfg.get("tipo_cambio_mxn_usd")
            tipo_cambio = _D(tc_str) if tc_str else _D("17.50")

            if tipo_suministro == TIPO_ELECTRICO_CALIFICADO:
                ppa_invoices, gas_invoices, _, _ = _cargar_ultimas_ppa_cogen(cliente_id)
                r = calcular_cogen_ppa(ppa_invoices, gas_invoices, params, tipo_cambio=tipo_cambio)
            else:
                cfe_invoices, gas_invoices, _, _ = _cargar_ultimas_facturas_cogen(cliente_id)
                _pm_str = cliente.get("precio_gas_manual_mxn_gj_pcs")
                _pm = _D(_pm_str) if _pm_str else None
                if len(cfe_invoices) >= 12 and len(gas_invoices) < 12 and _pm:
                    r = calcular_cogen_precio_manual(cfe_invoices, _pm, params, tipo_cambio=tipo_cambio)
                else:
                    r = calcular_cogen(cfe_invoices, gas_invoices, params, tipo_cambio=tipo_cambio)
        except Exception as _e:
            logger.exception("Error en cogeneracion/export-datos: %s", _e)
            return f"Error generando Excel: {_e}", 500

        wb = Workbook()
        _HDR_FILL = PatternFill("solid", fgColor="E8F4ED")
        _HDR_FONT = Font(bold=True, color="155936")
        _TOT_FILL = PatternFill("solid", fgColor="F0F4F1")
        _TOT_FONT = Font(bold=True)
        _FMT_MXN  = '$#,##0'
        _FMT_KWH  = '#,##0'
        _FMT_GJ   = '#,##0.00'
        _FMT_PCT  = '0.0%'

        def _hdr_row(ws, cols, row=1):
            for c, title in enumerate(cols, 1):
                cell = ws.cell(row=row, column=c, value=title)
                cell.font = _HDR_FONT
                cell.fill = _HDR_FILL
                cell.alignment = Alignment(horizontal="center")

        def _autofit(ws):
            for col in ws.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 32)

        nombre_cliente = cliente["nombre"]

        # ── Hoja 1: KPIs ──────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "KPIs"
        kpis = [
            ("Cliente", nombre_cliente),
            ("kWh total anual", float(r.kwh_total_anual)),
            ("kWh cubiertos anual", float(r.kwh_cubiertos_anual)),
            ("Ahorro eléctrico anual (MXN)", float(r.ahorro_electricidad_anual_mxn)),
            ("Ahorro caldera anual (MXN)", float(r.ahorro_caldera_anual_mxn)),
            ("Costo gas cogeneración anual (MXN)", float(r.costo_gas_cogen_anual_mxn)),
            ("Gasto O&M anual (MXN)", float(r.gasto_om_anual_mxn)),
            ("EBITDA anual (MXN)", float(r.ebitda_anual_mxn)),
            ("Capacidad nominal (kW)", float(r.capacidad_nominal_kw) if r.capacidad_nominal_kw else ""),
            ("Inversión estimada (MXN)", float(r.inversion_mxn) if r.inversion_mxn else ""),
        ]
        for r_i, (k, v) in enumerate(kpis, 1):
            ws.cell(r_i, 1, k).font = Font(bold=True)
            ws.cell(r_i, 2, v)
        _autofit(ws)

        # ── Hoja 2: Parámetros ────────────────────────────────────────────────
        ws2 = wb.create_sheet("Parámetros motor")
        params_rows = [
            ("Cobertura eléctrica", float(r.params.cobertura_electrica)),
            ("Rendimiento eléctrico", float(r.params.rendimiento_electrico)),
            ("Rendimiento térmico", float(r.params.rendimiento_termico)),
            ("Eficiencia caldera ref.", float(r.params.eficiencia_caldera)),
        ]
        for r_i, (k, v) in enumerate(params_rows, 1):
            ws2.cell(r_i, 1, k).font = Font(bold=True)
            ws2.cell(r_i, 2, v).number_format = _FMT_PCT

        # ── Hoja 3: Tabla mensual ─────────────────────────────────────────────
        ws3 = wb.create_sheet("Tabla mensual")
        cols3 = ["Mes", "kWh Total", "Costo CFE (MXN)", "kWh Cubiertos",
                 "GJ Gas Cogen", "Costo Gas Cogen (MXN)",
                 "Ahorro Eléctrico (MXN)", "Ahorro Caldera (MXN)",
                 "O&M (MXN)", "EBITDA Mes (MXN)"]
        _hdr_row(ws3, cols3)
        for i, m in enumerate(r.meses, 2):
            ws3.cell(i, 1, m.periodo_inicio.strftime("%b %Y"))
            ws3.cell(i, 2, float(m.kwh_total)).number_format = _FMT_KWH
            ws3.cell(i, 3, float(m.costo_cfe_mxn)).number_format = _FMT_MXN
            ws3.cell(i, 4, float(m.kwh_cubiertos)).number_format = _FMT_KWH
            ws3.cell(i, 5, float(m.gj_gas_cogen)).number_format = _FMT_GJ
            ws3.cell(i, 6, float(m.costo_gas_cogen_mxn)).number_format = _FMT_MXN
            ws3.cell(i, 7, float(m.ahorro_electricidad_mxn)).number_format = _FMT_MXN
            ws3.cell(i, 8, float(m.ahorro_caldera_mxn)).number_format = _FMT_MXN
            ws3.cell(i, 9, float(m.gasto_om_mes_mxn)).number_format = _FMT_MXN
            ws3.cell(i, 10, float(m.ebitda_mes_mxn)).number_format = _FMT_MXN
        # Fila TOTAL
        tr = len(r.meses) + 2
        ws3.cell(tr, 1, "TOTAL").font = _TOT_FONT
        ws3.cell(tr, 1).fill = _TOT_FILL
        for col_idx, val in enumerate([
            float(r.kwh_total_anual), float(sum(m.costo_cfe_mxn for m in r.meses)),
            float(r.kwh_cubiertos_anual), float(r.gj_gas_cogen_anual),
            float(r.costo_gas_cogen_anual_mxn), float(r.ahorro_electricidad_anual_mxn),
            float(r.ahorro_caldera_anual_mxn), float(r.gasto_om_anual_mxn),
            float(r.ebitda_anual_mxn),
        ], 2):
            c = ws3.cell(tr, col_idx, val)
            c.font = _TOT_FONT
            c.fill = _TOT_FILL
        _autofit(ws3)

        # ── Hoja 4: Cascada ahorro ────────────────────────────────────────────
        ws4 = wb.create_sheet("Cascada ahorro")
        cascada = [
            ("Ahorro Energía (MXN)",        float(r.ahorro_energia_anual_mxn)),
            ("Ahorro Capacidad (MXN)",       float(r.ahorro_capacidad_anual_mxn)),
            ("Ahorro Distribución (MXN)",    float(r.ahorro_distribucion_anual_mxn)),
            ("Ahorro Otros Servicios (MXN)", float(r.ahorro_otros_servicios_anual_mxn)),
            ("Ahorro Caldera (MXN)",         float(r.ahorro_caldera_anual_mxn)),
            ("Costo Gas Cogen (MXN)",        -float(r.costo_gas_cogen_anual_mxn)),
            ("Gasto O&M (MXN)",              -float(r.gasto_om_anual_mxn)),
            ("EBITDA Anual (MXN)",           float(r.ebitda_anual_mxn)),
        ]
        _hdr_row(ws4, ["Concepto", "Importe (MXN)"])
        for i, (k, v) in enumerate(cascada, 2):
            ws4.cell(i, 1, k)
            ws4.cell(i, 2, v).number_format = _FMT_MXN
        _autofit(ws4)

        # ── Hoja 5: Flujo 15 años ─────────────────────────────────────────────
        if r.inversion_mxn and r.inversion_mxn > 0:
            ws5 = wb.create_sheet("Flujo 15 años")
            _hdr_row(ws5, ["Año", "Flujo Anual (MXN)", "Flujo Acumulado (MXN)"])
            flujo_vals = calcular_flujo_acumulado(r.inversion_mxn, r.ebitda_anual_mxn)
            annual_vals = [-float(r.inversion_mxn)] + [float(r.ebitda_anual_mxn)] * 15
            for i, (anual, acum) in enumerate(zip(annual_vals, flujo_vals), 2):
                ws5.cell(i, 1, i - 2)
                ws5.cell(i, 2, anual).number_format = _FMT_MXN
                ws5.cell(i, 3, float(acum)).number_format = _FMT_MXN
            _autofit(ws5)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        slug = nombre_cliente.lower().replace(" ", "_")
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"cogeneracion_{slug}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/export/excel")
    def export_excel():
        import tempfile
        from flask import flash
        from reports.excel import generar_excel

        activo_id = session.get("cliente_activo_id")
        if activo_id is None:
            flash("Sin cliente activo. Selecciona un cliente primero.", "warning")
            return redirect(url_for("clientes.listado"))

        cfe_invoices, gas_invoices, _, _ = _cargar_facturas_seleccionadas(activo_id)
        r = calcular_cogen(cfe_invoices, gas_invoices, CoGenParams())
        if not r.meses:
            return "Sin datos para exportar (sin meses pareados).", 503

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = Path(f.name)
        generar_excel(r, tmp_path)
        return send_file(
            tmp_path,
            as_attachment=True,
            download_name="analisis_cogen.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/clientes/<int:cliente_id>/grafica/<grafica_id>/excel")
    def cliente_grafica_excel(cliente_id: int, grafica_id: str):
        """Descarga Excel con los datos crudos de la gráfica indicada."""
        import io
        import re
        from openpyxl import Workbook

        cliente, err = _verificar_cliente_activo(cliente_id)
        if err:
            return err

        _GRAFICAS = {
            "ahorro_neto_mensual":    "Cogeneración — Detalle Mensual",
            "demanda_por_horario":    "Demanda por Horario CFE",
            "consumo_por_horario":    "Consumo por Horario CFE",
            "costo_unitario_promedio": "Costos Detallados CFE",
            "composicion_costo":      "Indicadores CFE",
            "gas_consumo":            "Histórico Gas Natural",
            "gas_costos":             "Histórico Gas Natural",
        }
        if grafica_id not in _GRAFICAS:
            return "Gráfica no encontrada.", 404

        nombre_tabla = _GRAFICAS[grafica_id]
        cfe_invoices, gas_invoices, _, _ = _cargar_facturas_seleccionadas(cliente_id)

        wb = Workbook()
        ws = wb.active
        ws.title = nombre_tabla[:31]

        if grafica_id == "ahorro_neto_mensual":
            r = calcular_cogen(cfe_invoices, gas_invoices, CoGenParams())
            ws.append([
                "Periodo", "kWh Total", "Costo CFE (MXN)", "$/kWh Prom.",
                "GJ Consumido", "$/GJ Gas", "Costo Gas Actual (MXN)",
                "kWh Cubiertos", "GJ Gas Cogeneración", "Costo Gas Cogeneración (MXN)",
                "Ahorro Electricidad (MXN)", "Calor Recuperado (GJ)",
                "Ahorro Caldera (MXN)", "O&M (MXN)", "EBITDA Mensual (MXN)",
            ])
            for m in r.meses:
                ws.append([
                    m.periodo_inicio.strftime("%b %Y"),
                    float(m.kwh_total), float(m.costo_cfe_mxn),
                    float(m.costo_promedio_kwh), float(m.gj_consumido),
                    float(m.costo_unitario_gj), float(m.costo_gas_actual_mxn),
                    float(m.kwh_cubiertos), float(m.gj_gas_cogen),
                    float(m.costo_gas_cogen_mxn), float(m.ahorro_electricidad_mxn),
                    float(m.calor_recuperado_gj), float(m.ahorro_caldera_mxn),
                    float(m.gasto_om_mes_mxn), float(m.ebitda_mes_mxn),
                ])

        elif grafica_id in ("demanda_por_horario", "consumo_por_horario"):
            historico = calcular_historico_cfe(cfe_invoices)
            labels = historico["labels"]
            if grafica_id == "demanda_por_horario":
                ws.append(["Mes", "Demanda Punta (kW)", "Demanda Intermedia (kW)",
                            "Demanda Base (kW)", "Costo Unitario Prom. ($/kWh)"])
                for i, lbl in enumerate(labels):
                    ws.append([lbl, historico["demanda_punta"][i],
                                historico["demanda_intermedio"][i],
                                historico["demanda_base"][i],
                                historico["costo_unit_mes"][i]])
            else:
                ws.append(["Mes", "Consumo Punta (kWh)", "Consumo Intermedia (kWh)",
                            "Consumo Base (kWh)", "Costo Unitario Prom. ($/kWh)"])
                for i, lbl in enumerate(labels):
                    ws.append([lbl, historico["consumo_punta"][i],
                                historico["consumo_intermedio"][i],
                                historico["consumo_base"][i],
                                historico["costo_unit_mes"][i]])

        elif grafica_id == "costo_unitario_promedio":
            tablas = calcular_tablas_cfe(cfe_invoices)
            ws.append([
                "Mes", "CE Base (MXN)", "CE Intermedia (MXN)", "CE Punta (MXN)", "CE Total (MXN)",
                "Cargo Distribución (MXN)", "Cargo Capacidad (MXN)", "Total Demanda (MXN)",
                "CT Base (MXN)", "CT Intermedia (MXN)", "CT Punta (MXN)",
                "CU Base ($/kWh)", "CU Intermedia ($/kWh)", "CU Punta ($/kWh)",
                "Factor Potencia (MXN)", "Subtotal (MXN)",
            ])
            for f in tablas["costos_detallados"]:
                ws.append([
                    f["mes"], f["ce_base"], f["ce_inter"], f["ce_punta"], f["ce_total"],
                    f["costo_dist"], f["costo_cap"], f["costo_dem"],
                    f["ct_base"], f["ct_inter"], f["ct_punta"],
                    f["cu_base_total"], f["cu_inter_total"], f["cu_punta_total"],
                    f["cargo_fp"], f["subtotal"],
                ])

        elif grafica_id == "composicion_costo":
            tablas = calcular_tablas_cfe(cfe_invoices)
            ws.append(["Mes", "Costo Unitario ($/kWh)", "% Energía", "% Demanda",
                        "Factor Carga (%)", "Demanda Promedio (kW)"])
            for f in tablas["indicadores"]:
                ws.append([f["mes"], f["costo_unit"], f["pct_energia"],
                            f["pct_demanda"], f["factor_carga"], f["demanda_prom"]])

        elif grafica_id in ("gas_consumo", "gas_costos"):
            historico_gas = calcular_historico_gas(gas_invoices)
            if historico_gas is None:
                return "Sin datos de gas disponibles.", 404
            ws.append([
                "Mes", "Consumo (GJ)", "Molécula ($/GJ)", "Transporte ($/GJ)",
                "Costo Molécula (MXN)", "Costo Transporte (MXN)", "Costo Total (MXN)",
                "Costo Unitario ($/GJ)", "Costo Unitario ($/kWh)", "PCS (GJ/m³)", "PCS (kWh/m³)",
            ])
            for f in historico_gas["filas"]:
                ws.append([
                    f["mes"], f["consumo_gj"], f["molecula_precio_gj"],
                    f["transporte_precio_gj"], f["costo_molecula_mxn"],
                    f["costo_transporte_mxn"], f["costo_total_mxn"],
                    f["costo_unit_gj"], f["costo_unit_kwh"],
                    f["pcs_gj_m3"], f["pcs_kwh_m3"],
                ])
            tot = historico_gas["total"]
            ws.append([
                "TOTAL", tot["consumo_gj"], tot["molecula_precio_gj"],
                tot["transporte_precio_gj"], tot["costo_molecula_mxn"],
                tot["costo_transporte_mxn"], tot["costo_total_mxn"],
                tot["costo_unit_gj"], tot["costo_unit_kwh"],
                tot["pcs_gj_m3"], tot["pcs_kwh_m3"],
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nombre_cliente = cliente["nombre"]
        filename = re.sub(r'[\\/*?:"<>|]', "", f"{nombre_cliente} - {nombre_tabla}.xlsx")
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/admin/configuracion", methods=["GET", "POST"])
    def admin_configuracion():
        """Página de configuración global del sistema."""
        from decimal import Decimal, InvalidOperation

        # Validaciones específicas por clave: (min, max)
        _RANGOS = {
            "tipo_cambio_mxn_usd":                    (Decimal("10"),  Decimal("30")),
            "factor_emision_electricidad_kg_co2_kwh": (Decimal("0.1"), Decimal("2.0")),
            "factor_emision_gas_kg_co2_gj":           (Decimal("10"),  Decimal("200")),
        }

        if request.method == "POST":
            filas = list_configuracion()
            errores = []
            nuevos_valores = {}
            for fila in filas:
                clave = fila["clave"]
                raw = request.form.get(clave, "").strip()
                try:
                    val = Decimal(raw)
                    if val <= 0:
                        raise ValueError("no positivo")
                    if clave in _RANGOS:
                        lo, hi = _RANGOS[clave]
                        if val < lo or val > hi:
                            raise ValueError("fuera de rango")
                    nuevos_valores[clave] = str(val)
                except (InvalidOperation, ValueError):
                    desc = fila.get("descripcion") or clave
                    if clave in _RANGOS:
                        lo, hi = _RANGOS[clave]
                        errores.append(f"{desc}: valor inválido (rango {lo} – {hi}).")
                    else:
                        errores.append(f"{desc}: debe ser un número positivo.")

            claves_conocidas = {fila["clave"] for fila in filas}
            for clave_form in request.form:
                if clave_form not in claves_conocidas and not clave_form.startswith("csrf"):
                    logger.warning("admin_configuracion: clave desconocida en POST: %r", clave_form)

            if errores:
                for msg in errores:
                    flash(msg, "danger")
                return redirect(url_for("admin_configuracion"))

            for clave, valor in nuevos_valores.items():
                set_configuracion(clave, valor)
            flash("Configuración guardada correctamente.", "success")
            return redirect(url_for("admin_configuracion"))

        filas = list_configuracion()
        return render_template("admin/configuracion.html", filas=filas)

    @app.route("/changelog")
    def changelog():
        user = get_current_user()
        if not user or user.get("rol") not in ("admin", "master_admin"):
            abort(403)
        import markdown
        changelog_path = Path(__file__).resolve().parent.parent / "CHANGELOG.md"
        try:
            md_text = changelog_path.read_text(encoding="utf-8")
        except FileNotFoundError:
            md_text = "_Sin changelog disponible._"
        content = markdown.markdown(md_text, extensions=["nl2br"])
        return render_template("changelog.html", content=content)

    @app.route("/privacidad")
    def aviso_privacidad():
        return render_template("privacidad.html")

    @app.route("/healthz")
    def healthz():
        return "ok", 200

    @app.route("/health")
    def health():
        return "ok", 200

    @app.route("/admin/auditoria-logins", methods=["GET"])
    def admin_auditoria_logins():
        from web.auth import get_current_user as _gcu, ROL_MASTER_ADMIN, ROL_ADMIN
        user = _gcu()
        if not user or user["rol"] not in (ROL_MASTER_ADMIN, ROL_ADMIN):
            flash("Acceso no autorizado.", "danger")
            return redirect(url_for("dashboard"))
        from storage.repository import _supabase
        email_q = request.args.get("email", "").strip()
        success_q = request.args.get("success", "")
        q = _supabase.table("login_audit").select("*").order("created_at", desc=True).limit(100)
        if email_q:
            q = q.ilike("email", f"%{email_q}%")
        if success_q in ("1", "0"):
            q = q.eq("success", success_q == "1")
        registros = q.execute().data or []
        return render_template(
            "admin/auditoria_logins.html",
            registros=registros,
            email_q=email_q,
            success_q=success_q,
        )

    @app.route("/admin/usuarios", methods=["GET"])
    def admin_usuarios():
        from web.auth import get_current_user as _gcu, ROL_MASTER_ADMIN, ROL_ADMIN
        user = _gcu()
        if not user or user["rol"] not in (ROL_MASTER_ADMIN, ROL_ADMIN):
            flash("Acceso no autorizado.", "danger")
            return redirect(url_for("dashboard"))
        from storage.repository import _supabase, get_all_clientes_con_conteos
        # Obtener perfiles
        res = _supabase.table("user_profiles").select("*").order("email").execute()
        perfiles = res.data or []
        # admin no ve al master_admin
        if user["rol"] == ROL_ADMIN:
            perfiles = [p for p in perfiles if p.get("rol") != ROL_MASTER_ADMIN]
        # Enriquecer con nombre de empresa
        clientes_list = get_all_clientes_con_conteos()
        empresa_map = {c["id"]: c["nombre"] for c in clientes_list}
        for p in perfiles:
            p["empresa_nombre"] = empresa_map.get(p.get("empresa_id"), "")
        return render_template(
            "admin/usuarios.html",
            usuarios=perfiles,
            clientes=clientes_list,
        )

    @app.route("/admin/usuarios/crear", methods=["POST"])
    def admin_usuarios_crear():
        import secrets
        import string
        from web.auth import get_current_user as _gcu, ROL_MASTER_ADMIN, ROL_ADMIN
        user = _gcu()
        if not user or user["rol"] not in (ROL_MASTER_ADMIN, ROL_ADMIN):
            flash("Acceso no autorizado.", "danger")
            return redirect(url_for("dashboard"))

        from storage.repository import _supabase

        email = request.form.get("email", "").strip().lower()
        rol = request.form.get("rol", "").strip()
        password_input = request.form.get("password", "").strip()
        generar = request.form.get("generar_password") == "on"
        nombre_nuevo = request.form.get("nombre", "").strip() or None
        apellido_nuevo = request.form.get("apellido", "").strip() or None

        # Multi-cliente para usuario_normal
        if rol == "usuario_normal":
            cliente_ids_raw = request.form.getlist("cliente_ids")
            cliente_ids = [int(x) for x in cliente_ids_raw if x.isdigit()]
            empresa_id = cliente_ids[0] if len(cliente_ids) == 1 else None
        else:
            cliente_ids = []
            empresa_id = None

        # Validaciones
        if not email or "@" not in email:
            flash("Email inválido.", "danger")
            return redirect(url_for("admin_usuarios"))
        if rol not in ("admin", "usuario_normal"):
            flash("Rol no válido.", "danger")
            return redirect(url_for("admin_usuarios"))
        # admin solo puede crear usuario_normal
        if user["rol"] == ROL_ADMIN and rol != "usuario_normal":
            flash("Solo puedes crear usuarios de tipo Cliente.", "danger")
            return redirect(url_for("admin_usuarios"))

        # Verificar duplicado
        try:
            _supabase.postgrest.auth(os.environ["SUPABASE_KEY"])
            dup = _supabase.table("user_profiles").select("id").eq("email", email).limit(1).execute()
            if dup.data:
                flash(f"El email {email} ya está registrado.", "danger")
                return redirect(url_for("admin_usuarios"))
        except Exception:
            pass

        # Determinar contraseña
        if generar or not password_input:
            alfabeto = string.ascii_letters + string.digits + "!@#$%^&*"
            password = "".join(secrets.choice(alfabeto) for _ in range(12))
            password_generada = True
        else:
            if len(password_input) < 8:
                flash("La contraseña debe tener al menos 8 caracteres.", "danger")
                return redirect(url_for("admin_usuarios"))
            password = password_input
            password_generada = False

        # Crear usuario en Supabase Auth
        try:
            res = _supabase.auth.admin.create_user({
                "email": email,
                "password": password,
                "email_confirm": True,
            })
            user_id = res.user.id
        except Exception as exc:
            logger.error("Error creando usuario auth %s: %s", email, exc)
            flash(f"Error creando usuario: {exc}", "danger")
            return redirect(url_for("admin_usuarios"))

        # Crear perfil
        try:
            _supabase.postgrest.auth(os.environ["SUPABASE_KEY"])
            _supabase.table("user_profiles").insert({
                "id": user_id,
                "email": email,
                "rol": rol,
                "empresa_id": empresa_id,
                "activo": True,
                "nombre": nombre_nuevo,
                "apellido": apellido_nuevo,
            }).execute()
        except Exception as exc:
            logger.error("Error creando perfil %s: %s", email, exc)
            try:
                _supabase.auth.admin.delete_user(user_id)
            except Exception:
                pass
            flash(f"Error creando perfil: {exc}", "danger")
            return redirect(url_for("admin_usuarios"))

        # Asignar clientes en usuario_clientes (para usuario_normal)
        if rol == "usuario_normal" and cliente_ids:
            try:
                from storage.repository import set_clientes_de_usuario as _scdu
                _scdu(user_id, cliente_ids)
            except Exception as exc:
                logger.error("Error asignando clientes a usuario %s: %s", user_id, exc)
                flash("Usuario creado, pero hubo un error asignando los clientes. Edita el usuario para corregirlo.", "warning")

        if password_generada:
            flash(
                f"Usuario {email} creado. Contraseña generada: {password}. "
                f"Copia esta contraseña ahora — no se mostrará de nuevo.",
                "password_generada",
            )
        else:
            flash(f"Usuario {email} creado correctamente.", "success")
        return redirect(url_for("admin_usuarios"))

    @app.route("/admin/usuarios/<user_id>/cambiar-password", methods=["POST"])
    def admin_usuarios_cambiar_password(user_id: str):
        import secrets
        import string
        from web.auth import get_current_user as _gcu, ROL_MASTER_ADMIN, ROL_ADMIN

        actor = _gcu()
        if not actor or actor["rol"] not in (ROL_MASTER_ADMIN, ROL_ADMIN):
            flash("Acceso no autorizado.", "danger")
            return redirect(url_for("dashboard"))

        from storage.repository import _supabase
        try:
            _supabase.postgrest.auth(os.environ["SUPABASE_KEY"])
            res = _supabase.table("user_profiles").select("*").eq("id", user_id).limit(1).execute()
            target = res.data[0] if res.data else None
        except Exception:
            target = None

        if not target:
            flash("Usuario no encontrado.", "warning")
            return redirect(url_for("admin_usuarios"))

        # Validar matriz de permisos
        if target["rol"] == ROL_MASTER_ADMIN and target["id"] != actor["user_id"]:
            flash("No puedes cambiar la contraseña del Master Admin.", "danger")
            return redirect(url_for("admin_usuarios"))
        if target["rol"] == ROL_ADMIN and actor["rol"] == ROL_ADMIN and target["id"] != actor["user_id"]:
            flash("No puedes cambiar la contraseña de otro Admin.", "danger")
            return redirect(url_for("admin_usuarios"))

        nueva_password = request.form.get("password", "").strip()
        generar = request.form.get("generar_password") == "on"

        if generar or not nueva_password:
            alfabeto = string.ascii_letters + string.digits + "!@#$%^&*"
            nueva_password = "".join(secrets.choice(alfabeto) for _ in range(12))
            password_generada = True
        else:
            if len(nueva_password) < 8:
                flash("Contraseña mínimo 8 caracteres.", "danger")
                return redirect(url_for("admin_usuarios"))
            password_generada = False

        try:
            _supabase.auth.admin.update_user_by_id(user_id, {"password": nueva_password})
        except Exception as exc:
            logger.error("Error cambiando password %s: %s", user_id, exc)
            flash(f"Error cambiando contraseña: {exc}", "danger")
            return redirect(url_for("admin_usuarios"))

        from storage.repository import incrementar_session_version
        incrementar_session_version(user_id)

        if password_generada:
            flash(
                f"Contraseña actualizada para {target['email']}. Nueva contraseña: {nueva_password}. "
                f"Copia esta contraseña ahora — no se mostrará de nuevo.",
                "password_generada",
            )
        else:
            flash(f"Contraseña actualizada para {target['email']}.", "success")
        return redirect(url_for("admin_usuarios"))

    @app.route("/admin/usuarios/<user_id>/editar", methods=["GET", "POST"])
    def admin_usuarios_editar(user_id: str):
        from web.auth import get_current_user as _gcu, ROL_MASTER_ADMIN, ROL_ADMIN
        actor = _gcu()
        if not actor or actor["rol"] not in (ROL_MASTER_ADMIN, ROL_ADMIN):
            flash("Acceso no autorizado.", "danger")
            return redirect(url_for("dashboard"))
        from storage.repository import _supabase, get_all_clientes_con_conteos
        try:
            _supabase.postgrest.auth(os.environ["SUPABASE_KEY"])
            res = _supabase.table("user_profiles").select("*").eq("id", user_id).limit(1).execute()
            target = res.data[0] if res.data else None
        except Exception:
            target = None
        if not target:
            flash("Usuario no encontrado.", "warning")
            return redirect(url_for("admin_usuarios"))
        if target["rol"] == ROL_MASTER_ADMIN:
            flash("No se puede editar al Master Admin.", "warning")
            return redirect(url_for("admin_usuarios"))
        # admin no puede editar a otro admin (solo a sí mismo o a usuario_normal)
        if actor["rol"] == ROL_ADMIN and target["rol"] == ROL_ADMIN and target["id"] != actor["user_id"]:
            flash("No puedes editar a otro Administrador.", "danger")
            return redirect(url_for("admin_usuarios"))

        # admin no puede cambiar el rol del target
        actor_puede_cambiar_rol = (actor["rol"] == ROL_MASTER_ADMIN)

        if request.method == "POST":
            if actor_puede_cambiar_rol:
                rol = request.form.get("rol", "").strip()
                if rol not in ("admin", "usuario_normal"):
                    flash("Rol no válido.", "danger")
                    return redirect(url_for("admin_usuarios_editar", user_id=user_id))
            else:
                rol = target["rol"]  # mantener el rol actual
            nombre_ed = request.form.get("nombre", "").strip() or None
            apellido_ed = request.form.get("apellido", "").strip() or None
            if rol == "usuario_normal":
                cliente_ids_raw = request.form.getlist("cliente_ids")
                cliente_ids = [int(x) for x in cliente_ids_raw if x.isdigit()]
                # empresa_id legacy: apuntar al único si hay exactamente uno
                empresa_id_legacy = cliente_ids[0] if len(cliente_ids) == 1 else None
            else:
                cliente_ids = []
                empresa_id_legacy = None
            try:
                _supabase.postgrest.auth(os.environ["SUPABASE_KEY"])
                _supabase.table("user_profiles").update({
                    "rol": rol,
                    "empresa_id": empresa_id_legacy,
                    "nombre": nombre_ed,
                    "apellido": apellido_ed,
                }).eq("id", user_id).execute()
                flash(f"Usuario {target['email']} actualizado correctamente.", "success")
            except Exception as exc:
                logger.error("Error actualizando usuario %s: %s", user_id, exc)
                flash(f"Error actualizando usuario: {exc}", "danger")
                return redirect(url_for("admin_usuarios"))

            # Siempre actualizar asignaciones en usuario_clientes
            # Para rol admin: lista vacía (limpia huérfanos)
            # Para usuario_normal: lista de checkboxes seleccionados
            try:
                from storage.repository import set_clientes_de_usuario as _scdu
                _scdu(user_id, cliente_ids)
            except Exception as exc:
                logger.error("Error actualizando clientes de usuario %s: %s", user_id, exc)
                flash("Los datos del usuario se guardaron, pero hubo un error actualizando los clientes asignados.", "warning")

            return redirect(url_for("admin_usuarios"))

        clientes_list = get_all_clientes_con_conteos()
        from storage.repository import get_clientes_de_usuario as _gcdu
        clientes_asignados = _gcdu(user_id)
        clientes_asignados_ids = [c["id"] for c in clientes_asignados]
        return render_template("admin/editar_usuario.html",
                               target=target, clientes=clientes_list,
                               clientes_asignados_ids=clientes_asignados_ids,
                               form_rol=target["rol"],
                               form_empresa_id=target.get("empresa_id"),
                               form_nombre=target.get("nombre"),
                               form_apellido=target.get("apellido"),
                               actor_puede_cambiar_rol=actor_puede_cambiar_rol)

    @app.route("/admin/usuarios/<user_id>/borrar", methods=["POST"])
    def admin_usuarios_borrar(user_id: str):
        from web.auth import get_current_user as _gcu, ROL_MASTER_ADMIN, ROL_ADMIN
        from web.auth_permissions import validar_borrar_usuario
        actor = _gcu()
        if not actor or actor["rol"] not in (ROL_MASTER_ADMIN, ROL_ADMIN):
            flash("Acceso restringido.", "danger")
            return redirect(url_for("dashboard"))
        from storage.repository import _supabase
        try:
            res = _supabase.table("user_profiles").select("id,email,rol").eq("id", user_id).maybe_single().execute()
            target = res.data
        except Exception:
            target = None
        if not target:
            flash("Usuario no encontrado.", "warning")
            return redirect(url_for("admin_usuarios"))
        err = validar_borrar_usuario(actor, target)
        if err:
            flash(err, "danger")
            return redirect(url_for("admin_usuarios"))
        try:
            _supabase.auth.admin.delete_user(user_id)
            flash(f"Usuario {target['email']} eliminado.", "success")
        except Exception as exc:
            logger.error("Error borrando usuario %s: %s", user_id, exc)
            flash(f"Error al borrar el usuario: {exc}", "danger")
        return redirect(url_for("admin_usuarios"))

    @app.route("/admin/usuarios/<user_id>/desactivar", methods=["POST"])
    def admin_usuarios_desactivar(user_id: str):
        from web.auth import get_current_user as _gcu, ROL_MASTER_ADMIN, ROL_ADMIN
        actor = _gcu()
        if not actor or actor["rol"] not in (ROL_MASTER_ADMIN, ROL_ADMIN):
            flash("Acceso restringido.", "danger")
            return redirect(url_for("dashboard"))
        from storage.repository import _supabase
        try:
            res = _supabase.table("user_profiles").select("activo,email,rol").eq("id", user_id).maybe_single().execute()
            perfil = res.data
            if not perfil:
                flash("Usuario no encontrado.", "warning")
                return redirect(url_for("admin_usuarios"))
            if perfil.get("rol") == ROL_MASTER_ADMIN:
                flash("No se puede modificar al Master Admin.", "danger")
                return redirect(url_for("admin_usuarios"))
            if actor["rol"] == ROL_ADMIN and perfil.get("rol") == ROL_ADMIN:
                flash("No puedes desactivar a otro Administrador.", "danger")
                return redirect(url_for("admin_usuarios"))
            nuevo_activo = not perfil.get("activo", True)
            _supabase.table("user_profiles").update({"activo": nuevo_activo}).eq("id", user_id).execute()
            accion = "activado" if nuevo_activo else "desactivado"
            flash(f"Usuario {perfil['email']} {accion}.", "success")
        except Exception as exc:
            logger.error("Error desactivando usuario %s: %s", user_id, exc)
            flash(f"Error al actualizar el usuario: {exc}", "danger")
        return redirect(url_for("admin_usuarios"))

    @app.route("/mi-perfil")
    def mi_perfil():
        from web.auth import get_current_user as _gcu, is_authenticated as _ia
        if not _ia():
            return redirect(url_for("auth.login"))
        user = _gcu()
        from storage.repository import _supabase
        empresa_nombre = None
        if user and user.get("empresa_id"):
            try:
                _supabase.postgrest.auth(os.environ["SUPABASE_KEY"])
                res = _supabase.table("clientes").select("nombre").eq("id", user["empresa_id"]).limit(1).execute()
                if res.data:
                    empresa_nombre = res.data[0]["nombre"]
            except Exception:
                pass
        return render_template("mi_perfil.html", user=user, empresa_nombre=empresa_nombre)

    @app.route("/mi-perfil/cambiar-password", methods=["POST"])
    def mi_perfil_cambiar_password():
        from web.auth import get_current_user as _gcu, is_authenticated as _ia
        if not _ia():
            return redirect(url_for("auth.login"))
        user = _gcu()
        nueva = request.form.get("password", "").strip()
        confirmar = request.form.get("confirmar", "").strip()
        if len(nueva) < 8:
            flash("La contraseña debe tener al menos 8 caracteres.", "danger")
            return redirect(url_for("mi_perfil"))
        if nueva != confirmar:
            flash("Las contraseñas no coinciden.", "danger")
            return redirect(url_for("mi_perfil"))
        from storage.repository import _supabase
        try:
            _supabase.auth.admin.update_user_by_id(user["user_id"], {"password": nueva})
            from storage.repository import incrementar_session_version, get_session_version
            incrementar_session_version(user["user_id"])
            # Refrescar la versión en la sesión actual para que el usuario no se desloguee a sí mismo
            session["_session_version"] = get_session_version(user["user_id"]) or 1
            session.pop("_sv_check", None)
            flash("Contraseña actualizada correctamente.", "success")
        except Exception as exc:
            flash(f"Error: {exc}", "danger")
        return redirect(url_for("mi_perfil"))

    @app.route("/mi-perfil/cambiar-datos", methods=["POST"])
    def mi_perfil_cambiar_datos():
        from web.auth import get_current_user as _gcu, is_authenticated as _ia
        if not _ia():
            return redirect(url_for("auth.login"))
        user = _gcu()
        nombre_val = request.form.get("nombre", "").strip() or None
        apellido_val = request.form.get("apellido", "").strip() or None
        from storage.repository import _supabase
        try:
            _supabase.postgrest.auth(os.environ["SUPABASE_KEY"])
            _supabase.table("user_profiles").update({
                "nombre": nombre_val,
                "apellido": apellido_val,
            }).eq("id", user["user_id"]).execute()
            from flask import session as _sess
            _sess["_nombre"] = nombre_val
            _sess["_apellido"] = apellido_val
            flash("Datos actualizados correctamente.", "success")
        except Exception as exc:
            flash(f"Error: {exc}", "danger")
        return redirect(url_for("mi_perfil"))

    @app.route("/upload", methods=["POST"])
    def upload_facturas():
        from flask import jsonify
        return jsonify({
            "error": "Este endpoint fue eliminado. Usa /clientes/<id>/contratos/<id>/upload."
        }), 410

    # ── Error handlers ────────────────────────────────────────────────────────

    @app.errorhandler(403)
    def handle_403(e):
        log_error("error_403", str(e), codigo_http=403)
        return render_template("error.html", codigo=403, mensaje="Acceso no autorizado."), 403

    @app.errorhandler(404)
    def handle_404(e):
        log_error("error_404", str(e), codigo_http=404)
        return render_template("error.html", codigo=404, mensaje="Página no encontrada."), 404

    @app.errorhandler(Exception)
    def handle_500(e):
        import traceback
        log_error("error_500", str(e), exc=e, codigo_http=500)
        app.logger.error("Unhandled exception: %s", traceback.format_exc())
        return render_template("error.html", codigo=500, mensaje="Error interno del servidor."), 500

    # ── Ruta de registro de errores ────────────────────────────────────────────

    @app.route("/admin/errores")
    def admin_errores():
        from web.auth import get_current_user as _gcu, ROL_MASTER_ADMIN
        from storage.repository import _supabase
        import os
        import math

        actor = _gcu()
        if not actor or actor["rol"] != ROL_MASTER_ADMIN:
            flash("Acceso restringido a Super Admin.", "danger")
            return redirect(url_for("dashboard"))

        POR_PAGINA = 50
        NIVELES = ["error_500", "error_403", "error_404", "validacion", "negocio"]

        filtros = {
            "nivel":       request.args.get("nivel", "").strip() or None,
            "email":       request.args.get("email", "").strip() or None,
            "ruta":        request.args.get("ruta", "").strip() or None,
            "fecha_desde": request.args.get("fecha_desde", "").strip() or None,
            "fecha_hasta": request.args.get("fecha_hasta", "").strip() or None,
        }
        pagina = max(1, int(request.args.get("pagina", 1)))

        try:
            _supabase.postgrest.auth(os.environ["SUPABASE_KEY"])
            q = _supabase.table("error_logs").select("*", count="exact")

            if filtros["nivel"]:
                q = q.eq("nivel", filtros["nivel"])
            if filtros["email"]:
                q = q.ilike("usuario_email", f"%{filtros['email']}%")
            if filtros["ruta"]:
                q = q.ilike("ruta", f"%{filtros['ruta']}%")
            if filtros["fecha_desde"]:
                q = q.gte("created_at", filtros["fecha_desde"])
            if filtros["fecha_hasta"]:
                q = q.lte("created_at", filtros["fecha_hasta"] + "T23:59:59")

            offset = (pagina - 1) * POR_PAGINA
            res = q.order("created_at", desc=True).range(offset, offset + POR_PAGINA - 1).execute()
            registros = res.data or []
            total = res.count or 0
            total_paginas = max(1, math.ceil(total / POR_PAGINA))
        except Exception as exc:
            logger.error("Error consultando error_logs: %s", exc)
            registros, total, total_paginas = [], 0, 1

        return render_template(
            "admin/errores.html",
            registros=registros,
            total=total,
            pagina=pagina,
            total_paginas=total_paginas,
            niveles=NIVELES,
            filtros={k: v for k, v in filtros.items() if v},
            nav_active="errores",
        )

    # ── FASE 2: Telemetría ────────────────────────────────────────────────────

    @app.route("/admin/telemetria")
    def telemetria_index():
        if not app.config.get("FASE2_HABILITADA", False):
            abort(404)
        from web.auth import get_current_user as _gcu
        actor = _gcu()
        if not actor or actor["rol"] != "master_admin":
            flash("Acceso restringido.", "danger")
            return redirect(url_for("dashboard"))
        from storage.repository import get_all_clientes_con_conteos, obtener_medidores_por_cliente
        clientes = get_all_clientes_con_conteos()
        cliente_id = request.args.get("cliente_id", type=int)
        cliente_sel = None
        medidores = []
        if cliente_id:
            cliente_sel = next((c for c in clientes if c["id"] == cliente_id), None)
            if cliente_sel:
                medidores = obtener_medidores_por_cliente(cliente_id)
        return render_template(
            "telemetria/index.html",
            clientes=clientes,
            cliente_sel=cliente_sel,
            medidores=medidores,
            nav_active="telemetria",
        )

    @app.route("/admin/telemetria/medidor/<int:medidor_id>")
    def telemetria_medidor(medidor_id):
        if not app.config.get("FASE2_HABILITADA", False):
            abort(404)
        from web.auth import get_current_user as _gcu
        actor = _gcu()
        if not actor or actor["rol"] != "master_admin":
            flash("Acceso restringido.", "danger")
            return redirect(url_for("dashboard"))
        from storage.repository import obtener_medidor, obtener_mediciones_recientes
        from datetime import datetime, timedelta, timezone
        medidor = obtener_medidor(medidor_id)
        if medidor is None:
            flash("Medidor no encontrado.", "warning")
            return redirect(url_for("telemetria_index"))
        hasta = datetime.now(timezone.utc)
        desde = hasta - timedelta(hours=24)
        todas = obtener_mediciones_recientes(
            medidor_id,
            desde=desde.isoformat(),
            hasta=hasta.isoformat(),
        )
        mediciones = todas[-200:] if len(todas) > 200 else todas
        return render_template(
            "telemetria/medidor.html",
            medidor=medidor,
            mediciones=mediciones,
            nav_active="telemetria",
        )

    @app.route("/admin/telemetria/medidor/<int:medidor_id>/sembrar", methods=["POST"])
    def telemetria_sembrar(medidor_id):
        if not app.config.get("FASE2_HABILITADA", False):
            abort(404)
        from web.auth import get_current_user as _gcu
        actor = _gcu()
        if not actor or actor["rol"] != "master_admin":
            flash("Acceso restringido.", "danger")
            return redirect(url_for("dashboard"))
        from storage.repository import obtener_medidor, insertar_mediciones_batch
        from datetime import datetime, timedelta, timezone
        from telemetria.seed import generar_mediciones_sinteticas
        medidor = obtener_medidor(medidor_id)
        if medidor is None:
            flash("Medidor no encontrado.", "warning")
            return redirect(url_for("telemetria_index"))
        hasta = datetime.now(timezone.utc)
        desde = hasta - timedelta(hours=24)
        mediciones = generar_mediciones_sinteticas(medidor_id, desde_utc=desde)
        total = insertar_mediciones_batch(mediciones)
        flash(f"Sembrado correcto: {total} mediciones insertadas.", "success")
        return redirect(url_for("telemetria_medidor", medidor_id=medidor_id))

    # ── Dashboard Modelado CHP ────────────────────────────────────────────────

    @app.route("/clientes/<int:cliente_id>/dashboard/modelado-chp")
    def cliente_dashboard_modelado_chp(cliente_id: int):
        """Vista del dashboard de Modelado CHP con medición cincominutal."""
        cliente, err = _verificar_cliente_activo(cliente_id)
        if err:
            return err

        mediciones = get_mediciones_por_cliente(cliente_id)
        if not mediciones:
            flash("Este cliente no tiene mediciones cincominutal cargadas.", "warning")
            return redirect(url_for("cliente_dashboard_contabilidad", cliente_id=cliente_id))

        import json as _json_app
        chp_params = get_cliente_chp_params(cliente_id)
        from storage.repository import get_chp_session_params as _get_chp_session_params
        chp_session_params = _get_chp_session_params(cliente_id)

        medicion_activa_id = session.get("medicion_activa_id") or mediciones[0]["id"]
        medicion_activa = next(
            (m for m in mediciones if m["id"] == medicion_activa_id),
            mediciones[0],
        )

        return render_template(
            "dashboard_modelado_chp.html",
            cliente_id=cliente_id,
            cliente_nombre=cliente["nombre"],
            logo_url=obtener_logo_cliente(cliente),
            mediciones=mediciones,
            medicion_activa=medicion_activa,
            chp_params=chp_params,
            motores_config_json=_json_app.dumps(chp_params.get("motores_config") or "null"),
            chp_session_params=chp_session_params or {},
            nav_active="modelado_chp",
        )

    # ── Dashboard Telemetría (Fase 2 D2) ─────────────────────────────────────

    @app.route("/clientes/<int:cliente_id>/dashboard/telemetria")
    def cliente_dashboard_telemetria(cliente_id: int):
        """Vista de Telemetría: árbol de medidores del cliente."""
        if not app.config.get("FASE2_HABILITADA", False):
            abort(404)
        cliente, err = _verificar_cliente_activo(cliente_id)
        if err:
            return err

        from storage.repository import obtener_arbol_medidores as _oam
        arbol_medidores = _oam(cliente_id)

        return render_template(
            "telemetria/dashboard.html",
            cliente=cliente,
            arbol_medidores=arbol_medidores,
            nav_active="telemetria_cliente",
        )

    @app.route("/clientes/<int:cliente_id>/dashboard/telemetria/data")
    def cliente_dashboard_telemetria_data(cliente_id: int):
        """JSON para el dashboard de telemetría: sunburst, serie temporal y KPIs."""
        from flask import jsonify
        if not app.config.get("FASE2_HABILITADA", False):
            abort(404)
        cliente, err = _verificar_cliente_activo(cliente_id)
        if err:
            return jsonify({"error": "acceso denegado"}), 403

        from storage.repository import (
            obtener_arbol_medidores as _oam,
            obtener_descendientes_ids as _odi,
            obtener_mediciones_para_rango as _omfr,
        )
        from calc.telemetria_kpis import determinar_periodo_anterior as _dpa
        from datetime import datetime, timedelta, timezone

        nodo_id_raw = request.args.get("nodo_id")
        rango = request.args.get("rango", "24h")

        # Cargar árbol completo
        todos = _oam(cliente_id)
        if not todos:
            return jsonify({"error": "sin_medidores"}), 404

        # Indexar por id
        por_id = {m["id"]: m for m in todos}

        # Acometida raíz: primer medidor sin padre (punto_medicion == 'acometida_cfe')
        acometida = next(
            (m for m in todos if m.get("punto_medicion") == "acometida_cfe"),
            todos[0]
        )

        # Calcular ruta de breadcrumbs (hacia arriba), inyectando subestaciones virtuales
        def _breadcrumbs(nodo_dict):
            ruta = []
            cur = nodo_dict
            while cur:
                ruta.append({"id": cur["id"], "nombre": cur["nombre"]})
                padre_id = cur.get("medidor_padre_id")
                cur = por_id.get(padre_id) if padre_id else None
            ruta = list(reversed(ruta))
            # Inyectar nodo virtual SE-N entre el padre y cada transformador T-N.*
            result = []
            for idx, seg in enumerate(ruta):
                result.append(seg)
                if idx + 1 < len(ruta):
                    nxt_id = ruta[idx + 1]["id"]
                    nxt = por_id.get(nxt_id, {})
                    nxt_nombre = nxt.get("nombre", "")
                    if (nxt_nombre.startswith("T-") and "." in nxt_nombre
                            and nxt.get("medidor_padre_id") == seg["id"]):
                        se_num = nxt_nombre.split("-")[1].split(".")[0]
                        result.append({"id": f"grupo:SE-{se_num}", "nombre": f"SE-{se_num}"})
            return result

        # --- Nodo virtual de subestación ---
        # El frontend envía "grupo:SE-N" para subestaciones que no existen como medidor.
        # Se agrega sobre sus transformadores hijo (T-N.*) y las cargas_final de estos.
        _nodo_virtual = None  # dict con id, nombre, punto_medicion, ruta_breadcrumbs si es virtual

        if nodo_id_raw and nodo_id_raw.startswith("grupo:"):
            codigo_se = nodo_id_raw[len("grupo:"):]  # ej. "SE-4"
            # Número de SE: "SE-4" → "4"
            se_num = codigo_se.split("-")[-1] if "-" in codigo_se else codigo_se
            prefijo_tx = f"T-{se_num}."
            # Transformadores de esta SE
            txs = [m for m in todos if (m.get("nombre") or "").startswith(prefijo_tx)]
            hojas_ids_nodo = []
            for tx in txs:
                desc_ids = _odi(tx["id"])
                hojas_ids_nodo += [
                    mid for mid in desc_ids
                    if por_id.get(mid, {}).get("punto_medicion") == "carga_final"
                ]
            if not hojas_ids_nodo and txs:
                hojas_ids_nodo = [txs[0]["id"]]
            elif not hojas_ids_nodo:
                hojas_ids_nodo = [acometida["id"]]
            _nodo_virtual = {
                "id": nodo_id_raw,
                "nombre": codigo_se,
                "punto_medicion": "subestacion_virtual",
                "ruta_breadcrumbs": [
                    {"id": acometida["id"], "nombre": acometida["nombre"]},
                    {"id": nodo_id_raw, "nombre": codigo_se},
                ],
            }
            nodo = acometida  # referencia interna no usada en JSON cuando hay _nodo_virtual
        else:
            nodo_id = int(nodo_id_raw) if nodo_id_raw else None
            if nodo_id is None:
                nodo_id = acometida["id"]
            nodo = por_id.get(nodo_id, acometida)

            # Hojas del nodo seleccionado: determinan KPIs, serie y comparativa
            if nodo.get("punto_medicion") == "carga_final":
                hojas_ids_nodo = [nodo_id]
            else:
                desc_ids = _odi(nodo_id)
                hojas_ids_nodo = [
                    mid for mid in desc_ids
                    if por_id.get(mid, {}).get("punto_medicion") == "carga_final"
                ]
                if not hojas_ids_nodo:
                    hojas_ids_nodo = [nodo_id]

        # Todas las hojas del árbol completo: necesarias para que el sunburst
        # muestre energía correcta en todos los nodos, independientemente del
        # nodo seleccionado.
        todas_hojas_ids = [m["id"] for m in todos if m.get("punto_medicion") == "carga_final"]
        if not todas_hojas_ids:
            todas_hojas_ids = hojas_ids_nodo

        # Calcular ventana temporal.
        # DEUDA TÉCNICA: usa max(timestamp) como ancla para que la demo con datos
        # sintéticos siempre muestre información sin re-seeds periódicos.
        # Revertir a datetime.now(timezone.utc) cuando entren medidores físicos con MQTT.
        from storage.repository import obtener_ultimo_timestamp_cliente as _outc
        _ts_max = _outc(cliente_id)
        ahora = _ts_max if _ts_max is not None else datetime.now(timezone.utc)
        modo_temporal = "sintetico" if _ts_max is not None else "tiempo_real"
        if rango == "7d":
            desde = ahora - timedelta(days=7)
        elif rango == "30d":
            desde = ahora - timedelta(days=30)
        else:
            desde = ahora - timedelta(hours=24)

        desde_iso = desde.strftime("%Y-%m-%dT%H:%M:%SZ")
        hasta_iso = ahora.strftime("%Y-%m-%dT%H:%M:%SZ")

        # Calcular periodo anterior antes del fetch paralelo
        desde_ant, hasta_ant, etiqueta_ant = _dpa(rango, ahora)
        desde_ant_iso = desde_ant.strftime("%Y-%m-%dT%H:%M:%SZ")
        hasta_ant_iso = hasta_ant.strftime("%Y-%m-%dT%H:%M:%SZ")

        # Fetch secuencial por medidor: actual + anterior en la misma iteración.
        # Serializado deliberadamente para no saturar el pool de sockets de Supabase
        # (Errno 11 EAGAIN bajo carga concurrente en Render free tier).
        todas_hojas_set = set(todas_hojas_ids)
        hojas_nodo_set = set(hojas_ids_nodo)
        all_medidores = list(todas_hojas_set | hojas_nodo_set)

        def _fmt_rows(rows):
            return [
                {
                    "ts": r["timestamp"],
                    "kw": float(r.get("potencia_activa_kw") or 0),
                    "fp": float(r.get("factor_potencia") or 0),
                }
                for r in rows
            ]

        mediciones_por_hoja = {}
        mediciones_ant = {}
        for hid in all_medidores:
            if hid in todas_hojas_set:
                mediciones_por_hoja[hid] = _fmt_rows(_omfr(hid, desde_iso, hasta_iso, rango))
            if hid in hojas_nodo_set:
                mediciones_ant[hid] = _fmt_rows(_omfr(hid, desde_ant_iso, hasta_ant_iso, rango))

        # Diagnóstico: contar filas por hoja para detectar ventanas vacías
        _n_filas_total = sum(len(v) for v in mediciones_por_hoja.values())
        # Si todas las hojas están vacías, el sunburst mostrará 0 kWh
        # (datos fuera del rango de fechas — re-ejecutar seed_iberica.py --forzar)

        # Garantizar que hojas_ids_nodo esté cubierto en mediciones_por_hoja
        # (edge case: nodo seleccionado es un tx sin cargas hijo — fallback a nodo_id
        # que no es carga_final y por tanto no está en todas_hojas_ids)
        ids_sin_datos = [hid for hid in hojas_ids_nodo if hid not in mediciones_por_hoja]
        for hid in ids_sin_datos:
            rows = _omfr(hid, desde_iso, hasta_iso, rango)
            mediciones_por_hoja[hid] = [
                {
                    "ts": r["timestamp"],
                    "kw": float(r.get("potencia_activa_kw") or 0),
                    "fp": float(r.get("factor_potencia") or 0),
                }
                for r in rows
            ]

        # Agregar serie temporal: sumar kW solo de las hojas del nodo seleccionado
        from collections import defaultdict
        bucket_kw = defaultdict(float)
        bucket_fp_peso = defaultdict(float)
        bucket_kw_peso = defaultdict(float)
        for hid in hojas_ids_nodo:
            for r in mediciones_por_hoja.get(hid, []):
                bucket_kw[r["ts"]] += r["kw"]
                bucket_kw_peso[r["ts"]] += r["kw"]
                bucket_fp_peso[r["ts"]] += r["fp"] * r["kw"]

        ts_sorted = sorted(bucket_kw.keys())
        potencia_serie = [round(bucket_kw[ts], 3) for ts in ts_sorted]

        # KPIs
        num_muestras = len(ts_sorted)
        demanda_pico = max(potencia_serie) if potencia_serie else 0.0

        # Integral trapezoidal para energía (kW × h)
        energia_kwh = 0.0
        if len(ts_sorted) >= 2:
            from datetime import datetime as _dt
            for i in range(1, len(ts_sorted)):
                try:
                    t0 = _dt.fromisoformat(ts_sorted[i-1].replace("Z", "+00:00"))
                    t1 = _dt.fromisoformat(ts_sorted[i].replace("Z", "+00:00"))
                    dt_h = (t1 - t0).total_seconds() / 3600.0
                    energia_kwh += (potencia_serie[i-1] + potencia_serie[i]) / 2.0 * dt_h
                except Exception:
                    pass

        # FP promedio ponderado
        total_peso_kw = sum(bucket_kw_peso[ts] for ts in ts_sorted)
        fp_prom = (
            sum(bucket_fp_peso[ts] for ts in ts_sorted) / total_peso_kw
            if total_peso_kw > 0 else 0.0
        )

        # Estructura sunburst: reconstruir árbol desde todos los medidores
        def _energia_nodo(mid):
            """Suma de kWh de las hojas descendientes del nodo mid."""
            if por_id.get(mid, {}).get("punto_medicion") == "carga_final":
                rows = mediciones_por_hoja.get(mid, [])
                kwh = 0.0
                for i in range(1, len(rows)):
                    try:
                        from datetime import datetime as _dt2
                        t0 = _dt2.fromisoformat(rows[i-1]["ts"].replace("Z", "+00:00"))
                        t1 = _dt2.fromisoformat(rows[i]["ts"].replace("Z", "+00:00"))
                        dt_h = (t1 - t0).total_seconds() / 3600.0
                        kwh += (rows[i-1]["kw"] + rows[i]["kw"]) / 2.0 * dt_h
                    except Exception:
                        pass
                return round(kwh, 3)
            else:
                hijos_ids = [m["id"] for m in todos if m.get("medidor_padre_id") == mid]
                return round(sum(_energia_nodo(h) for h in hijos_ids), 3)

        # ── Costo del periodo actual ───────────────────────────────────────
        from calc.telemetria_kpis import (
            atribuir_produccion_a_nodo as _apn,
            calcular_baseline_movil as _cbm,
            calcular_kpis_economicos as _cke,
            calcular_kpis_energeticos as _cken,
            calcular_kpis_produccion as _ckp,
            generar_sparkline as _gs,
        )
        from storage.repository import obtener_produccion_diaria as _opd
        from calc.telemetria_costos import calcular_costo_periodo as _ccp
        costo_info = _ccp(cliente_id, energia_kwh, desde, ahora)

        # ── Comparativa periodo anterior ───────────────────────────────────
        # Agregar serie anterior y calcular energía
        bucket_ant = defaultdict(float)
        for hid, rows in mediciones_ant.items():
            for r in rows:
                bucket_ant[r["ts"]] += r["kw"]

        ts_ant = sorted(bucket_ant.keys())
        pot_ant = [bucket_ant[ts] for ts in ts_ant]

        # Muestras esperadas: misma cantidad que el periodo actual
        muestras_esperadas = max(num_muestras, 1)
        disponible_ant = len(ts_ant) >= muestras_esperadas * 0.5

        energia_ant = 0.0
        if len(ts_ant) >= 2:
            from datetime import datetime as _dt3
            for i in range(1, len(ts_ant)):
                try:
                    t0 = _dt3.fromisoformat(ts_ant[i-1].replace("Z", "+00:00"))
                    t1 = _dt3.fromisoformat(ts_ant[i].replace("Z", "+00:00"))
                    dt_h = (t1 - t0).total_seconds() / 3600.0
                    energia_ant += (pot_ant[i-1] + pot_ant[i]) / 2.0 * dt_h
                except Exception:
                    pass

        if disponible_ant and energia_kwh > 0:
            energia_delta_pct = round((energia_kwh - energia_ant) / energia_ant * 100, 1) if energia_ant > 0 else None
        else:
            energia_delta_pct = None

        costo_ant_info = _ccp(cliente_id, energia_ant, desde_ant, hasta_ant) if disponible_ant else None
        costo_ant = costo_ant_info["costo_mxn"] if costo_ant_info else None
        costo_delta_pct = None
        if costo_info.get("costo_mxn") and costo_ant and costo_ant > 0:
            costo_delta_pct = round((costo_info["costo_mxn"] - costo_ant) / costo_ant * 100, 1)

        # ── Sunburst con costo por nodo ────────────────────────────────────
        precio_kwh = costo_info.get("precio_mxn_kwh")

        def _costo_nodo(kwh):
            return round(kwh * precio_kwh, 2) if precio_kwh is not None else None

        def _arbol_sunburst_con_costo(mid):
            m = por_id.get(mid, {})
            hijos_ids_local = [x["id"] for x in todos if x.get("medidor_padre_id") == mid]
            kwh = _energia_nodo(mid)
            return {
                "id": mid,
                "nombre": m.get("nombre", ""),
                "punto_medicion": m.get("punto_medicion", ""),
                "tipo_carga": m.get("tipo_carga"),
                "potencia_nominal_kw": m.get("potencia_nominal_kw"),
                "energia_kwh": kwh,
                "costo_mxn": _costo_nodo(kwh),
                "hijos": [_arbol_sunburst_con_costo(h) for h in hijos_ids_local],
            }

        arbol_sunburst = _arbol_sunburst_con_costo(acometida["id"])

        # ── KPIs de paneles ────────────────────────────────────────────────
        _N_SPARK = {"24h": 24, "7d": 7, "30d": 30}.get(rango, 24)

        # Serie agregada actual (ya disponible como bucket_kw + bucket_fp_peso)
        meds_actuales = [
            {
                "ts": ts,
                "kw": bucket_kw[ts],
                "fp": (bucket_fp_peso[ts] / bucket_kw_peso[ts]
                       if bucket_kw_peso[ts] > 0 else 0.0),
            }
            for ts in ts_sorted
        ]
        meds_anteriores = [
            {"ts": ts, "kw": bucket_ant[ts], "fp": 0.0}
            for ts in sorted(bucket_ant.keys())
        ]

        # Potencia nominal del nodo seleccionado (solo carga_final la tiene)
        pot_nom = nodo.get("potencia_nominal_kw")
        pot_nom = float(pot_nom) if pot_nom else None

        # Calcular KPIs energéticos
        ken_act = _cken(meds_actuales, pot_nom)
        ken_ant = _cken(meds_anteriores, None) if disponible_ant else {}

        def _delta_pct(act_val, ant_val):
            if act_val is None or ant_val is None or ant_val == 0:
                return None
            return round((act_val - ant_val) / abs(ant_val) * 100, 1)

        # Producción diaria
        desde_str = desde.strftime("%Y-%m-%d")
        hasta_str = ahora.strftime("%Y-%m-%d")
        desde_ant_str = desde_ant.strftime("%Y-%m-%d")
        hasta_ant_str = hasta_ant.strftime("%Y-%m-%d")

        prod_act = _opd(cliente_id, desde_str, hasta_str)
        prod_ant = _opd(cliente_id, desde_ant_str, hasta_ant_str)

        m2_planta_act = sum(float(r.get("m2_producidos") or 0) for r in prod_act)
        m2_planta_ant = sum(float(r.get("m2_producidos") or 0) for r in prod_ant)

        # Energía total de la acometida (para atribuir producción proporcionalmente)
        energia_total_planta = _energia_nodo(acometida["id"])

        m2_nodo_act = _apn(m2_planta_act, energia_kwh, energia_total_planta)
        m2_nodo_ant = _apn(m2_planta_ant, energia_ant, energia_total_planta)

        # Baseline = energía del mismo periodo del mes anterior
        baseline_kwh = _cbm(meds_anteriores) if disponible_ant else None

        # KPIs económicos
        precio = costo_info.get("precio_mxn_kwh")
        costo_total_act = costo_info.get("costo_mxn")
        costo_total_ant = costo_ant_info.get("costo_mxn") if costo_ant_info else None
        costo_planta_act = (
            round(energia_total_planta * precio, 2) if precio else None
        )
        kec_act = _cke(energia_kwh, precio, costo_planta_act, baseline_kwh)
        kec_ant = _cke(energia_ant, precio, None, None) if disponible_ant else {}

        # KPIs producción
        kp_act = _ckp(energia_kwh, costo_total_act, m2_nodo_act)
        kp_ant = _ckp(energia_ant, costo_total_ant, m2_nodo_ant) if disponible_ant else {}

        # Sparklines (24 puntos)
        sp_energia_act = _gs(meds_actuales, _N_SPARK)
        sp_energia_ant = _gs(meds_anteriores, _N_SPARK) if disponible_ant else None

        def _kpi_bloque(act_val, ant_val, spark_act, spark_ant, **extra):
            return {
                "actual": act_val,
                "anterior": ant_val if disponible_ant else None,
                "delta_pct": _delta_pct(act_val, ant_val) if disponible_ant else None,
                "sparkline_actual": spark_act,
                "sparkline_anterior": spark_ant,
                **extra,
            }

        kpis_paneles = {
            "energeticos": {
                "energia_kwh": _kpi_bloque(
                    ken_act.get("energia_kwh"), ken_ant.get("energia_kwh"),
                    sp_energia_act, sp_energia_ant,
                    es_favorable_menor=True,
                ),
                "demanda_pico_kw": _kpi_bloque(
                    ken_act.get("demanda_pico_kw"), ken_ant.get("demanda_pico_kw"),
                    None, None,
                    es_favorable_menor=True,
                ),
                "demanda_promedio_kw": _kpi_bloque(
                    ken_act.get("demanda_promedio_kw"), ken_ant.get("demanda_promedio_kw"),
                    None, None,
                    es_favorable_menor=True,
                ),
                "factor_potencia": _kpi_bloque(
                    ken_act.get("factor_potencia_promedio"),
                    ken_ant.get("factor_potencia_promedio"),
                    None, None,
                    es_favorable_menor=False,
                    es_gauge=True,
                    rango_min=0.0,
                    rango_max=1.0,
                ),
            },
            "economicos": {
                "costo_total_mxn": _kpi_bloque(
                    kec_act.get("costo_total_mxn"), kec_ant.get("costo_total_mxn"),
                    None, None,
                    es_favorable_menor=True,
                ),
                "costo_unitario_mxn_kwh": _kpi_bloque(
                    kec_act.get("costo_unitario_mxn_kwh"),
                    kec_ant.get("costo_unitario_mxn_kwh"),
                    None, None,
                    es_favorable_menor=True,
                    fuente_precio=costo_info.get("fuente"),
                ),
                "pct_sobre_factura": _kpi_bloque(
                    kec_act.get("pct_sobre_factura"), kec_ant.get("pct_sobre_factura"),
                    None, None,
                    es_favorable_menor=True,
                    oculto_en_nodo=["acometida_cfe"],
                ),
            },
            "produccion": {
                "solo_en_rango": ["30d"],
                "consumo_especifico_kwh_m2": _kpi_bloque(
                    kp_act.get("consumo_especifico_kwh_m2"),
                    kp_ant.get("consumo_especifico_kwh_m2"),
                    None, None,
                    es_favorable_menor=True,
                ),
                "costo_especifico_mxn_m2": _kpi_bloque(
                    kp_act.get("costo_especifico_mxn_m2"),
                    kp_ant.get("costo_especifico_mxn_m2"),
                    None, None,
                    es_favorable_menor=True,
                ),
                "produccion_m2": _kpi_bloque(
                    kp_act.get("m2_producidos"), kp_ant.get("m2_producidos"),
                    None, None,
                    es_favorable_menor=False,
                ),
            },
            "meta": {
                "periodo_actual_desde": desde_iso,
                "periodo_actual_hasta": hasta_iso,
                "periodo_anterior_desde": desde_ant_iso,
                "periodo_anterior_hasta": hasta_ant_iso,
                "periodo_anterior_etiqueta": etiqueta_ant,
                "n_puntos_sparkline": _N_SPARK,
                "modo_temporal": modo_temporal,
            },
        }

        return jsonify({
            "nodo_seleccionado": _nodo_virtual if _nodo_virtual else {
                "id": nodo["id"],
                "nombre": nodo["nombre"],
                "punto_medicion": nodo.get("punto_medicion"),
                "ruta_breadcrumbs": _breadcrumbs(nodo),
            },
            "serie_temporal": {
                "labels": ts_sorted,
                "potencia_kw": potencia_serie,
            },
            "kpis": {
                "energia_total_kwh": round(energia_kwh, 2),
                "demanda_pico_kw": round(demanda_pico, 2),
                "factor_potencia_promedio": round(fp_prom, 3),
                "num_muestras": num_muestras,
                "costo_mxn": costo_info.get("costo_mxn"),
                "precio_mxn_kwh": costo_info.get("precio_mxn_kwh"),
                "precio_fuente": costo_info.get("fuente"),
                "precio_mes_referencia": costo_info.get("mes_referencia"),
            },
            "comparativa_mes_anterior": {
                "energia_kwh_anterior": round(energia_ant, 2),
                "energia_delta_pct": energia_delta_pct,
                "costo_mxn_anterior": costo_ant,
                "costo_delta_pct": costo_delta_pct,
                "disponible": disponible_ant,
            },
            "arbol_sunburst": arbol_sunburst,
            "kpis_paneles": kpis_paneles,
        })

    @app.route("/clientes/<int:cliente_id>/telemetria/produccion", methods=["POST"])
    def telemetria_produccion_post(cliente_id: int):
        """Captura manual de producción mensual: distribuye m² entre días del mes.

        Body JSON requerido: {"anio": int, "mes": int, "m2_mes": float}
        Ponderación: L-V = 1.0, Sáb = 0.6, Dom = 0.0.
        Retorna: {"ok": True, "registros": N}
        """
        from flask import jsonify, request
        if not app.config.get("FASE2_HABILITADA", False):
            abort(404)
        from web.auth import is_authenticated
        if not is_authenticated():
            return jsonify({"error": "No autenticado"}), 401

        payload = request.get_json(silent=True) or {}
        anio = payload.get("anio")
        mes = payload.get("mes")
        m2_mes = payload.get("m2_mes")

        if not isinstance(anio, int) or anio < 2000 or anio > 2100:
            return jsonify({"error": "anio inválido"}), 400
        if not isinstance(mes, int) or mes < 1 or mes > 12:
            return jsonify({"error": "mes inválido (1-12)"}), 400
        if not isinstance(m2_mes, (int, float)) or m2_mes <= 0:
            return jsonify({"error": "m2_mes debe ser > 0"}), 400
        if m2_mes > 100_000_000:
            return jsonify({"error": "m2_mes excede límite permitido (100 M m²)"}), 400

        from storage.repository import upsert_produccion_mes
        n = upsert_produccion_mes(cliente_id, anio, mes, float(m2_mes))
        return jsonify({"ok": True, "registros": n})

    return app
